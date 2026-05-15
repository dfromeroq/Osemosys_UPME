"""Genera un Excel de seguimiento con las diferencias entre el modelo regionalizado
y el modelo nacional, categorizadas y con diagnóstico + acción recomendada.

Una hoja por set (TECHNOLOGY, FUEL, EMISSION, MODE_OF_OPERATION, YEAR, TIMESLICE)
+ una hoja para PARAM_CSVS y una hoja Resumen.

Uso:
    python build_set_diff_excel.py \\
        --regional /tmp/osemosys_regionalizado \\
        --nacional ./backend/tmp/run_notebook_clean \\
        --output ./diferencias_regional_vs_nacional.xlsx
"""

from __future__ import annotations

import argparse
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─────────────────────────────────────────────────────────────────────────────
# Prefijos regionales y de transporte
# ─────────────────────────────────────────────────────────────────────────────
REGION_PREFIXES = ("NE", "AN", "SO", "OR", "CA", "IN", "SE")
TRANSPORT_PREFIXES = ("TRNELC", "TRNNGS", "TRNOIL", "TRN")
_REGION_RE = re.compile(r"^(" + "|".join(REGION_PREFIXES) + r")_(.+)$")
_TRANSPORT_RE = re.compile(r"^(" + "|".join(TRANSPORT_PREFIXES) + r")(_|$)")
_FUEL_SUFFIX_RE = re.compile(r"00\d$")


def strip_region(code: str) -> str:
    m = _REGION_RE.match(code)
    return m.group(2) if m else code


def is_transport(code: str) -> bool:
    return bool(_TRANSPORT_RE.match(code))


def strip_fuel_suffix(code: str) -> str:
    return _FUEL_SUFFIX_RE.sub("", code)


def load_set(path: Path) -> list[str]:
    if not path.exists():
        return []
    return [l.strip() for l in path.read_text().splitlines()[1:] if l.strip()]


# ─────────────────────────────────────────────────────────────────────────────
# Catálogo de reglas de clasificación
# Para cada elemento "solo en X", clasificamos por patrón con un (categoría,
# diagnóstico, acción recomendada).
# ─────────────────────────────────────────────────────────────────────────────
TECH_RULES: list[tuple[re.Pattern, str, str, str]] = [
    (re.compile(r"^TRN(ELC|NGS|OIL|)_"),
     "Transporte interregional",
     "Tecnología que representa flujo entre regiones; no existe en el modelo nacional por construcción.",
     "Excluir del agregado nacional. Reportar como categoría separada si interesa medir pérdidas de transmisión interregional."),
    (re.compile(r"^BACKSTOP"),
     "Backstop (válvula de escape)",
     "Backstop desagregado por subsector en el regional vs solo 2 genéricos en el nacional.",
     "Decidir política: (a) colapsar todos los BACKSTOP_* en una sola categoría 'BACKSTOP_TOTAL' al agregar; o (b) mantenerlos para diagnóstico de qué subsector necesita la válvula."),
    (re.compile(r"^DEMRESZNI|ZNI"),
     "Zonas No Interconectadas (ZNI)",
     "El nacional tiene tecnologías específicas para ZNI; el regional consolida ZNI dentro de las regiones geográficas o las omite.",
     "Verificar con el modelador si las ZNI se incorporaron dentro de alguna región o si están omitidas. Probable falta de cobertura."),
    (re.compile(r"^GRD"),
     "Red de distribución (Grid)",
     "Tecnologías Grid (GRDTYDELC, GRDZNIELC, GRDNGSDST, …) presentes solo en nacional. En regional el rol equivalente lo cumplen las TRN*.",
     "No-acción si TRN_* cubre el rol. Validar que las pérdidas de distribución estén modeladas en la activity ratio de las TRN_*."),
    (re.compile(r"^MIN(SOL|ONW|OFW|DAM|ROR|GEO|CSP)"),
     "Minería de recurso renovable",
     "El nacional separa el recurso renovable (MINSOL, MINONW, MINROR, ...) de la planta generadora. El regional los integra en una sola tech PWR*.",
     "No-acción si la generación renovable queda contabilizada en PWR* del regional. Documentar la diferencia de formulación."),
    (re.compile(r"^MIN"),
     "Minería / extracción",
     "Otras MIN* (no renovables) presentes solo en uno.",
     "Revisar manualmente. Posible falta de cobertura de recurso primario."),
    (re.compile(r"^EXP|^IMP"),
     "Comercio exterior (importación / exportación)",
     "Diferencias en el manejo de comercio de combustibles.",
     "Verificar si el regional consolida con otra tech o si simplemente no modela ese flujo."),
    (re.compile(r"^UPSREF"),
     "Refinería",
     "Variantes de refinería (Barranca/Cartagena) explícitas en nacional vs versión genérica en regional.",
     "Si interesa desagregación por refinería, revisar si están ocultas dentro de regiones específicas (CA_UPSREF*, AN_UPSREF*)."),
    (re.compile(r"^DEMDERHDG|^BBGDST"),
     "Demanda / Distribución específica",
     "Tecnologías de demanda de hidrógeno o distribución de biogás presentes solo en nacional.",
     "Verificar si están modeladas con otro nombre en regional."),
    (re.compile(r"^TYD"),
     "Transformación / Distribución (TYD)",
     "Familia TYD existe solo en regional; agrupa transformación local de fueles.",
     "Documentar la familia. Decidir si en la agregación nacional se mapea a UPS* o se reporta aparte."),
    (re.compile(r"^DEM"),
     "Demanda final (otros sectores)",
     "DEM* desagregado adicional en regional (mayor granularidad rural/urbana o subsector).",
     "Al agregar, mapear el sufijo extra (_RUR/_URB) sumándolos a la versión base del nacional."),
    (re.compile(r"^PWR"),
     "Generación eléctrica",
     "Diferencia menor en PWR* — pocas techs nuevas o renombradas.",
     "Revisar caso a caso: probable renombre o nueva tecnología agregada en una sola versión del modelo."),
]

FUEL_RULES: list[tuple[re.Pattern, str, str, str]] = [
    (re.compile(r"^(BDL|BET|COA|DSL|ELC|GSL|HDG|JET|LPG|NGS|OIL|SAF)$"),
     "Combustible base sin sufijo",
     "El nacional usa el código base (ELC, NGS, DSL), mientras el regional añade sufijo numérico 001/002/003 para diferenciar calidades o usos.",
     "Mapear en script de agregación: NGS001→NGS, ELC001→ELC, etc. Está implementado el patrón pero falta confirmar que 001 vs 002 no representen calidades distintas."),
    (re.compile(r"^OIL_[123][A-Z]+$|^OIL\d+_[123][A-Z]+$"),
     "Crudo por tipo (LIV/MED/PES)",
     "Ambos modelos diferencian crudo por densidad (1LIV, 2MED, 3PES) pero el regional añade prefijo OIL001_*. Además 'MED' vs 'MID' inconsistente.",
     "Mapear OIL001_2MED → OIL_2MID y verificar si 'MED' (mediano) y 'MID' significan lo mismo. Probable typo histórico."),
    (re.compile(r"^(ROR|DAM|SOL|SOL_RTP|ONW|OFW|OFW_FLO|CSP|GEO|HDGDER|ELCZNI)$"),
     "Recurso primario renovable",
     "Fueles que representan el recurso renovable (radiación solar, viento, agua) explícitos en nacional. El regional los integra implícitamente en PWR*.",
     "No-acción para el agregado de producción. Documentar la diferencia de formulación; los outputs cuantitativos (PJ generados) deberían coincidir aunque el nombre del fuel difiera."),
    (re.compile(r"^RES(CLIM|DHT|OTH|REF)_?(ZNI)?$"),
     "Servicio residencial / terciario",
     "Fueles que representan servicios (climatización, calor, refrigeración) presentes en uno u otro con nomenclatura distinta.",
     "Revisar correspondencia semántica con el modelador. Probables equivalencias 1:1 con renombre."),
    (re.compile(r"^OTH(DSL|ELC|GSL)$"),
     "'Otros' por combustible",
     "Solo en regional. Categoría 'Otros' de DSL/ELC/GSL que captura usos misceláneos.",
     "Al agregar, sumar OTHDSL al DSL nacional (probable). Confirmar con modelador."),
    (re.compile(r"^ELCEV"),
     "Electricidad para vehículos eléctricos",
     "Fuel específico para movilidad eléctrica solo en regional.",
     "Mapear a ELC nacional sumando con flag de seguimiento."),
    (re.compile(r"^COQELC"),
     "Coque eléctrico",
     "Fuel solo en regional. Posiblemente coque de petróleo destinado a generación.",
     "Validar; podría mapear a COA o COQ nacional, o quedar sin contraparte."),
    (re.compile(r"^ELC004$"),
     "ELC numerado adicional",
     "Solo en nacional. Posible variante de electricidad (alta tensión/baja tensión).",
     "Validar mapeo con modelador."),
]

EMISSION_RULES: list[tuple[re.Pattern, str, str, str]] = [
    (re.compile(r"^EMIC0?2$|^EMICO2$"),
     "CO2 — typo entre datasets",
     "Nacional usa 'EMIC02' (cero), regional usa 'EMICO2' (letra O). Significan lo mismo pero no se cruzan textualmente.",
     "ALTA PRIORIDAD: estandarizar a 'EMICO2' en ambos. Definir el mapeo {'EMIC02':'EMICO2'} en el script de agregación."),
    (re.compile(r"^EMI(CH4|N2O)$"),
     "GEI común (CH4, N2O)",
     "Emisión presente y consistente en ambos.",
     "Sin acción. Comparables directamente."),
    (re.compile(r"^EMI(FGA|REN)$"),
     "Agregado solo regional (FGA / REN)",
     "EMIFGA (gases fluorados) y EMIREN (renovables/'avoided') existen solo en regional como categorías agregadas.",
     "No tienen contraparte directa en nacional. Documentar y excluir de la comparación 1-a-1."),
    (re.compile(r"^EMI(BC|CO|COVDM|NOx|NH3|PM10|PM2_5|SOx)$"),
     "Contaminantes locales (solo nacional)",
     "Nacional desagrega contaminantes atmosféricos locales (PM, SOx, NOx, NH3, BC, COVDM). El regional no los modela.",
     "Brecha de cobertura: si el análisis requiere contaminantes locales, no se puede obtener del regional. Reportar como 'no disponible'."),
]

PARAM_RULES: list[tuple[re.Pattern, str, str, str]] = [
    (re.compile(r"^(UDC|UDCMultiplier|UDCConstant|UDCTag)"),
     "User-Defined Constraints (UDC)",
     "Solo en regional. UDC reemplaza la Reserve Margin nativa.",
     "Verificar que las cotas equivalentes se modelen vía RETag* en nacional o vía UDC en regional. Garantizar que no se dupliquen restricciones."),
    (re.compile(r"^(Storage|Capital.*Storage|Operational.*Storage|Residual.*Storage|MinStorage|TechnologyFromStorage|TechnologyToStorage|STORAGE)"),
     "Almacenamiento (Storage)",
     "Familia de parámetros de storage declarada en regional (CSVs presentes aunque vacíos) y ausente en nacional.",
     "Storage está habilitado a nivel de declaración pero los CSV vienen vacíos. Validar si se planea activar storage en el regional."),
    (re.compile(r"^(Conversion|DaySplit|DAILYTIMEBRACKET|DAYTYPE|SEASON|DaysInDayType|STORAGEINTRA)"),
     "Estructura temporal extendida",
     "DAYTYPE, DAILYTIMEBRACKET, SEASON, Conversion* solo en regional.",
     "Soporte para múltiples timeslices y storage intraday/intrayear, aunque no usado actualmente (TS=1)."),
    (re.compile(r"^(TechnologyActivity(By|Increase|Decrease)Mode|EmissionToActivityChangeRatio|InputTo(New|Total)CapacityRatio|DiscountRateIdv)"),
     "Restricciones MUIO y ratios extendidos",
     "MUIO (LU1–LU4) y ratios adicionales presentes solo en regional.",
     "MUIO declarado pero NO cargado al instance_builder (ver CLAUDE.md). Confirmar si se usa o se puede remover."),
    (re.compile(r"^TradeRoute"),
     "TradeRoute",
     "Solo en regional pero viene vacío. Indica intención futura de modelar trade routes vía OSeMOSYS nativo (no usado actualmente).",
     "No-acción inmediata. Si se planea activar trade entre regiones nativo (no UDC), confirmar formato y cargar valores."),
    (re.compile(r"^(RETag|REMin)"),
     "Renewable Energy Targets",
     "Solo en nacional. El regional probablemente codifica targets RE vía UDC en lugar de los tags nativos.",
     "Verificar que los targets RE estén replicados en UDC del regional, sino la comparación de objetivos RE no es directa."),
    (re.compile(r"^ReserveMargin"),
     "Reserve Margin nativo",
     "Solo en nacional. El regional lo reemplaza con UDC_Margin.",
     "Comparar resultados de margen de reserva entre formulaciones. Ya documentado en CLAUDE.md."),
    (re.compile(r"^(AnnualExogenousEmission|ModelPeriodExogenousEmission)"),
     "Emisiones exógenas",
     "Solo en nacional. Permite añadir emisiones no modeladas (e.g. históricas) al total.",
     "Si interesa contabilidad total de emisiones, el regional carece de este mecanismo. Validar."),
    (re.compile(r"^DepreciationMethod"),
     "Método de depreciación",
     "Solo en nacional. Selecciona linear vs sinking-fund.",
     "Verificar que el regional use el default del modelo (1 = linear) y no requiera variantes."),
]


# ─────────────────────────────────────────────────────────────────────────────
# Aplicación de reglas
# ─────────────────────────────────────────────────────────────────────────────
def classify(code: str, rules: list[tuple[re.Pattern, str, str, str]]) -> tuple[str, str, str]:
    for pat, cat, diag, action in rules:
        if pat.search(code):
            return cat, diag, action
    return ("Otros / Sin clasificar",
            "Elemento sin coincidencia con reglas catalogadas. Requiere revisión manual.",
            "Revisar con el modelador y agregar a la regla correspondiente.")


# ─────────────────────────────────────────────────────────────────────────────
# Excel building helpers
# ─────────────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBTITLE_FONT = Font(bold=True, size=14, color="1F4E78")
INFO_FILL = PatternFill("solid", fgColor="DDEBF7")
ALERT_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, max_width=80):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        widths = [len(str(c.value)) for c in col if c.value]
        if not widths:
            continue
        ws.column_dimensions[letter].width = min(max(widths) + 2, max_width)


def _write_header(ws, title: str, subtitle: str, summary: list[tuple[str, str]]):
    ws["A1"] = title
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = subtitle
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:F2")
    row = 4
    ws[f"A{row}"] = "Métrica"
    ws[f"B{row}"] = "Valor"
    ws[f"A{row}"].font = HEADER_FONT
    ws[f"B{row}"].font = HEADER_FONT
    ws[f"A{row}"].fill = HEADER_FILL
    ws[f"B{row}"].fill = HEADER_FILL
    row += 1
    for k, v in summary:
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        ws[f"A{row}"].fill = INFO_FILL
        row += 1
    return row + 2


def _write_table(ws, start_row: int, columns: list[str], rows: list[list[str]], table_name: str):
    if not rows:
        ws[f"A{start_row}"] = "— Sin elementos en esta categoría —"
        ws[f"A{start_row}"].font = Font(italic=True, color="808080")
        return start_row + 2
    # encabezado
    for j, c in enumerate(columns):
        cell = ws.cell(row=start_row, column=j + 1, value=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = BORDER
    # filas
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            cell = ws.cell(row=start_row + 1 + i, column=j + 1, value=v)
            cell.alignment = WRAP
            cell.border = BORDER
            if j >= 2:  # categoría / diagnóstico / acción
                cell.font = Font(size=10)
    # crear Table object (filtros + estilos)
    end_row = start_row + len(rows)
    end_col_letter = get_column_letter(len(columns))
    ref = f"A{start_row}:{end_col_letter}{end_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)
    return end_row + 2


# ─────────────────────────────────────────────────────────────────────────────
# Construcción del libro
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(regional: Path, nacional: Path, output: Path) -> None:
    wb = Workbook()
    # Quitamos hoja default
    wb.remove(wb.active)

    # Cargas
    reg_tech_raw = load_set(regional / "TECHNOLOGY.csv")
    reg_fuel_raw = load_set(regional / "FUEL.csv")
    reg_emi = load_set(regional / "EMISSION.csv")
    reg_mod = load_set(regional / "MODE_OF_OPERATION.csv")
    reg_year = load_set(regional / "YEAR.csv")
    reg_ts = load_set(regional / "TIMESLICE.csv")

    nat_tech = load_set(nacional / "TECHNOLOGY.csv")
    nat_fuel = load_set(nacional / "FUEL.csv")
    nat_emi = load_set(nacional / "EMISSION.csv")
    nat_mod = load_set(nacional / "MODE_OF_OPERATION.csv")
    nat_year = load_set(nacional / "YEAR.csv")
    nat_ts = load_set(nacional / "TIMESLICE.csv")

    # Strip helpers
    reg_tech_trn = sorted({c for c in reg_tech_raw if is_transport(c)})
    reg_tech_stripped = sorted({strip_region(c) for c in reg_tech_raw if not is_transport(c)})
    reg_fuel_stripped = sorted({strip_region(c) for c in reg_fuel_raw})
    nat_tech_set = sorted(set(nat_tech))
    nat_fuel_set = sorted(set(nat_fuel))

    summary_rows: list[tuple[str, str, int, int, int]] = []  # (sheet, set, only_reg, only_nat, common)

    # ─────────── Hoja: Resumen (la generamos al final) ───────────
    summary_ws = wb.create_sheet("Resumen", 0)

    # ─────────── Hoja: TECHNOLOGY ───────────
    ws = wb.create_sheet("TECHNOLOGY")
    only_reg = sorted(set(reg_tech_stripped) - set(nat_tech_set))
    only_nat = sorted(set(nat_tech_set) - set(reg_tech_stripped))
    common = sorted(set(reg_tech_stripped) & set(nat_tech_set))
    summary_rows.append(("TECHNOLOGY", "TECHNOLOGY", len(only_reg), len(only_nat), len(common)))

    next_row = _write_header(
        ws,
        title="TECHNOLOGY — diferencias",
        subtitle=(
            "Diferencias entre el modelo regionalizado y el nacional. El regional se compara "
            "tras eliminar el prefijo regional (NE_/AN_/SO_/OR_/CA_/IN_/SE_) y separar las "
            "tecnologías de transporte interregional (TRN*). Las TRN* no aparecen en la comparación "
            "porque no tienen contraparte por construcción."
        ),
        summary=[
            ("Total nacional", len(nat_tech_set)),
            ("Total regional (raw)", len(reg_tech_raw)),
            ("Total regional (post-strip único)", len(reg_tech_stripped)),
            ("Transporte interregional (TRN*)", len(reg_tech_trn)),
            ("Solo en REGIONAL (post-strip)", len(only_reg)),
            ("Solo en NACIONAL", len(only_nat)),
            ("En común", len(common)),
        ],
    )

    ws.cell(row=next_row, column=1, value="Elementos solo en REGIONAL (post-strip)").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Regional"] + list(classify(c, TECH_RULES)) + [""] for c in only_reg]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_tech_reg",
    )

    ws.cell(row=next_row, column=1, value="Elementos solo en NACIONAL").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Nacional"] + list(classify(c, TECH_RULES)) + [""] for c in only_nat]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_tech_nat",
    )

    ws.cell(row=next_row, column=1, value="Transporte interregional (referencia)").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Regional (transporte)",
             "Transporte interregional",
             "Tecnología de flujo entre regiones; sin contraparte en nacional por diseño.",
             "Excluido del agregado por defecto. Disponible con --include-transport.",
             ""] for c in reg_tech_trn]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_tech_trn",
    )
    _autosize(ws)
    # Ajuste manual de columnas anchas
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["B"].width = 18

    # ─────────── Hoja: FUEL ───────────
    ws = wb.create_sheet("FUEL")
    # Comparación cruda (post-strip prefijo regional)
    only_reg_raw = sorted(set(reg_fuel_stripped) - set(nat_fuel_set))
    only_nat_raw = sorted(set(nat_fuel_set) - set(reg_fuel_stripped))
    # Comparación adicional: removiendo sufijo numérico
    reg_fuel_base = sorted({strip_fuel_suffix(f) for f in reg_fuel_stripped})
    nat_fuel_base = sorted({strip_fuel_suffix(f) for f in nat_fuel_set})
    only_reg_base = sorted(set(reg_fuel_base) - set(nat_fuel_base))
    only_nat_base = sorted(set(nat_fuel_base) - set(reg_fuel_base))
    common_base = sorted(set(reg_fuel_base) & set(nat_fuel_base))
    summary_rows.append(("FUEL", "FUEL", len(only_reg_raw), len(only_nat_raw), len(set(reg_fuel_stripped) & set(nat_fuel_set))))

    next_row = _write_header(
        ws,
        title="FUEL — diferencias",
        subtitle=(
            "Comparación tras quitar prefijo regional. Se incluye una segunda comparación que "
            "además remueve el sufijo numérico 00X (NGS001 → NGS), porque el regional añade "
            "sufijos a casi todos los combustibles fósiles. La comparación 'por base' muestra "
            "cuántos códigos son realmente comparables semánticamente."
        ),
        summary=[
            ("Total nacional", len(nat_fuel_set)),
            ("Total regional (post-strip único)", len(reg_fuel_stripped)),
            ("Solo en regional (post-strip)", len(only_reg_raw)),
            ("Solo en nacional (post-strip)", len(only_nat_raw)),
            ("Solo en regional (base, sin sufijo 00X)", len(only_reg_base)),
            ("Solo en nacional (base, sin sufijo 00X)", len(only_nat_base)),
            ("Comparables (base, sin sufijo)", len(common_base)),
        ],
    )

    ws.cell(row=next_row, column=1, value="Solo en REGIONAL (post-strip prefijo regional)").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Regional"] + list(classify(c, FUEL_RULES)) + [""] for c in only_reg_raw]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_fuel_reg",
    )

    ws.cell(row=next_row, column=1, value="Solo en NACIONAL").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Nacional"] + list(classify(c, FUEL_RULES)) + [""] for c in only_nat_raw]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_fuel_nat",
    )

    ws.cell(row=next_row, column=1, value="Solo en NACIONAL — comparación por código BASE (sin sufijo 00X)").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Nacional (base)"] + list(classify(c, FUEL_RULES)) + [""] for c in only_nat_base]
    next_row = _write_table(
        ws, next_row,
        ["Código base", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_fuel_nat_base",
    )

    ws.cell(row=next_row, column=1, value="Solo en REGIONAL — comparación por código BASE (sin sufijo 00X)").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Regional (base)"] + list(classify(c, FUEL_RULES)) + [""] for c in only_reg_base]
    next_row = _write_table(
        ws, next_row,
        ["Código base", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_fuel_reg_base",
    )

    _autosize(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["B"].width = 18

    # ─────────── Hoja: EMISSION ───────────
    ws = wb.create_sheet("EMISSION")
    only_reg = sorted(set(reg_emi) - set(nat_emi))
    only_nat = sorted(set(nat_emi) - set(reg_emi))
    common = sorted(set(reg_emi) & set(nat_emi))
    summary_rows.append(("EMISSION", "EMISSION", len(only_reg), len(only_nat), len(common)))

    next_row = _write_header(
        ws,
        title="EMISSION — diferencias",
        subtitle=(
            "Los sets de emisiones son casi disjuntos. Hay un typo en CO₂ (EMIC02 con cero vs "
            "EMICO2 con letra O) que impide cruzar el dataset si no se mapea explícitamente."
        ),
        summary=[
            ("Total nacional", len(nat_emi)),
            ("Total regional", len(reg_emi)),
            ("Solo regional", len(only_reg)),
            ("Solo nacional", len(only_nat)),
            ("En común (texto exacto)", len(common)),
        ],
    )

    ws.cell(row=next_row, column=1, value="Solo en REGIONAL").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Regional"] + list(classify(c, EMISSION_RULES)) + [""] for c in only_reg]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_emi_reg",
    )

    ws.cell(row=next_row, column=1, value="Solo en NACIONAL").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Nacional"] + list(classify(c, EMISSION_RULES)) + [""] for c in only_nat]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_emi_nat",
    )

    ws.cell(row=next_row, column=1, value="En común (texto exacto)").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Común", "GEI común", "Presente y consistente en ambos modelos.", "Sin acción.", ""] for c in common]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_emi_common",
    )

    _autosize(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["B"].width = 18

    # ─────────── Hoja: MODE_OF_OPERATION ───────────
    ws = wb.create_sheet("MODE_OF_OPERATION")
    only_reg = sorted(set(reg_mod) - set(nat_mod))
    only_nat = sorted(set(nat_mod) - set(reg_mod))
    common = sorted(set(reg_mod) & set(nat_mod))
    summary_rows.append(("MODE_OF_OPERATION", "MODE_OF_OPERATION", len(only_reg), len(only_nat), len(common)))

    next_row = _write_header(
        ws,
        title="MODE_OF_OPERATION — diferencias",
        subtitle=(
            "El nacional opera en modo único (1). El regional define dos modos (1, 2). "
            "Las tecnologías usan los modos para distinguir flujos alternativos (por ejemplo, "
            "cogeneración elec/calor)."
        ),
        summary=[
            ("Modos en nacional", ", ".join(nat_mod) or "—"),
            ("Modos en regional", ", ".join(reg_mod) or "—"),
            ("Solo regional", len(only_reg)),
            ("Solo nacional", len(only_nat)),
            ("Común", len(common)),
        ],
    )

    rows = [
        [c, "Regional", "Modo adicional",
         "Modo 2 solo definido en regional; usado para separar productos secundarios (calor, cogeneración) o variantes de operación.",
         "Al agregar a nacional: sumar `RateOfActivity` sobre el modo. Verificar que no se pierda información útil (e.g. dispatch por modo).",
         ""]
        for c in only_reg
    ]
    next_row = _write_table(
        ws, next_row,
        ["Código", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_mod_reg",
    )

    _autosize(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["B"].width = 18

    # ─────────── Hoja: YEAR ───────────
    ws = wb.create_sheet("YEAR")
    only_reg = sorted(set(reg_year) - set(nat_year))
    only_nat = sorted(set(nat_year) - set(reg_year))
    common = sorted(set(reg_year) & set(nat_year))
    summary_rows.append(("YEAR", "YEAR", len(only_reg), len(only_nat), len(common)))

    next_row = _write_header(
        ws,
        title="YEAR — diferencias",
        subtitle="Verificación de horizonte temporal.",
        summary=[
            ("Horizonte nacional",
             f"{nat_year[0] if nat_year else '—'}–{nat_year[-1] if nat_year else '—'} ({len(nat_year)})"),
            ("Horizonte regional",
             f"{reg_year[0] if reg_year else '—'}–{reg_year[-1] if reg_year else '—'} ({len(reg_year)})"),
            ("Años solo regional", len(only_reg)),
            ("Años solo nacional", len(only_nat)),
            ("Años en común", len(common)),
        ],
    )
    if only_reg or only_nat:
        rows = [
            [y, "Regional", "Año extra", "Año presente solo en regional.", "Recortar al horizonte común para comparación.", ""]
            for y in only_reg
        ] + [
            [y, "Nacional", "Año extra", "Año presente solo en nacional.", "Recortar al horizonte común para comparación.", ""]
            for y in only_nat
        ]
        next_row = _write_table(
            ws, next_row,
            ["Año", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
            rows, "tbl_year",
        )
    else:
        ws.cell(row=next_row, column=1,
                value="Horizonte idéntico en ambos modelos. Sin acción.").font = Font(italic=True, color="2E7D32")
    _autosize(ws)

    # ─────────── Hoja: TIMESLICE ───────────
    ws = wb.create_sheet("TIMESLICE")
    only_reg = sorted(set(reg_ts) - set(nat_ts))
    only_nat = sorted(set(nat_ts) - set(reg_ts))
    common = sorted(set(reg_ts) & set(nat_ts))
    summary_rows.append(("TIMESLICE", "TIMESLICE", len(only_reg), len(only_nat), len(common)))

    next_row = _write_header(
        ws,
        title="TIMESLICE — diferencias",
        subtitle="Resolución temporal intra-anual.",
        summary=[
            ("Timeslices nacional", ", ".join(nat_ts) or "—"),
            ("Timeslices regional", ", ".join(reg_ts) or "—"),
            ("Solo regional", len(only_reg)),
            ("Solo nacional", len(only_nat)),
            ("Común", len(common)),
        ],
    )
    if not only_reg and not only_nat:
        ws.cell(row=next_row, column=1,
                value="Ambos modelos usan TS_0 (timeslice único). Sin acción.").font = Font(italic=True, color="2E7D32")
    _autosize(ws)

    # ─────────── Hoja: PARAM_CSVS ───────────
    ws = wb.create_sheet("PARAM_CSVS")
    nat_params = {p.name for p in nacional.glob("*.csv")}
    reg_params = {p.name for p in regional.glob("*.csv")}
    only_reg = sorted(reg_params - nat_params)
    only_nat = sorted(nat_params - reg_params)
    common = sorted(reg_params & nat_params)
    summary_rows.append(("PARAM_CSVS", "Param CSVs presentes", len(only_reg), len(only_nat), len(common)))

    next_row = _write_header(
        ws,
        title="PARAM CSVs — presencia",
        subtitle=(
            "Diferencias en qué archivos de parámetros vienen con cada dataset. Un CSV "
            "ausente significa que ese parámetro queda en su valor default del modelo."
        ),
        summary=[
            ("CSV solo regional", len(only_reg)),
            ("CSV solo nacional", len(only_nat)),
            ("CSV comunes", len(common)),
        ],
    )

    ws.cell(row=next_row, column=1, value="Solo en REGIONAL").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Regional"] + list(classify(c.replace(".csv", ""), PARAM_RULES)) + [""] for c in only_reg]
    next_row = _write_table(
        ws, next_row,
        ["CSV", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_param_reg",
    )

    ws.cell(row=next_row, column=1, value="Solo en NACIONAL").font = SUBTITLE_FONT
    next_row += 1
    rows = [[c, "Nacional"] + list(classify(c.replace(".csv", ""), PARAM_RULES)) + [""] for c in only_nat]
    next_row = _write_table(
        ws, next_row,
        ["CSV", "Origen", "Categoría", "Diagnóstico", "Acción recomendada", "Estado / Notas"],
        rows, "tbl_param_nat",
    )

    _autosize(ws)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["B"].width = 18

    # ─────────── Hoja: Resumen ───────────
    ws = summary_ws
    ws["A1"] = "Resumen — diferencias entre modelo regionalizado y nacional"
    ws["A1"].font = SUBTITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"Regional:  {regional}\n"
        f"Nacional:  {nacional}\n"
        f"Generado por backend/scripts/build_set_diff_excel.py"
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 45

    headers = ["Hoja", "Set", "Solo regional", "Solo nacional", "En común", "Severidad"]
    for j, h in enumerate(headers):
        c = ws.cell(row=4, column=j + 1, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = WRAP; c.border = BORDER

    def severity(only_r, only_n, common):
        total = only_r + only_n + common
        if total == 0:
            return "—"
        diff_ratio = (only_r + only_n) / total
        if diff_ratio < 0.05:
            return "🟢 Bajo"
        if diff_ratio < 0.25:
            return "🟡 Medio"
        return "🔴 Alto"

    for i, (sheet, setname, or_, on_, com) in enumerate(summary_rows):
        ws.cell(row=5 + i, column=1, value=sheet).border = BORDER
        ws.cell(row=5 + i, column=2, value=setname).border = BORDER
        ws.cell(row=5 + i, column=3, value=or_).border = BORDER
        ws.cell(row=5 + i, column=4, value=on_).border = BORDER
        ws.cell(row=5 + i, column=5, value=com).border = BORDER
        sev_cell = ws.cell(row=5 + i, column=6, value=severity(or_, on_, com))
        sev_cell.border = BORDER
        sev_cell.alignment = Alignment(horizontal="center")

    # Notas finales
    notes_row = 5 + len(summary_rows) + 2
    ws.cell(row=notes_row, column=1, value="Notas").font = SUBTITLE_FONT
    notes = [
        "• Severidad calculada como (solo_reg + solo_nac) / (total). 🔴 ≥25 %, 🟡 ≥5 %, 🟢 <5 %.",
        "• La hoja TECHNOLOGY excluye las 34 tecnologías de transporte interregional (TRN*) del diff y las muestra como tabla aparte.",
        "• La hoja FUEL incluye dos comparaciones: (a) post-strip prefijo regional, (b) tras también quitar sufijo numérico 00X.",
        "• EMISSION es el caso de mayor severidad por typo CO2 + diferencia de modelado (GEI agregados vs contaminantes locales detallados).",
        "• Las acciones marcadas con 'ALTA PRIORIDAD' o que afecten reconciliación cuantitativa deben resolverse antes de cualquier comparación numérica.",
        "• La columna 'Estado / Notas' está vacía a propósito: úsala para registrar decisiones, responsables o links a tickets de seguimiento.",
    ]
    for k, n in enumerate(notes):
        c = ws.cell(row=notes_row + 1 + k, column=1, value=n)
        c.alignment = WRAP
        ws.merge_cells(start_row=notes_row + 1 + k, start_column=1, end_row=notes_row + 1 + k, end_column=6)

    for col_letter, width in zip("ABCDEF", [22, 22, 14, 14, 12, 14]):
        ws.column_dimensions[col_letter].width = width

    # Guardado
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"✔ Excel generado: {output}")


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--regional", required=True, type=Path)
    p.add_argument("--nacional", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    args = p.parse_args()
    build_excel(args.regional, args.nacional, args.output)


if __name__ == "__main__":
    main()
