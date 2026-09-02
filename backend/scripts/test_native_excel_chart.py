#!/usr/bin/env python3
"""Genera XLSX de prueba con gráficas nativas Excel (stacked column, line, area)."""
from __future__ import annotations

import io
import sys
from pathlib import Path

# Permitir importar app.* desde backend/
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook

from app.visualization.chart_service import (
    _add_native_excel_chart,
    _write_data_table,
)


def _minimal_png_bytes() -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(4, 2))
    ax.bar(["2024", "2030"], [100, 200])
    ax.set_title("PNG backup")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=72)
    plt.close(fig)
    return buf.getvalue()


def _build_sample_table() -> tuple[list[str], list[list]]:
    headers = ["Categoria", "Biogas", "Carbon", "Gas Natural", "Hidro"]
    rows = [
        ["2022", 10, 50, 120, 80],
        ["2024", 12, 45, 130, 85],
        ["2030", 15, 30, 150, 90],
        ["2050", 20, 10, 180, 95],
    ]
    return headers, rows


def _write_sheet_with_native_chart(
    wb: Workbook,
    sheet_name: str,
    chart_type: str,
    *,
    with_none_values: bool = False,
) -> None:
    headers, rows = _build_sample_table()
    if with_none_values:
        rows[1][2] = None  # simular celda vacía corregida a 0 en producción

    ws = wb.create_sheet(sheet_name)
    header_row, data_start_row, data_end_row, _ = _write_data_table(
        ws, 1, headers, rows, y_axis_label="PJ"
    )

    series_colors = ["#8B4513", "#2F4F4F", "#4169E1", "#228B22"]
    chart_anchor_row = data_end_row + 2
    _add_native_excel_chart(
        ws,
        categories_col=1,
        data_start_col=2,
        data_end_col=len(headers),
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        series_colors=series_colors,
        chart_type=chart_type,
        title=f"Prueba {chart_type} — {sheet_name}",
        y_axis_label="PJ",
        anchor_cell=f"A{chart_anchor_row}",
    )


def main() -> None:
    out_path = Path("/tmp/test_native_chart.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet_with_native_chart(wb, "StackedColumn", "column")
    _write_sheet_with_native_chart(wb, "Line", "line")
    _write_sheet_with_native_chart(wb, "StackedArea", "area")

    wb.save(out_path)
    print(f"OK: {out_path} ({out_path.stat().st_size} bytes)")
    print("Hojas:", wb.sheetnames)


if __name__ == "__main__":
    main()
