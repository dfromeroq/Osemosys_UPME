"""Build a pivot-ready employment workbook from exported simulation results.

The script expects exported result workbooks with a ``Resultados`` sheet in the
wide format produced by the simulation output exporter. It concatenates the
scenario results, applies direct employment factors, and writes a workbook with
raw long-form results plus employment outputs.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_DIR.parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.services.employment_factors_service import (  # noqa: E402
    CONSTRUCTION_FACTOR_TYPE,
    DIRECT_JOB_TYPE,
    OM_FACTOR_TYPE,
    OPTIMIZATION_START_YEAR,
    load_model_employment_factors,
    load_technology_mapping,
    pj_per_year_to_mw,
)

RESULT_SHEET = "Resultados"
INPUT_VARIABLES = {
    "NewCapacity": CONSTRUCTION_FACTOR_TYPE,
    "AccumulatedNewCapacity": OM_FACTOR_TYPE,
}
EMPLOYMENT_VARIABLE_BY_FACTOR_TYPE = {
    CONSTRUCTION_FACTOR_TYPE: "EmploymentConstructionManufacturingDirect",
    OM_FACTOR_TYPE: "EmploymentOMDirect",
}
FORMULA_COLUMNS = {
    "value_pj_per_year",
    "capacity_mw",
    "component_multiplier",
    "component_capacity_mw",
    "factor_value",
    "factor_unit",
    "factor_source",
    "default_source",
    "employment_fte",
    "capacity_pj_per_year",
}


def _infer_scenario(path: Path) -> str:
    match = re.search(r"_filtered_([^._]+)", path.stem)
    if match:
        return match.group(1)
    return path.stem


def _year_columns(columns: list[Any]) -> list[Any]:
    return [col for col in columns if str(col).isdigit()]


def load_results_long(paths: list[Path]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for path in paths:
        df = pd.read_excel(path, sheet_name=RESULT_SHEET)
        years = _year_columns(list(df.columns))
        id_vars = [col for col in df.columns if col not in years]
        long = df.melt(
            id_vars=id_vars,
            value_vars=years,
            var_name="year",
            value_name="value_pj_per_year",
        )
        long.insert(0, "scenario", _infer_scenario(path))
        long.insert(1, "source_file", path.name)
        long["year"] = pd.to_numeric(long["year"], errors="coerce").astype("Int64")
        long["value_pj_per_year"] = pd.to_numeric(long["value_pj_per_year"], errors="coerce")
        frames.append(long)

    if not frames:
        return pd.DataFrame()
    raw_long = pd.concat(frames, ignore_index=True)
    raw_long.insert(
        0,
        "raw_lookup_key",
        raw_long.apply(
            lambda row: _join_key(
                row["scenario"],
                row["variable_name"],
                row["region"],
                row["technology"],
                row["year"],
            ),
            axis=1,
        ),
    )
    return raw_long


def _eligible_factors(factors: pd.DataFrame) -> pd.DataFrame:
    filtered = factors[
        (factors["Job_Type"] == DIRECT_JOB_TYPE)
        & (factors["Factor_Type"].isin([CONSTRUCTION_FACTOR_TYPE, OM_FACTOR_TYPE]))
    ].copy()
    filtered["Year"] = pd.to_numeric(filtered["Year"], errors="coerce").astype("Int64")
    filtered["Value_Numeric"] = pd.to_numeric(filtered["Value_Numeric"], errors="coerce")
    return filtered.dropna(subset=["Year", "Value_Numeric"])


def _flatten_mapping(mapping) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for model_technology, components in mapping.items():
        for component in components:
            rows.append(
                {
                    "mapping_lookup_key": _join_key(
                        model_technology,
                        component.employment_technology,
                    ),
                    "technology": model_technology,
                    "employment_technology": component.employment_technology,
                    "component_multiplier": component.multiplier,
                    "mapping_note": component.note,
                }
            )
    return pd.DataFrame(rows)


def _join_key(*parts: Any) -> str:
    return "|".join("" if pd.isna(part) else str(part) for part in parts)


def build_employment_results(raw_long: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    mapping = load_technology_mapping()
    mapping_df = _flatten_mapping(mapping)
    factors = _eligible_factors(load_model_employment_factors())

    capacity = raw_long[
        raw_long["variable_name"].isin(INPUT_VARIABLES)
        & (raw_long["year"] >= OPTIMIZATION_START_YEAR)
        & raw_long["value_pj_per_year"].notna()
    ].copy()
    capacity["factor_type"] = capacity["variable_name"].map(INPUT_VARIABLES)
    capacity["employment_variable"] = capacity["factor_type"].map(
        EMPLOYMENT_VARIABLE_BY_FACTOR_TYPE
    )
    capacity["capacity_mw"] = capacity["value_pj_per_year"].map(pj_per_year_to_mw)

    expanded = capacity.merge(mapping_df, on="technology", how="left", indicator=True)
    unmatched = expanded[expanded["_merge"] == "left_only"].copy()
    expanded = expanded[expanded["_merge"] == "both"].drop(columns=["_merge"])
    expanded["component_capacity_mw"] = (
        expanded["capacity_mw"] * expanded["component_multiplier"]
    )

    factor_cols = [
        "Technology",
        "Year",
        "Factor_Type",
        "Job_Type",
        "Unit",
        "Source",
        "Value_Numeric",
        "Default_Source",
    ]
    factors_for_merge = factors[factor_cols].rename(
        columns={
            "Technology": "employment_technology",
            "Year": "year",
            "Factor_Type": "factor_type",
            "Job_Type": "job_type",
            "Unit": "factor_unit",
            "Source": "factor_source",
            "Value_Numeric": "factor_value",
            "Default_Source": "default_source",
        }
    )
    factors_for_merge.insert(
        0,
        "factor_lookup_key",
        factors_for_merge.apply(
            lambda row: _join_key(
                row["employment_technology"],
                row["year"],
                row["factor_type"],
                row["job_type"],
            ),
            axis=1,
        ),
    )
    employment = expanded.merge(
        factors_for_merge,
        on=["employment_technology", "year", "factor_type"],
        how="left",
    )
    missing_factor = employment[employment["factor_value"].isna()].copy()
    employment = employment[employment["factor_value"].notna()].copy()
    employment["employment_fte"] = (
        employment["component_capacity_mw"] * employment["factor_value"]
    )
    employment["raw_lookup_key"] = employment.apply(
        lambda row: _join_key(
            row["scenario"],
            row["variable_name"],
            row["region"],
            row["technology"],
            row["year"],
        ),
        axis=1,
    )
    employment["factor_lookup_key"] = employment.apply(
        lambda row: _join_key(
            row["employment_technology"],
            row["year"],
            row["factor_type"],
            row["job_type"],
        ),
        axis=1,
    )
    employment["mapping_lookup_key"] = employment.apply(
        lambda row: _join_key(row["technology"], row["employment_technology"]),
        axis=1,
    )

    ordered_cols = [
        "raw_lookup_key",
        "factor_lookup_key",
        "mapping_lookup_key",
        "scenario",
        "source_file",
        "region",
        "technology",
        "year",
        "variable_name",
        "employment_variable",
        "factor_type",
        "job_type",
        "value_pj_per_year",
        "capacity_mw",
        "employment_technology",
        "component_multiplier",
        "component_capacity_mw",
        "factor_value",
        "factor_unit",
        "factor_source",
        "default_source",
        "employment_fte",
        "mapping_note",
    ]
    employment = employment[[col for col in ordered_cols if col in employment.columns]]

    aggregated = (
        employment.groupby(
            [
                "scenario",
                "region",
                "technology",
                "year",
                "variable_name",
                "employment_variable",
                "factor_type",
                "job_type",
            ],
            dropna=False,
            as_index=False,
        )
        .agg(
            capacity_pj_per_year=("value_pj_per_year", "first"),
            capacity_mw=("capacity_mw", "first"),
            employment_fte=("employment_fte", "sum"),
        )
        .sort_values(["scenario", "employment_variable", "technology", "year"])
    )
    aggregated.insert(
        0,
        "raw_lookup_key",
        aggregated.apply(
            lambda row: _join_key(
                row["scenario"],
                row["variable_name"],
                row["region"],
                row["technology"],
                row["year"],
            ),
            axis=1,
        ),
    )

    diagnostics = pd.concat(
        [
            unmatched.assign(issue="missing_technology_mapping"),
            missing_factor.assign(issue="missing_employment_factor"),
        ],
        ignore_index=True,
        sort=False,
    )
    return employment, aggregated, diagnostics


def write_workbook(
    *,
    raw_long: pd.DataFrame,
    employment: pd.DataFrame,
    aggregated: pd.DataFrame,
    diagnostics: pd.DataFrame,
    output_path: Path,
) -> None:
    mapping_df = _flatten_mapping(load_technology_mapping())
    factors = _eligible_factors(load_model_employment_factors())
    matched_factor_keys = employment[
        ["employment_technology", "year", "factor_type"]
    ].drop_duplicates()
    factor_inputs = factors.merge(
        matched_factor_keys,
        left_on=["Technology", "Year", "Factor_Type"],
        right_on=["employment_technology", "year", "factor_type"],
        how="inner",
    ).drop(columns=["employment_technology", "year", "factor_type"])
    factor_inputs.insert(
        0,
        "factor_lookup_key",
        factor_inputs.apply(
            lambda row: _join_key(
                row["Technology"],
                row["Year"],
                row["Factor_Type"],
                row["Job_Type"],
            ),
            axis=1,
        ),
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        raw_long.to_excel(writer, sheet_name="RawResultsLong", index=False)
        employment.to_excel(writer, sheet_name="EmploymentResults", index=False)
        aggregated.to_excel(writer, sheet_name="EmploymentAggregated", index=False)
        mapping_df.to_excel(writer, sheet_name="TechnologyMapping", index=False)
        factor_inputs.to_excel(writer, sheet_name="FactorInputs", index=False)
        diagnostics.to_excel(writer, sheet_name="Diagnostics", index=False)
    _add_formulas_and_tables(output_path)


def _column_map(ws) -> dict[str, int]:
    return {str(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def _col_letter(ws, column_name: str) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(_column_map(ws)[column_name])


def _formula_key(*refs: str) -> str:
    return "&\"|\"&".join(refs)


def _quote(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _set_employment_formulas(wb) -> None:
    ws = wb["EmploymentResults"]
    cols = _column_map(ws)
    raw = wb["RawResultsLong"]
    factor = wb["FactorInputs"]
    mapping = wb["TechnologyMapping"]
    raw_value_col = _col_letter(raw, "value_pj_per_year")
    raw_scenario_col = _col_letter(raw, "scenario")
    raw_variable_col = _col_letter(raw, "variable_name")
    raw_region_col = _col_letter(raw, "region")
    raw_technology_col = _col_letter(raw, "technology")
    raw_year_col = _col_letter(raw, "year")
    factor_value_col = _col_letter(factor, "Value_Numeric")
    factor_technology_col = _col_letter(factor, "Technology")
    factor_year_col = _col_letter(factor, "Year")
    factor_type_col = _col_letter(factor, "Factor_Type")
    factor_job_type_col = _col_letter(factor, "Job_Type")
    factor_key_col = _col_letter(factor, "factor_lookup_key")
    factor_unit_col = _col_letter(factor, "Unit")
    factor_source_col = _col_letter(factor, "Source")
    factor_default_source_col = _col_letter(factor, "Default_Source")
    mapping_multiplier_col = _col_letter(mapping, "component_multiplier")
    mapping_technology_col = _col_letter(mapping, "technology")
    mapping_employment_technology_col = _col_letter(mapping, "employment_technology")

    for row in range(2, ws.max_row + 1):
        scenario = f"{_col_letter(ws, 'scenario')}{row}"
        variable = f"{_col_letter(ws, 'variable_name')}{row}"
        region = f"{_col_letter(ws, 'region')}{row}"
        technology = f"{_col_letter(ws, 'technology')}{row}"
        year = f"{_col_letter(ws, 'year')}{row}"
        employment_technology = f"{_col_letter(ws, 'employment_technology')}{row}"
        factor_type = f"{_col_letter(ws, 'factor_type')}{row}"
        job_type = f"{_col_letter(ws, 'job_type')}{row}"
        factor_key = f"{_col_letter(ws, 'factor_lookup_key')}{row}"
        capacity_mw = f"{_col_letter(ws, 'capacity_mw')}{row}"
        component_multiplier = f"{_col_letter(ws, 'component_multiplier')}{row}"
        component_capacity_mw = f"{_col_letter(ws, 'component_capacity_mw')}{row}"
        factor_value = f"{_col_letter(ws, 'factor_value')}{row}"

        ws.cell(row, cols["raw_lookup_key"]).value = (
            f"={_formula_key(scenario, variable, region, technology, year)}"
        )
        ws.cell(row, cols["factor_lookup_key"]).value = (
            f"={_formula_key(employment_technology, year, factor_type, job_type)}"
        )
        ws.cell(row, cols["mapping_lookup_key"]).value = (
            f"={_formula_key(technology, employment_technology)}"
        )
        ws.cell(row, cols["value_pj_per_year"]).value = (
            f'=SUMIFS(RawResultsLong!${raw_value_col}:${raw_value_col},'
            f'RawResultsLong!${raw_scenario_col}:${raw_scenario_col},{scenario},'
            f'RawResultsLong!${raw_variable_col}:${raw_variable_col},{variable},'
            f'RawResultsLong!${raw_region_col}:${raw_region_col},{region},'
            f'RawResultsLong!${raw_technology_col}:${raw_technology_col},{technology},'
            f'RawResultsLong!${raw_year_col}:${raw_year_col},{year})'
        )
        ws.cell(row, cols["capacity_mw"]).value = (
            f"={_col_letter(ws, 'value_pj_per_year')}{row}*1000/31.5576"
        )
        ws.cell(row, cols["component_multiplier"]).value = (
            f'=SUMIFS(TechnologyMapping!${mapping_multiplier_col}:${mapping_multiplier_col},'
            f'TechnologyMapping!${mapping_technology_col}:${mapping_technology_col},{technology},'
            f'TechnologyMapping!${mapping_employment_technology_col}:${mapping_employment_technology_col},{employment_technology})'
        )
        ws.cell(row, cols["component_capacity_mw"]).value = (
            f"={capacity_mw}*{component_multiplier}"
        )
        ws.cell(row, cols["factor_value"]).value = (
            f'=SUMIFS(FactorInputs!${factor_value_col}:${factor_value_col},'
            f'FactorInputs!${factor_technology_col}:${factor_technology_col},{employment_technology},'
            f'FactorInputs!${factor_year_col}:${factor_year_col},{year},'
            f'FactorInputs!${factor_type_col}:${factor_type_col},{factor_type},'
            f'FactorInputs!${factor_job_type_col}:${factor_job_type_col},{job_type})'
        )
        ws.cell(row, cols["factor_unit"]).value = (
            f'=XLOOKUP({factor_key},FactorInputs!${factor_key_col}:${factor_key_col},'
            f'FactorInputs!${factor_unit_col}:${factor_unit_col},"")'
        )
        ws.cell(row, cols["factor_source"]).value = (
            f'=XLOOKUP({factor_key},FactorInputs!${factor_key_col}:${factor_key_col},'
            f'FactorInputs!${factor_source_col}:${factor_source_col},"")'
        )
        ws.cell(row, cols["default_source"]).value = (
            f'=XLOOKUP({factor_key},FactorInputs!${factor_key_col}:${factor_key_col},'
            f'FactorInputs!${factor_default_source_col}:${factor_default_source_col},"")'
        )
        ws.cell(row, cols["employment_fte"]).value = (
            f"={component_capacity_mw}*{factor_value}"
        )


def _set_aggregated_formulas(wb) -> None:
    ws = wb["EmploymentAggregated"]
    cols = _column_map(ws)
    raw = wb["RawResultsLong"]
    employment = wb["EmploymentResults"]
    raw_value_col = _col_letter(raw, "value_pj_per_year")
    raw_scenario_col = _col_letter(raw, "scenario")
    raw_variable_col = _col_letter(raw, "variable_name")
    raw_region_col = _col_letter(raw, "region")
    raw_technology_col = _col_letter(raw, "technology")
    raw_year_col = _col_letter(raw, "year")
    employment_fte_col = _col_letter(employment, "employment_fte")
    employment_scenario_col = _col_letter(employment, "scenario")
    employment_region_col = _col_letter(employment, "region")
    employment_technology_col = _col_letter(employment, "technology")
    employment_year_col = _col_letter(employment, "year")
    employment_variable_col = _col_letter(employment, "employment_variable")
    employment_factor_type_col = _col_letter(employment, "factor_type")
    employment_job_type_col = _col_letter(employment, "job_type")

    for row in range(2, ws.max_row + 1):
        scenario = f"{_col_letter(ws, 'scenario')}{row}"
        region = f"{_col_letter(ws, 'region')}{row}"
        technology = f"{_col_letter(ws, 'technology')}{row}"
        year = f"{_col_letter(ws, 'year')}{row}"
        variable = f"{_col_letter(ws, 'variable_name')}{row}"
        employment_variable = f"{_col_letter(ws, 'employment_variable')}{row}"
        factor_type = f"{_col_letter(ws, 'factor_type')}{row}"
        job_type = f"{_col_letter(ws, 'job_type')}{row}"

        ws.cell(row, cols["raw_lookup_key"]).value = (
            f"={_formula_key(scenario, variable, region, technology, year)}"
        )
        ws.cell(row, cols["capacity_pj_per_year"]).value = (
            f'=SUMIFS(RawResultsLong!${raw_value_col}:${raw_value_col},'
            f'RawResultsLong!${raw_scenario_col}:${raw_scenario_col},{scenario},'
            f'RawResultsLong!${raw_variable_col}:${raw_variable_col},{variable},'
            f'RawResultsLong!${raw_region_col}:${raw_region_col},{region},'
            f'RawResultsLong!${raw_technology_col}:${raw_technology_col},{technology},'
            f'RawResultsLong!${raw_year_col}:${raw_year_col},{year})'
        )
        ws.cell(row, cols["capacity_mw"]).value = (
            f"={_col_letter(ws, 'capacity_pj_per_year')}{row}*1000/31.5576"
        )
        ws.cell(row, cols["employment_fte"]).value = (
            f'=SUMIFS(EmploymentResults!${employment_fte_col}:${employment_fte_col},'
            f'EmploymentResults!${employment_scenario_col}:${employment_scenario_col},{scenario},'
            f'EmploymentResults!${employment_region_col}:${employment_region_col},{region},'
            f'EmploymentResults!${employment_technology_col}:${employment_technology_col},{technology},'
            f'EmploymentResults!${employment_year_col}:${employment_year_col},{year},'
            f'EmploymentResults!${employment_variable_col}:${employment_variable_col},{employment_variable},'
            f'EmploymentResults!${employment_factor_type_col}:${employment_factor_type_col},{factor_type},'
            f'EmploymentResults!${employment_job_type_col}:${employment_job_type_col},{job_type})'
        )


def _add_formulas_and_tables(output_path: Path) -> None:
    wb = load_workbook(output_path)
    _set_employment_formulas(wb)
    _set_aggregated_formulas(wb)
    for name in wb.sheetnames:
        wb[name].freeze_panes = "A2"
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--inputs",
        nargs="+",
        type=Path,
        default=[
            REPO_ROOT / "simulation_328_results_filtered_PD.xlsx",
            REPO_ROOT / "simulation_327_results_filtered_CN.xlsx",
            REPO_ROOT / "simulation_332_results_filtered_PA.xlsx",
        ],
        help="Input scenario result workbooks.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=REPO_ROOT / "employment_results_PD_CN_PA.xlsx",
        help="Output workbook path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    inputs = [path.resolve() for path in args.inputs]
    missing = [path for path in inputs if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing input workbook(s): {missing}")

    raw_long = load_results_long(inputs)
    employment, aggregated, diagnostics = build_employment_results(raw_long)
    write_workbook(
        raw_long=raw_long,
        employment=employment,
        aggregated=aggregated,
        diagnostics=diagnostics,
        output_path=args.output.resolve(),
    )
    print(f"raw rows: {len(raw_long)}")
    print(f"employment component rows: {len(employment)}")
    print(f"employment aggregated rows: {len(aggregated)}")
    print(f"diagnostic rows: {len(diagnostics)}")
    print(f"saved: {args.output.resolve()}")


if __name__ == "__main__":
    main()
