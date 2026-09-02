"""Chart service — lógica central de visualización.

Reemplaza ``graficas.py`` y ``graficas_comparacion.py`` del paquete
``osemosys_src`` para funcionar directamente contra la BD.  Cada función
pública devuelve un Pydantic schema listo para serialización en FastAPI.

Funciones públicas:
  - ``build_chart_data``       → gráfica single-escenario
  - ``build_comparison_data``  → gráfica multi-escenario / subplots por año
  - ``get_result_summary``     → KPIs de cabecera
  - ``get_chart_catalog``      → catálogo de tipos de gráfica disponibles
"""

from __future__ import annotations

import json
import logging
import os
from typing import Any

import pandas as pd
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.models import OsemosysOutputParamValue, OsemosysParamValue, Region, SimulationJob, Technology, Timeslice


# ──────────────────────────────────────────────────────────────────────────
#  Tipografía: registramos Nunito (bundled en `fonts/`) y la dejamos como
#  font.family default para todos los renderers matplotlib del módulo.
# ──────────────────────────────────────────────────────────────────────────
def _register_nunito_font() -> None:
    try:
        import matplotlib

        matplotlib.use("Agg")
        from matplotlib import font_manager

        fonts_dir = os.path.join(os.path.dirname(__file__), "fonts")
        if not os.path.isdir(fonts_dir):
            return
        for fname in os.listdir(fonts_dir):
            if fname.lower().endswith((".ttf", ".otf")):
                font_manager.fontManager.addfont(os.path.join(fonts_dir, fname))
        # Si Nunito quedó disponible, lo activamos como default.
        names = {f.name for f in font_manager.fontManager.ttflist}
        if "Nunito" in names:
            matplotlib.rcParams["font.family"] = "Nunito"
            matplotlib.rcParams["font.sans-serif"] = [
                "Nunito",
                "DejaVu Sans",
                "sans-serif",
            ]
    except Exception:  # pragma: no cover — no romper el módulo si falla
        pass


_register_nunito_font()
from app.schemas.scenario import ScenarioTagPublic
from app.schemas.visualization import (
    ChartCatalogItem,
    ChartDataResponse,
    ChartSeries,
    CompareChartFacetResponse,
    CompareChartResponse,
    FacetData,
    ParetoChartResponse,
    ResultSummaryResponse,
    SubplotData,
)
from app.visualization.colors import (
    asignar_grupo,
    generar_colores_tecnologias,
    _color_electricidad,
    _color_electrolisis,
    _color_por_grupo_fijo,
    _color_por_sector,
    _color_por_emision,
    _color_por_region,
    _color_transporte_grupo,
    _color_por_modo,
)
from app.visualization.labels import get_label
from app.visualization.regional import REGION_LABELS, REGIONAL_PREFIXES, strip_region, transform_regional_df
from app.visualization.configs import (
    CONFIGS,
    CONFIGS_CON_ALIAS_PWR,
    PWR_TECH_ALIASES,
    TITULOS_VARIABLES_CAPACIDAD,
    NOMBRES_COMBUSTIBLES,
    _filtro_recursos_crudo,
    _filtro_recursos_gas,
    _filtro_recursos_carbon,
    _map_electrolisis_verde,
    _map_h2_verde_azul_gris,
    _map_h2_consumo_grupo,
)
from app.visualization.configs_comparacion import CONFIGS_COMPARACION
from app.visualization.catalog_reader import (
    get_colores_emisiones,
    get_colores_grupos,
    get_colores_sector,
    get_configs_comparacion,
    get_mapa_sector,
)

logger = logging.getLogger(__name__)

# Variables principales (columnas tipadas en la BD)
_MAIN_TYPED_VARIABLES = {"Dispatch", "NewCapacity", "UnmetDemand", "AnnualEmissions"}


def _is_regional_job(db: Session, job_id: int) -> bool:
    """``True`` si el simulation_job se ejecutó en modo REGIONAL.

    Se lee directamente de ``simulation_job.simulation_type`` (heredado del
    escenario al hacer submit). Una sola query escalar.
    """
    sim_type = (
        db.query(SimulationJob.simulation_type)
        .filter(SimulationJob.id == job_id)
        .scalar()
    )
    return sim_type == "REGIONAL"


def _apply_regional_transform(
    db: Session,
    job_id: int,
    df: pd.DataFrame,
    *,
    region_filter: str | None = None,
    agrupar_por: str | None = None,
) -> pd.DataFrame:
    """No-op en NACIONAL; aplica ``transform_regional_df`` en REGIONAL.

    Debe llamarse INMEDIATAMENTE después de ``_load_variable_data`` y ANTES
    de cualquier filtro/alias/groupby que dependa de TECHNOLOGY o FUEL —
    los filtros de ``CONFIGS`` asumen códigos sin prefijo regional.
    """
    if df is None or df.empty or not _is_regional_job(db, job_id):
        return df
    return transform_regional_df(
        df, region_filter=region_filter, agrupar_por=agrupar_por
    )


# ═══════════════════════════════════════════════════════════════════════════
# 1. DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════


def _get_simulation_max_year(db: Session, job_id: int) -> int | None:
    """Último año del horizonte de simulación del job.

    Misma fuente que la página "Periodo del modelo" del escenario
    (``ScenarioService.get_scenario_period_info``): ``max(year)`` de las filas
    ``YearSplit`` en ``osemosys_param_value`` del escenario asociado al job.
    Es el set ``YEAR`` que entra realmente al modelo Pyomo (ver
    ``data_processing.py``), por lo que coincide con lo que el usuario ve al
    editar el escenario.

    Para jobs en modo Excel/SAND (sin ``scenario_id``) o cuando no hay
    ``YearSplit`` cargado, se cae al ``max(year)`` de los outputs del job
    como aproximación.
    """
    scenario_id = (
        db.query(SimulationJob.scenario_id)
        .filter(SimulationJob.id == job_id)
        .scalar()
    )
    if scenario_id is not None:
        val = (
            db.query(func.max(OsemosysParamValue.year))
            .filter(
                OsemosysParamValue.id_scenario == scenario_id,
                OsemosysParamValue.param_name == "YearSplit",
            )
            .scalar()
        )
        if val is not None:
            return int(val)
    # Fallback: Excel/SAND o YearSplit ausente.
    val = (
        db.query(func.max(OsemosysOutputParamValue.year))
        .filter(OsemosysOutputParamValue.id_simulation_job == job_id)
        .scalar()
    )
    return int(val) if val is not None else None


def _extend_years_to_sim_end(años: list[int], max_sim_year: int | None) -> list[int]:
    """Extiende una lista ordenada de años hasta ``max_sim_year`` inclusive."""
    if max_sim_year is None or not años:
        return años
    last = años[-1]
    if max_sim_year <= last:
        return años
    return años + list(range(last + 1, max_sim_year + 1))


def _load_variable_data(
    db: Session,
    job_id: int,
    variable_name: str,
) -> pd.DataFrame:
    """Carga datos de ``osemosys_output_param_value`` y devuelve un DataFrame.

    • Variables principales (Dispatch, NewCapacity, …): usa columnas tipadas
      (``technology_name``, ``fuel_name``, ``year``, ``value``).
    • Variables intermedias (ProductionByTechnology, TotalCapacityAnnual, …):
      extrae TECHNOLOGY, FUEL, YEAR del campo ``index_json``.

    Parámetros
    ----------
    db : Session
        Sesión SQLAlchemy activa.
    job_id : int
        ID del simulation_job.
    variable_name : str
        Nombre exacto de la variable a cargar.

    Retorna
    -------
    pd.DataFrame
        Columnas garantizadas: TECHNOLOGY, YEAR, VALUE.
        Columna opcional: FUEL (cuando está disponible).
    """

    # ── Consulta BD ──────────────────────────────────────────────────────
    # Hacemos LEFT JOIN con `timeslice` para recuperar el código de TS cuando
    # el output tiene `id_timeslice` poblado (Dispatch tipado; intermedias
    # como ProductionByTechnology / UseByTechnology / RateOfActivity). Las
    # filas sin TS quedan con cadena vacía y el filtro de `timeslice` en
    # `build_chart_data` simplemente no las restringe.
    rows = (
        db.query(OsemosysOutputParamValue, Timeslice.code.label("ts_code"))
        .outerjoin(Timeslice, OsemosysOutputParamValue.id_timeslice == Timeslice.id)
        .filter(
            OsemosysOutputParamValue.id_simulation_job == job_id,
            OsemosysOutputParamValue.variable_name == variable_name,
        )
        .all()
    )

    if not rows:
        return pd.DataFrame(columns=["TECHNOLOGY", "FUEL", "TIMESLICE", "YEAR", "VALUE"])

    # ── Construir DataFrame ──────────────────────────────────────────────
    if variable_name in _MAIN_TYPED_VARIABLES:
        region_ids = {r.id_region for r, _ in rows if r.id_region is not None}
        region_names = (
            dict(db.query(Region.id, Region.name).filter(Region.id.in_(region_ids)).all())
            if region_ids
            else {}
        )
        records = []
        for r, ts_code in rows:
            records.append(
                {
                    "TECHNOLOGY": r.technology_name or "",
                    "FUEL": r.fuel_name or "",
                    "REGION": region_names.get(r.id_region, "") if r.id_region is not None else "",
                    "TIMESLICE": ts_code or "",
                    "YEAR": r.year,
                    "VALUE": float(r.value),
                }
            )
        df = pd.DataFrame(records)

    else:
        # Variable intermedia → extraer de index_json
        records = []
        for r, ts_code in rows:
            idx_raw = r.index_json if r.index_json else []
            idx = idx_raw if isinstance(idx_raw, (list, tuple)) else []
            # Convenciones del pipeline:
            #   ProductionByTechnology / UseByTechnology / TotalCapacityAnnual /
            #   AccumulatedNewCapacity / AnnualTechnologyEmission:
            #     index_json = [REGION, TECHNOLOGY, FUEL?, YEAR?, ...]
            # La posición del YEAR puede variar: posición 2, 3 o 4.
            technology = str(idx[1]) if len(idx) > 1 else ""
            fuel = ""
            year = None

            if len(idx) >= 5:
                # [REGION, TECH, FUEL, ?, YEAR]  (5-element index)
                fuel = str(idx[2]) if idx[2] is not None else ""
                year = _safe_int(idx[4]) or _safe_int(idx[3])
            elif len(idx) >= 4:
                # [REGION, TECH, FUEL, YEAR]  (4-element index)
                fuel = str(idx[2]) if idx[2] is not None else ""
                year = _safe_int(idx[3])
            elif len(idx) >= 3:
                # [REGION, TECH, YEAR]  (3-element index)
                year = _safe_int(idx[2])

            records.append(
                {
                    "TECHNOLOGY": technology,
                    "FUEL": fuel,
                    "TIMESLICE": ts_code or "",
                    "YEAR": year,
                    "VALUE": float(r.value),
                }
            )
        df = pd.DataFrame(records)

    # Limpiar: descartar filas sin YEAR útil
    df = df.dropna(subset=["YEAR"])
    df["YEAR"] = df["YEAR"].astype(int)

    return df


def _load_resource_cap_input(
    db: Session, job_id: int, tech_prefix: str
) -> dict[str, float]:
    """Carga TotalTechnologyModelPeriodActivityUpperLimit desde input params.

    Para tecnologías cuyo nombre empieza con ``tech_prefix`` (ej. MINOIL).
    Retorna ``{technology_name: cap_value}`` solo para valores no-default
    (< 9,999,990). Si no hay valores configurados o no existe escenario,
    retorna dict vacío.
    """
    job = db.query(SimulationJob).filter(SimulationJob.id == job_id).first()
    if not job or not job.scenario_id:
        return {}

    names = [tech_prefix] + [f"{p}_{tech_prefix}" for p in REGIONAL_PREFIXES]
    results = (
        db.query(Technology.name, OsemosysParamValue.value)
        .join(Technology, Technology.id == OsemosysParamValue.id_technology)
        .filter(
            OsemosysParamValue.id_scenario == job.scenario_id,
            OsemosysParamValue.param_name
            == "TotalTechnologyModelPeriodActivityUpperLimit",
            Technology.name.in_(names),
        )
        .all()
    )

    caps: dict[str, float] = {}
    for name, value in results:
        v = float(value)
        if v < 9_999_990:
            key = strip_region(name)
            caps[key] = caps.get(key, 0.0) + v  # sum across regions
    return caps


def _unit_conversion_factor(un: str) -> float:
    """Factor de conversión desde PJ a la unidad solicitada."""
    if un == "GW":
        return 1.0 / 31.536
    elif un == "MW":
        return 1.0 / 0.031536
    elif un == "TWh":
        return 1.0 / 3.6
    elif un == "Gpc":
        return 1.0 / 1.0095581216
    return 1.0  # PJ


def _safe_int(val: Any) -> int | None:
    """Intenta convertir un valor a int, retorna None si falla."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def format_axis_3sig(v: Any) -> str:
    """Formatea un valor numérico como **entero** con separador de miles.

    Uso típico: ``ax.yaxis.set_major_formatter(FuncFormatter(format_axis_3sig))``.

    Reglas:
      * ``|v| >= 1`` → entero con separador de miles ("1,234")
      * ``|v| < 1`` y ≠ 0 → "0" (se trunca la parte decimal)
      * cero → "0"
    """
    import math

    if v is None:
        return "0"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if not math.isfinite(v):
        return str(v)
    if v == 0:
        return "0"
    return f"{v:,.0f}"


def _year_keep_indices(
    categories: list[Any],
    year_from: int | None,
    year_to: int | None,
) -> list[int]:
    """Índices de ``categories`` cuyo valor (parseable como año) cae en el rango.

    Categorías no parseables como entero se preservan (no son años).
    """
    if year_from is None and year_to is None:
        return list(range(len(categories)))
    keep: list[int] = []
    for i, c in enumerate(categories):
        try:
            y = int(str(c))
        except (TypeError, ValueError):
            keep.append(i)
            continue
        if year_from is not None and y < year_from:
            continue
        if year_to is not None and y > year_to:
            continue
        keep.append(i)
    return keep


def _aplicar_alias_pwr(df: pd.DataFrame) -> pd.DataFrame:
    """Reescribe ``TECHNOLOGY`` según ``PWR_TECH_ALIASES``.

    Aplica los aliases definidos en ``configs.PWR_TECH_ALIASES`` sobre la
    columna ``TECHNOLOGY``. Las filas cuya tecnología no esté en el mapa
    quedan intactas. Llamar DESPUÉS del filtro y ANTES del ``groupby``: así
    las variantes (p. ej. ``PWRSOLRTP_ZNI``) se suman bajo la tecnología
    padre (``PWRSOLRTP``) y la leyenda muestra una sola entrada con el label
    del padre.
    """
    if df is None or df.empty or "TECHNOLOGY" not in df.columns:
        return df
    if not PWR_TECH_ALIASES:
        return df
    df = df.copy()
    df["TECHNOLOGY"] = df["TECHNOLOGY"].replace(PWR_TECH_ALIASES)
    return df


def reorder_chart_series(chart: Any, order: list[str] | None) -> None:
    """Reordena ``chart.series`` in-place según ``order`` (lista de nombres).

    Series no listadas se mantienen al final en su orden natural. Series
    listadas pero no presentes en el chart se ignoran silenciosamente.

    El primer nombre del array queda arriba del stack — convención del
    proyecto (Highcharts ``yAxis.reversedStacks=true`` por defecto, y los
    renderers matplotlib iteran en ``reversed()``).
    """
    if not order:
        return
    series = getattr(chart, "series", None)
    if not series:
        return
    by_name = {s.name: s for s in series}
    used: set[str] = set()
    new_series: list[Any] = []
    for name in order:
        s = by_name.get(name)
        if s is not None and name not in used:
            new_series.append(s)
            used.add(name)
    for s in series:
        if s.name not in used:
            new_series.append(s)
    chart.series = new_series


def filter_chart_series(chart: Any, names: list[str] | None) -> None:
    """Restringe ``chart.series`` a las que están en ``names`` (in-place).

    Match exacto por ``series.name``. ``None`` o lista vacía → no-op.
    Mantiene el orden actual del chart, no el de ``names``.
    """
    if not names:
        return
    series = getattr(chart, "series", None)
    if not series:
        return
    allowed = set(names)
    chart.series = [s for s in series if s.name in allowed]


def filter_chart_categories(chart: Any, keep_categories: list[str] | None) -> None:
    """Restringe ``chart.categories`` a las que están en ``keep_categories``.

    El filtrado se aplica también a cada ``series.data`` por índice posicional
    para mantener la alineación. Match por ``str(category)`` (los años suelen
    venir como ``int``). ``None`` o lista vacía → no-op.
    """
    if not keep_categories:
        return
    cats = getattr(chart, "categories", None)
    series = getattr(chart, "series", None)
    if not cats or series is None:
        return
    allowed = {str(k) for k in keep_categories}
    keep_idx = [i for i, c in enumerate(cats) if str(c) in allowed]
    if not keep_idx or len(keep_idx) == len(cats):
        return
    chart.categories = [cats[i] for i in keep_idx]
    for s in series:
        data = getattr(s, "data", None)
        if data is None:
            continue
        s.data = [data[i] for i in keep_idx if i < len(data)]


def apply_period_years(chart: Any, period: int | None) -> None:
    """Filtra ``categories`` (años) tomando uno cada ``period``, in-place.

    El primer año visible es el primer índice donde el offset (idx desde el
    primer año-categoría) es múltiplo de ``period``. El último año siempre
    se preserva (aunque rompa la cadencia) para que la tabla cierre en el
    horizonte real del modelo.

    Categorías no parseables como año se preservan tal cual.
    Si ``period`` es ``None`` o ``< 2``, no se hace nada.
    """
    if period is None or period < 2:
        return
    cats = getattr(chart, "categories", None)
    series = getattr(chart, "series", None)
    if not cats or series is None:
        return
    # Encontrar índices de años parseables.
    year_indices: list[int] = []
    year_values: list[int] = []
    for i, c in enumerate(cats):
        try:
            y = int(str(c))
            year_indices.append(i)
            year_values.append(y)
        except (TypeError, ValueError):
            continue
    if not year_indices:
        return
    base = year_values[0]
    keep = set()
    # No-año (otros): siempre conservar.
    for i in range(len(cats)):
        if i not in year_indices:
            keep.add(i)
    # Cada ``period`` años desde el primer año visible.
    for idx, year in zip(year_indices, year_values):
        if (year - base) % period == 0:
            keep.add(idx)
    # Garantizar que el último año esté presente.
    keep.add(year_indices[-1])
    keep_sorted = sorted(keep)
    chart.categories = [cats[i] for i in keep_sorted]
    for s in series:
        data = getattr(s, "data", None)
        if data is None:
            continue
        s.data = [data[i] for i in keep_sorted if i < len(data)]


def apply_cumulative_series(chart: Any) -> None:
    """Reemplaza cada ``series.data`` por su suma acumulada, in-place.

    Útil para tablas de "capacidad acumulada", "emisiones acumuladas", etc.
    NaN/None se trata como 0 para no destruir el acumulador.
    """
    import math

    series = getattr(chart, "series", None)
    if not series:
        return
    for s in series:
        data = getattr(s, "data", None)
        if data is None:
            continue
        cum: list[float] = []
        running = 0.0
        for v in data:
            try:
                f = float(v)
                if not math.isfinite(f):
                    f = 0.0
            except (TypeError, ValueError):
                f = 0.0
            running += f
            cum.append(running)
        s.data = cum


def filter_chart_by_year_range(
    chart: Any,
    year_from: int | None,
    year_to: int | None,
) -> None:
    """Filtra in-place ``ChartDataResponse`` o ``CompareChartFacetResponse``.

    - ``ChartDataResponse``: corta ``categories`` y cada ``series.data``.
    - ``CompareChartFacetResponse``: aplica el corte por cada faceta.

    Si ambos extremos son ``None``, no hace nada. Categorías no-año se
    preservan tal cual.
    """
    if year_from is None and year_to is None:
        return
    # Single chart / line-total (categories + series)
    cats = getattr(chart, "categories", None)
    series = getattr(chart, "series", None)
    if cats is not None and series is not None:
        keep = _year_keep_indices(cats, year_from, year_to)
        chart.categories = [cats[i] for i in keep]
        for s in series:
            data = getattr(s, "data", None)
            if data is None:
                continue
            s.data = [data[i] for i in keep if i < len(data)]
    # Facet (per-facet categories + series)
    facets = getattr(chart, "facets", None)
    if facets is not None:
        for f in facets:
            f_cats = getattr(f, "categories", None)
            f_series = getattr(f, "series", None)
            if f_cats is None or f_series is None:
                continue
            keep = _year_keep_indices(f_cats, year_from, year_to)
            f.categories = [f_cats[i] for i in keep]
            for s in f_series:
                data = getattr(s, "data", None)
                if data is None:
                    continue
                s.data = [data[i] for i in keep if i < len(data)]


# ═══════════════════════════════════════════════════════════════════════════
# 2. HELPERS DE TRANSFORMACIÓN (ports de graficas_comparacion.py)
# ═══════════════════════════════════════════════════════════════════════════


def _fuel_to_group(row) -> str:
    """Normaliza una fila (FUEL + TECHNOLOGY) al código base de grupo.

    Códigos FUEL con sufijos numéricos (ELC002, NGS002…) se mapean a su
    clave base (ELC, NGS…) usando asignar_grupo, para que coincidan con
    COLORES_GRUPOS y DISPLAY_NAMES.  OIL se desambigua por tecnología.
    """
    fuel = row.get("FUEL", "")
    tech = str(row.get("TECHNOLOGY", ""))
    if fuel == "OIL":
        if "MINOIL_3PES" in tech:
            return "MINOIL_3PES"
        if "MINOIL_2MID" in tech:
            return "MINOIL_2MID"
        if "MINOIL_1LIV" in tech:
            return "MINOIL_1LIV"
        if "MINOIL" in tech:
            return "MINOIL"
    return asignar_grupo(fuel) if fuel else "OTRO"


_PREFIJO_TECH_LISTS: dict[str, str] = {
    "DEMIND": "TECNOLOGIAS_INDUSTRIALES",
    "DEMRES": "TECNOLOGIAS_RESIDENCIALES",
    "DEMTRA": "TECNOLOGIAS_TRANSPORTE",
    "DEMTER": "TECNOLOGIAS_TERCIARIO",
}


def _resolve_tech_list(prefijo: str | tuple[str, ...]) -> list[str]:
    from app.visualization.catalog_cache import get_catalog_cache

    resolver = get_catalog_cache().filter_resolver
    if isinstance(prefijo, tuple):
        combined: list[str] = []
        for p in prefijo:
            code = _PREFIJO_TECH_LISTS.get(p)
            if code:
                combined.extend(resolver.tech(code))
        return combined
    code = _PREFIJO_TECH_LISTS.get(prefijo)
    if code:
        return list(resolver.tech(code))
    return []


def _filtrar_df(
    df: pd.DataFrame,
    prefijo: str | tuple[str, ...],
    sub_filtro: str | None,
    loc: str | None,
) -> pd.DataFrame:
    """Aplica filtro por listas de tecnologías, sub_filtro y localización.
    """
    if df.empty:
        return df

    tech_list = _resolve_tech_list(prefijo)
    if not tech_list:
        return df.iloc[:0]

    mask = df["TECHNOLOGY"].isin(tech_list)

    if sub_filtro == "CARRETERA":
        from app.visualization.catalog_cache import get_catalog_cache

        carretera = get_catalog_cache().filter_resolver.tech(
            "TECNOLOGIAS_TRANSPORTE_CARRETERA"
        )
        mask &= df["TECHNOLOGY"].isin(carretera)
    elif sub_filtro:
        mask &= df["TECHNOLOGY"].isin(
            [t for t in tech_list if sub_filtro in t]
        )

    if loc == "URB":
        from app.visualization.catalog_cache import get_catalog_cache

        resolver = get_catalog_cache().filter_resolver
        mask &= df["TECHNOLOGY"].isin(resolver.tech("TEC_RES_URB"))
    elif loc == "RUR":
        from app.visualization.catalog_cache import get_catalog_cache

        resolver = get_catalog_cache().filter_resolver
        mask &= df["TECHNOLOGY"].isin(resolver.tech("TEC_RES_RUR"))
    elif loc == "ZNI":
        from app.visualization.catalog_cache import get_catalog_cache

        resolver = get_catalog_cache().filter_resolver
        mask &= df["TECHNOLOGY"].isin(resolver.tech("TEC_RES_ZNI"))

    return df[mask].copy()


# Mapeo de código de uso de transporte a nombre de grupo
_TRANSPORTE_USO_A_GRUPO: dict[str, str] = {
    "MOT": "Motos",
    "LDV": "Livianos",
    "TAX": "Livianos",
    "FWD": "Livianos",
    "BUS": "Buses",
    "MIC": "Microbuses",
    "TCK": "Carga",
    "STT": "Carga",
    "BOT": "Barcos",
    "SHP": "Barcos",
    "MET": "Metro",
    "AVI": "Aviación",
    "AIR": "Aviación",
}


def _transporte_techs() -> frozenset[str]:
    from app.visualization.catalog_cache import get_catalog_cache

    return get_catalog_cache().filter_resolver.tech("TECNOLOGIAS_TRANSPORTE")


def _export_carbon_techs() -> frozenset[str]:
    from app.visualization.catalog_cache import get_catalog_cache

    return get_catalog_cache().filter_resolver.tech("TECNOLOGIAS_EXPORTACION_CARBON")


def _map_transporte_grupo(tech_code: str) -> str:
    """Clasifica un código DEMTRA en grupo de transporte."""
    if not isinstance(tech_code, str) or tech_code not in _transporte_techs():
        return "Otros"
    rest = tech_code[len("DEMTRA"):]

    # Eliminar sufijos de eficiencia y área
    from app.visualization.labels import _EFIC, _AREA  # lazy: evita circular
    for suffix in _EFIC:
        if rest.endswith(suffix):
            rest = rest[:-len(suffix)]
            break
    for suffix in _AREA:
        if rest.endswith(suffix):
            rest = rest[:-len(suffix)]
            break

    # Buscar código de uso al final del restante
    for uso_code in _TRANSPORTE_USO_A_GRUPO:
        if rest.endswith(uso_code):
            return _TRANSPORTE_USO_A_GRUPO[uso_code]

    # Fallback: buscar cualquier código conocido en cualquier posición
    for uso_code in _TRANSPORTE_USO_A_GRUPO:
        if uso_code in rest:
            return _TRANSPORTE_USO_A_GRUPO[uso_code]

    return "Otros"


# Mapeo de código de uso de transporte a modo (agrupación MODO)
_MODOS_TRANSPORTE_MAP: dict[str, str] = {
    "MOT": "CARRETERA",
    "LDV": "CARRETERA",
    "TAX": "CARRETERA",
    "FWD": "CARRETERA",
    "BUS": "CARRETERA",
    "MIC": "CARRETERA",
    "TCK": "CARRETERA",
    "STT": "CARRETERA",
    "AVI": "AVI",
    "AIR": "AVI",
    "BOT": "BOT",
    "SHP": "BOT",
    "MET": "MET",
}


def _map_transporte_modo(tech_code: str) -> str:
    """Clasifica un código DEMTRA en modo de transporte (CARRETERA, AVI, BOT, MET)."""
    if not isinstance(tech_code, str) or tech_code not in _transporte_techs():
        return "Otros"
    rest = tech_code[len("DEMTRA"):]

    # Eliminar sufijos de eficiencia y área
    from app.visualization.labels import _EFIC, _AREA  # lazy: evita circular
    for suffix in _EFIC:
        if rest.endswith(suffix):
            rest = rest[:-len(suffix)]
            break
    for suffix in _AREA:
        if rest.endswith(suffix):
            rest = rest[:-len(suffix)]
            break

    # Buscar código de modo al final del restante
    for modo_code in _MODOS_TRANSPORTE_MAP:
        if rest.endswith(modo_code):
            return _MODOS_TRANSPORTE_MAP[modo_code]

    # Fallback: buscar cualquier código conocido en cualquier posición
    for modo_code in _MODOS_TRANSPORTE_MAP:
        if modo_code in rest:
            return _MODOS_TRANSPORTE_MAP[modo_code]

    return "Otros"


def _sector_labels(tech_series: pd.Series) -> pd.Series:
    """Asignación vectorizada de sector, incluyendo PWR → Generación Electricidad.

    Usa ``MAPA_SECTOR`` desde BD (sembrado al arrancar el API; fallback al
    dict hardcoded sólo si la BD aún no está poblada).
    """
    labels = tech_series.str[:6].map(get_mapa_sector())
    pwr_mask = labels.isna() & tech_series.str.startswith("PWR")
    labels = labels.where(~pwr_mask, "Generación Electricidad")
    return labels.fillna("Otros")


def _asignar_categoria(
    df: pd.DataFrame,
    agrupacion: str,
) -> pd.DataFrame:
    """Crea columna CATEGORIA según el tipo de agrupación.

    Port de ``graficas_comparacion._asignar_categoria``.
    """
    df = df.copy()

    if agrupacion == "TECNOLOGIA":
        df["CATEGORIA"] = df["TECHNOLOGY"]

    elif agrupacion in ("COMBUSTIBLE", "FUEL"):
        if "FUEL" in df.columns:
            df["_TECH_FUEL"] = (
                df["TECHNOLOGY"].astype(str) + "_" + df["FUEL"].astype(str)
            )
        else:
            df["_TECH_FUEL"] = df["TECHNOLOGY"].astype(str)

        df["CATEGORIA"] = df["_TECH_FUEL"].apply(asignar_grupo)
        df = df.drop(columns="_TECH_FUEL")

    elif agrupacion == "SECTOR":
        df["CATEGORIA"] = _sector_labels(df["TECHNOLOGY"])

    elif agrupacion == "EMISION":
        # Para AnnualTechnologyEmission, FUEL contiene el tipo de emisión
        df["CATEGORIA"] = df["FUEL"] if "FUEL" in df.columns else "?"

    elif agrupacion == "TRANSPORTE_GRUPO":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_transporte_grupo)

    elif agrupacion == "MODO":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_transporte_modo)

    elif agrupacion == "ELECTROLISIS":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_electrolisis_verde)

    elif agrupacion == "H2_CONSUMO":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_h2_consumo_grupo)

    return df


def _convertir_unidades(df: pd.DataFrame, un: str) -> pd.DataFrame:
    """Convierte columna VALUE a las unidades solicitadas.

    Port de ``graficas_comparacion._convertir_unidades``.
    """
    df = df.copy()
    if un == "GW":
        df["VALUE"] /= 31.536
    elif un == "MW":
        df["VALUE"] /= 0.031536
    elif un == "TWh":
        df["VALUE"] /= 3.6
    elif un == "Gpc":
        df["VALUE"] /= 1.0095581216
    # PJ: unidad base, sin conversión
    return df


def _convertir_por_tecnologia(df: pd.DataFrame, un: str, factor_map: dict[str, dict[str, float]]) -> pd.DataFrame:
    """Aplica factores de conversión por tecnología (p. ej. PJ → kton).

    Opera sobre el DataFrame pre-agregación, cuando aún existe la
    columna ``TECHNOLOGY``.  Si ``un`` no está en ``factor_map``,
    devuelve el DataFrame sin cambios.
    """
    tech_factors = factor_map.get(un)
    if tech_factors is None:
        return df
    df = df.copy()
    factor_series = df["TECHNOLOGY"].map(tech_factors).fillna(1.0)
    df["VALUE"] /= factor_series
    return df


def _convertir_unidades_emision(df: pd.DataFrame, un: str) -> pd.DataFrame:
    """Convierte emisiones GEI entre MtCO₂eq y ktCO₂eq.

    Base de datos: MtCO₂eq. Multiplicar × 1000 para obtener ktCO₂eq.
    """
    df = df.copy()
    if un == "ktCO2eq":
        df["VALUE"] *= 1000.0
    return df


def _emision_unit_label(un: str, es_emision_kt: bool) -> str:
    """Devuelve la etiqueta de unidad correcta para gráficas de emisión."""
    if es_emision_kt:
        return "kt"
    return "ktCO₂eq" if un == "ktCO2eq" else "MtCO₂eq"


def _color_map_comparison(
    agrupacion: str,
    categorias_unicas: list[str],
) -> dict[str, str]:
    """Devuelve ``{categoria: color_hex}`` para gráficas de comparación.

    Port de ``graficas_comparacion._color_map``.
    """
    if agrupacion in ("COMBUSTIBLE", "FUEL"):
        palette = get_colores_grupos()
        return {c: palette.get(c, "#999999") for c in categorias_unicas}

    if agrupacion == "SECTOR":
        palette = get_colores_sector()
        return {c: palette.get(c, "#999999") for c in categorias_unicas}

    if agrupacion == "EMISION":
        palette = get_colores_emisiones()
        return {c: palette.get(c, "#999999") for c in categorias_unicas}

    # TECNOLOGIA: reutiliza generar_colores_tecnologias de colors.py
    df_tmp = pd.DataFrame({"COLOR": list(categorias_unicas)})
    colores_lista, orden_lista = generar_colores_tecnologias(df_tmp, "COLOR")
    return dict(zip(orden_lista, colores_lista))


def _build_factor_planta_data(
    db: Session,
    job_id: int,
    chart_tipo: str,
    cfg: dict,
    title: str,
    sub_filtro: str | None,
    loc: str | None,
) -> ChartDataResponse:
    """CF = Producción[PJ] / TotalCapacityAnnual[PJ] × 100 %

    Ambas variables están en PJ (baseline del modelo).
    TotalCapacityAnnual[PJ] = capacidad[GW] × 31.536 = energía máxima anual posible.
    CF = Energía real / Energía máxima posible ∈ [0 %, 100 %].
    """
    filtro_fn = cfg.get("filtro")

    df_cap = _load_variable_data(db, job_id, "TotalCapacityAnnual")
    df_prd = _load_variable_data(db, job_id, "ProductionByTechnology")

    # Strip prefijos regionales antes del filtro (jobs REGIONAL).
    df_cap = _apply_regional_transform(db, job_id, df_cap)
    df_prd = _apply_regional_transform(db, job_id, df_prd)

    if filtro_fn is not None:
        df_cap = filtro_fn(df_cap, sub_filtro=sub_filtro, loc=loc)
        df_prd = filtro_fn(df_prd, sub_filtro=sub_filtro, loc=loc)

    if df_cap.empty or df_prd.empty:
        return ChartDataResponse(categories=[], series=[], title=title, yAxisLabel="%")

    cap_agg = df_cap.groupby(["TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    prd_agg = df_prd.groupby(["TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    prd_agg = prd_agg.rename(columns={"VALUE": "PRODUCTION"})

    df = cap_agg.merge(prd_agg, on=["TECHNOLOGY", "YEAR"], how="inner")
    df = df[df["VALUE"] > 1e-6].copy()

    if df.empty:
        return ChartDataResponse(categories=[], series=[], title=title, yAxisLabel="%")

    # Ambos en PJ → ratio directo, sin conversión adicional
    df["CF"] = (df["PRODUCTION"] / df["VALUE"] * 100.0).clip(0, 100)
    df["COLOR"] = df["TECHNOLOGY"]

    color_fn = cfg.get("color_fn")
    if color_fn is not None:
        colores_ordenados, orden_color = color_fn(df, "COLOR")
    else:
        orden_color = sorted(df["COLOR"].unique())
        colores_ordenados = ["#999999"] * len(orden_color)

    color_dict = dict(zip(orden_color, colores_ordenados))
    años = sorted(df["YEAR"].unique())
    categories = [str(a) for a in años]

    from app.services.chart_series_config_service import (
        apply_global_series_config,
        normalize_agrupar_por,
    )

    ap = normalize_agrupar_por(cfg.get("agrupar_por"), cfg.get("agrupar_por"))
    stack_items = apply_global_series_config(
        db,
        tipo=chart_tipo,
        agrupar_por=ap,
        orden_color=orden_color,
        color_dict=color_dict,
        default_name=lambda code: get_label(str(code)),
    )

    series: list[ChartSeries] = []
    for tech, series_color, series_name in stack_items:
        df_tech = df[df["COLOR"] == tech]
        valor_por_año = {int(row["YEAR"]): row["CF"] for _, row in df_tech.iterrows()}
        data = [round(valor_por_año.get(a, 0.0), 4) for a in años]
        series.append(
            ChartSeries(
                name=series_name,
                data=data,
                color=series_color,
                stack="default",
            )
        )

    return ChartDataResponse(
        categories=categories, series=series, title=title, yAxisLabel="%"
    )


# ═══════════════════════════════════════════════════════════════════════════
# 2b. build_recursos_vs_demanda_data — RECURSOS Y RESERVAS
# ═══════════════════════════════════════════════════════════════════════════

_TECH_DISPLAY: dict[str, str] = {
    "MINOIL_1LIV": "Crudo liviano",
    "MINOIL_2MID": "Crudo intermedio",
    "MINOIL_3PES": "Crudo pesado",
    "MINOIL": "Crudo genérico",
}

_PRODUCTION_COLORS: dict[str, str] = {
    "MINOIL_1LIV": "#3b82f6",
    "MINOIL_2MID": "#f59e0b",
    "MINOIL_3PES": "#6b7280",
    "MINOIL": "#a855f7",
}

_REMAINING_COLORS: dict[str, str] = {
    "MINOIL_1LIV": "#1e3a5f",
    "MINOIL_2MID": "#92400e",
    "MINOIL_3PES": "#374151",
    "MINOIL": "#581c87",
}

_GAS_TECH_DISPLAY: dict[str, str] = {
    "MINNGS": "Gas natural nacional",
}

_GAS_PRODUCTION_COLORS: dict[str, str] = {
    "MINNGS": "#10b981",
}

_GAS_REMAINING_COLORS: dict[str, str] = {
    "MINNGS": "#065f46",
}

_COAL_DOMESTIC_COLOR = "#dc2626"
_COAL_EXPORT_COLOR = "#2563eb"
_COAL_LIMIT_COLOR = "#000000"


def _get_recursos_title(tipo: str) -> str:
    """Retorna el título base para una gráfica de recursos."""
    M = {
        "recursos_vs_demanda": "Recursos y reservas vs Demanda (Crudo)",
        "recursos_vs_demanda_gas": "Recursos y reservas vs Demanda (Gas Natural)",
        "recursos_vs_demanda_carbon": "Recursos y reservas vs Demanda (Carbón)",
    }
    return M.get(tipo, "Recursos y reservas")


def _load_annual_activity_limit_input(
    db: Session, job_id: int, tech_prefix: str,
) -> dict[int, float]:
    """Carga TotalTechnologyAnnualActivityUpperLimit desde input params.

    Retorna ``{year: total_limit}`` sumando todas las tecnologías que
    empiezan con ``tech_prefix`` (ej. MINCOA). Filtra valores default
    (>= 9,999,990).
    """
    job = db.query(SimulationJob).filter(SimulationJob.id == job_id).first()
    if not job or not job.scenario_id:
        return {}

    names = [tech_prefix] + [f"{p}_{tech_prefix}" for p in REGIONAL_PREFIXES]
    results = (
        db.query(Technology.name, OsemosysParamValue.year, OsemosysParamValue.value)
        .join(Technology, Technology.id == OsemosysParamValue.id_technology)
        .filter(
            OsemosysParamValue.id_scenario == job.scenario_id,
            OsemosysParamValue.param_name
            == "TotalTechnologyAnnualActivityUpperLimit",
            Technology.name.in_(names),
        )
        .all()
    )

    limits: dict[int, float] = {}
    for name, year, value in results:
        v = float(value)
        if v < 9_999_990 and year is not None:
            limits[year] = limits.get(year, 0.0) + v
    return limits


def _build_recursos_production_total(
    db: Session, job_id: int, tipo: str, un: str = "PJ",
) -> dict[int, float] | None:
    """Retorna ``{year: total_production}`` en unidades ``un``.

    - ``recursos_vs_demanda_gas``: ProductionByTechnology(MINNGS)
    - ``recursos_vs_demanda_carbon``: UseByTechnology(FUEL=COA, no EXPCOA)
                                      + ProductionByTechnology(EXPCOA)
    """
    if tipo == "recursos_vs_demanda_gas":
        df_gas = _build_recursos_vs_demanda_gas_df(db, job_id, un=un)
        if df_gas is None or df_gas.empty:
            return None
        prod = df_gas.groupby("YEAR")["PRODUCTION"].sum().to_dict()
        return prod if prod else None

    if tipo == "recursos_vs_demanda_carbon":
        df_use = _load_variable_data(db, job_id, "UseByTechnology")
        if df_use.empty:
            return None
        df_use = _apply_regional_transform(db, job_id, df_use)
        df_use = _filtro_recursos_carbon(df_use)
        domestic = df_use[df_use["FUEL"].str.startswith("COA")].groupby("YEAR")["VALUE"].sum().to_dict()

        df_prod = _load_variable_data(db, job_id, "ProductionByTechnology")
        if not df_prod.empty:
            df_prod = _apply_regional_transform(db, job_id, df_prod)
            df_prod = _filtro_recursos_carbon(df_prod)
            export = df_prod[df_prod["TECHNOLOGY"].isin(_export_carbon_techs())].groupby("YEAR")["VALUE"].sum().to_dict()
        else:
            export = {}

        factor = _unit_conversion_factor(un)
        all_years = sorted(set(domestic.keys()) | set(export.keys()))
        result = {}
        for y in all_years:
            total = (domestic.get(y, 0.0) + export.get(y, 0.0)) * factor
            result[y] = round(total, 6)
        return result if result else None

    return None


def _build_recursos_vs_demanda_gas_df(
    db: Session, job_id: int, un: str = "PJ",
) -> pd.DataFrame | None:
    """Helper que retorna un DataFrame con columnas:
    YEAR, TECHNOLOGY, PRODUCTION, REMAINING, HAS_RESOURCE

    Retorna None si no hay datos.
    """
    resource_caps = _load_resource_cap_input(db, job_id, "MINNGS")

    df = _load_variable_data(db, job_id, "ProductionByTechnology")
    if df.empty:
        return None

    df = _apply_regional_transform(db, job_id, df)
    df_gas = _filtro_recursos_gas(df)
    if df_gas.empty:
        return None

    df_agg = df_gas.groupby(["TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    all_techs = sorted(df_agg["TECHNOLOGY"].unique())

    prod_totals = df_agg.groupby("TECHNOLOGY")["VALUE"].sum().to_dict()

    initial: dict[str, float] = {}
    for tech in all_techs:
        if tech in resource_caps:
            initial[tech] = resource_caps[tech]
        else:
            initial[tech] = prod_totals.get(tech, 0.0)

    factor = _unit_conversion_factor(un)
    años = sorted(df_agg["YEAR"].unique())

    rows: list[dict] = []
    for tech in all_techs:
        tech_rows = df_agg[df_agg["TECHNOLOGY"] == tech]
        prod_by_year = dict(zip(tech_rows["YEAR"], tech_rows["VALUE"]))
        cap = initial.get(tech, 0.0)
        cum = 0.0
        for a in años:
            prod = prod_by_year.get(a, 0.0)
            cum += prod
            rem = max(0.0, cap - cum)
            rows.append({
                "YEAR": a,
                "TECHNOLOGY": tech,
                "PRODUCTION": round(prod * factor, 6),
                "REMAINING": round(rem * factor, 6) if cap > 0 else 0.0,
                "HAS_RESOURCE": 1.0 if cap > 0 else 0.0,
            })

    return pd.DataFrame(rows)


def build_recursos_vs_demanda_gas_data(
    db: Session,
    job_id: int,
    un: str = "PJ",
) -> ChartDataResponse:
    """Construye datos para Recursos y reservas vs Demanda (Gas Natural).

    Para la tecnología MINNGS:
      1. Carga el recurso inicial desde
         TotalTechnologyModelPeriodActivityUpperLimit (input param).
      2. Carga la producción anual desde ProductionByTechnology (output).
      3. Calcula ``remaining[t] = initial - cumulative_production[t]``.

    Retorna 2 series:
      - 1 barra apilada (stack="produccion"): producción anual.
      - 1 línea (chart_type="line", sin stack): recurso remanente.
    """
    title = f"Recursos y reservas vs Demanda (Gas Natural) ({un})"

    df_gas = _build_recursos_vs_demanda_gas_df(db, job_id, un=un)
    if df_gas is None or df_gas.empty:
        return ChartDataResponse(
            categories=[], series=[], title=title, yAxisLabel=un
        )

    años = sorted(df_gas["YEAR"].unique())
    categories = [str(a) for a in años]
    all_techs = sorted(df_gas["TECHNOLOGY"].unique())

    series: list[ChartSeries] = []

    for tech in all_techs:
        tech_display = _GAS_TECH_DISPLAY.get(tech, tech).lower()
        tech_rows = df_gas[df_gas["TECHNOLOGY"] == tech]

        prod_by_year = dict(zip(tech_rows["YEAR"], tech_rows["PRODUCTION"]))
        rem_by_year = dict(zip(tech_rows["YEAR"], tech_rows["REMAINING"]))
        has_res = tech_rows["HAS_RESOURCE"].iloc[0] > 0

        # Serie de producción
        prod_data = [prod_by_year.get(a, 0.0) for a in años]
        series.append(
            ChartSeries(
                name=f"Producción {tech_display}",
                data=prod_data,
                color=_GAS_PRODUCTION_COLORS.get(tech, "#10b981"),
                stack="produccion",
            )
        )

        # Serie de recurso remanente
        if has_res:
            rem_data = [rem_by_year.get(a, 0.0) for a in años]
            series.append(
                ChartSeries(
                    name=f"Recurso remanente {tech_display}",
                    data=rem_data,
                    color=_GAS_REMAINING_COLORS.get(tech, "#065f46"),
                    stack=None,
                    chart_type="line",
                )
            )

    return ChartDataResponse(
        categories=categories,
        series=series,
        title=title,
        yAxisLabel=un,
    )


def build_recursos_vs_demanda_carbon_data(
    db: Session,
    job_id: int,
    un: str = "PJ",
) -> ChartDataResponse:
    """Construye datos para Recursos y reservas vs Demanda (Carbón).

    Barras apiladas (stack="carbon"):
      - Demanda interna: UseByTechnology(FUEL=COA, excluyendo EXPCOA)
      - Exportaciones:   ProductionByTechnology(EXPCOA)
    Línea:
      - Límite anual:    TotalTechnologyAnnualActivityUpperLimit(MINCOA)
    """
    title = f"Recursos y reservas vs Demanda (Carbón) ({un})"
    factor = _unit_conversion_factor(un)

    # 1. Demanda interna: UseByTechnology(FUEL=COA, excluir EXPCOA)
    df_use = _load_variable_data(db, job_id, "UseByTechnology")
    if df_use.empty:
        return ChartDataResponse(
            categories=[], series=[], title=title, yAxisLabel=un
        )
    df_use = _apply_regional_transform(db, job_id, df_use)
    df_use = _filtro_recursos_carbon(df_use)
    domestic_by_year = (
        df_use[df_use["FUEL"].str.startswith("COA")].groupby("YEAR")["VALUE"].sum().to_dict()
    )

    # 2. Exportaciones: ProductionByTechnology(EXPCOA)
    df_prod = _load_variable_data(db, job_id, "ProductionByTechnology")
    if not df_prod.empty:
        df_prod = _apply_regional_transform(db, job_id, df_prod)
        df_prod = _filtro_recursos_carbon(df_prod)
        export_by_year = (
            df_prod[df_prod["TECHNOLOGY"].isin(_export_carbon_techs())]
            .groupby("YEAR")["VALUE"].sum().to_dict()
        )
    else:
        export_by_year = {}

    # 3. Límite anual: TotalTechnologyAnnualActivityUpperLimit(MINCOA)
    annual_limits = _load_annual_activity_limit_input(db, job_id, "MINCOA")

    all_years = sorted(
        set(domestic_by_year.keys())
        | set(export_by_year.keys())
        | set(annual_limits.keys())
    )

    categories = [str(a) for a in all_years]

    series: list[ChartSeries] = []

    # Barra: Demanda interna
    dom_data = [round(domestic_by_year.get(a, 0.0) * factor, 6) for a in all_years]
    series.append(
        ChartSeries(
            name="Demanda interna de carbón",
            data=dom_data,
            color=_COAL_DOMESTIC_COLOR,
            stack="carbon",
        )
    )

    # Barra: Exportaciones
    exp_data = [round(export_by_year.get(a, 0.0) * factor, 6) for a in all_years]
    series.append(
        ChartSeries(
            name="Exportaciones de carbón",
            data=exp_data,
            color=_COAL_EXPORT_COLOR,
            stack="carbon",
        )
    )

    # Línea: Límite anual de producción
    lim_data = [round(annual_limits.get(a, 0.0) * factor, 6) for a in all_years]
    if any(v > 0 for v in lim_data):
        series.append(
            ChartSeries(
                name="Capacidad máxima de producción anual",
                data=lim_data,
                color=_COAL_LIMIT_COLOR,
                stack=None,
                chart_type="line",
            )
        )

    return ChartDataResponse(
        categories=categories,
        series=series,
        title=title,
        yAxisLabel=un,
    )


def build_recursos_vs_demanda_data(
    db: Session,
    job_id: int,
    un: str = "PJ",
) -> ChartDataResponse:
    """Construye datos para Recursos y reservas vs Demanda (Crudo).

    Para cada categoría MINOIL (1LIV, 2MID, 3PES):
      1. Carga el recurso inicial desde
         TotalTechnologyModelPeriodActivityUpperLimit (input param).
      2. Carga la producción anual desde ProductionByTechnology (output).
      3. Calcula ``remaining[t] = initial - cumulative_production[t]``.

    Retorna 6 series por categoría de crudo:
      - 3 áreas/barras apiladas (stack="produccion"): producción anual por tipo.
      - 3 líneas (chart_type="line", sin stack): recurso remanente por tipo.
    """
    title = f"Recursos y reservas vs Demanda (Crudo) ({un})"

    # 1. Cargar recursos iniciales desde input params (si existen)
    resource_caps = _load_resource_cap_input(db, job_id, "MINOIL")

    # 2. Cargar producción anual desde output
    df = _load_variable_data(db, job_id, "ProductionByTechnology")
    if df.empty:
        return ChartDataResponse(
            categories=[], series=[], title=title, yAxisLabel=un
        )

    # Strip prefijos regionales antes del filtro (jobs REGIONAL).
    df = _apply_regional_transform(db, job_id, df)

    # Filtrar solo tecnologías MINOIL
    df_minoil = _filtro_recursos_crudo(df)
    if df_minoil.empty:
        return ChartDataResponse(
            categories=[], series=[], title=title, yAxisLabel=un
        )

    df_agg = df_minoil.groupby(
        ["TECHNOLOGY", "YEAR"], as_index=False
    )["VALUE"].sum()

    all_techs = sorted(df_agg["TECHNOLOGY"].unique())

    # 3. Determinar recurso inicial por tecnología.
    # Si no hay cap configurado (< 9.999.990), usar producción total acumulada
    # como recurso (la gráfica muestra "este fue el total extraído").
    prod_totals = df_agg.groupby("TECHNOLOGY")["VALUE"].sum().to_dict()

    initial: dict[str, float] = {}
    for tech in all_techs:
        if tech in resource_caps:
            initial[tech] = resource_caps[tech]
        else:
            initial[tech] = prod_totals.get(tech, 0.0)

    años = sorted(df_agg["YEAR"].unique())
    categories = [str(a) for a in años]

    series: list[ChartSeries] = []
    factor = _unit_conversion_factor(un)

    # 4. Construir series: producción (área) + recurso remanente (línea) por tipo
    for tech in all_techs:
        tech_display = _TECH_DISPLAY.get(tech, tech).lower()
        tech_rows = df_agg[df_agg["TECHNOLOGY"] == tech]
        prod_by_year = dict(
            zip(tech_rows["YEAR"], tech_rows["VALUE"])
        )

        # 4a. Serie de producción como área/barra
        prod_data = [
            round(prod_by_year.get(a, 0.0) * factor, 6) for a in años
        ]
        series.append(
            ChartSeries(
                name=f"Producción {tech_display}",
                data=prod_data,
                color=_PRODUCTION_COLORS.get(tech, "#999999"),
                stack="produccion",
            )
        )

        # 4b. Serie de recurso remanente como línea
        cap = initial.get(tech, 0.0)
        if cap > 0:
            cum = 0.0
            remaining = []
            for a in años:
                cum += prod_by_year.get(a, 0.0)
                rem = max(0.0, cap - cum)
                remaining.append(round(rem * factor, 6))

            series.append(
                ChartSeries(
                    name=f"Recurso remanente {tech_display}",
                    data=remaining,
                    color=_REMAINING_COLORS.get(tech, "#666666"),
                    stack=None,
                    chart_type="line",
                )
            )

    return ChartDataResponse(
        categories=categories,
        series=series,
        title=title,
        yAxisLabel=un,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 3. build_chart_data — SINGLE ESCENARIO
# ═══════════════════════════════════════════════════════════════════════════


def get_available_fuels(
    db: Session,
    job_id: int,
    tipo: str,
    sub_filtro: str | None = None,
    loc: str | None = None,
    region: str | None = None,
) -> list[str]:
    """Retorna los códigos FUEL que sobreviven al pipeline de filtrado del chart.

    Útil para poblar el selector de combustible que reemplaza al de región
    cuando ``agrupar_por='REGION'``. Reusa ``_load_variable_data``,
    ``_apply_regional_transform`` y la filter function del config — cero
    lógica duplicada.
    """
    cfg = CONFIGS.get(tipo)
    if cfg is None:
        return []
    variable_name = cfg["variable_default"]
    df = _load_variable_data(db, job_id, variable_name)
    if df.empty:
        return []
    df = _apply_regional_transform(
        db, job_id, df, region_filter=region, agrupar_por="REGION"
    )
    filtro_fn = cfg.get("filtro")
    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)
    if "FUEL" not in df.columns:
        return []
    return sorted(df["FUEL"].dropna().unique().tolist())


def build_chart_data(
    db: Session,
    job_id: int,
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
    variable: str | None = None,
    agrupar_por: str | None = None,
    es_porcentaje_override: bool = False,
    region: str | None = None,
    timeslice: str | None = None,
    combustible: str | None = None,
) -> ChartDataResponse:
    """Construye la respuesta de gráfica para un solo escenario.

    Parámetros
    ----------
    db : Session
    job_id : int
    tipo : str
        Clave en ``CONFIGS`` (ej: ``'cap_electricidad'``, ``'gas_consumo'``).
    un : str
        Unidades de salida (PJ, GW, MW, TWh, Gpc).
    sub_filtro : str | None
        Filtro adicional dentro del sector.
    loc : str | None
        Localización (URB, RUR, ZNI).
    variable : str | None
        Override de variable para configs de capacidad.
    region : str | None
        Solo aplica si el job es REGIONAL. Filtra a una región específica
        (``'AN'..'SO'``). Ignorado cuando ``agrupar_por == 'REGION'``.
    combustible : str | None
        Solo aplica cuando ``agrupar_por == 'REGION'``. Filtra las filas
        por FUEL (p. ej. ``'NGS'``, ``'DSL'``). Se muestra en el título.
    """
    # ── Ruta especial: recursos vs demanda ────────────────────────────────
    if tipo == "recursos_vs_demanda":
        return build_recursos_vs_demanda_data(db, job_id, un=un)
    if tipo == "recursos_vs_demanda_gas":
        return build_recursos_vs_demanda_gas_data(db, job_id, un=un)
    if tipo == "recursos_vs_demanda_carbon":
        return build_recursos_vs_demanda_carbon_data(db, job_id, un=un)

    if tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no existe en CONFIGS.")

    cfg = CONFIGS[tipo]
    es_capacidad = cfg.get("es_capacidad", False)
    es_porcentaje = cfg.get("es_porcentaje", False) or es_porcentaje_override
    es_factor_planta = cfg.get("es_factor_planta", False)

    # Variable a consultar
    variable_name = variable if (variable and es_capacidad) else cfg["variable_default"]

    # ── Título ───────────────────────────────────────────────────────────
    if es_capacidad:
        titulo_var = TITULOS_VARIABLES_CAPACIDAD.get(variable_name, variable_name)
        title = f"{cfg['titulo_base']} — {titulo_var}"
    elif es_porcentaje or es_factor_planta:
        title = cfg.get("titulo_base", cfg.get("titulo", tipo))
    else:
        title = cfg.get("titulo", tipo)

    if sub_filtro:
        sub_label = NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)
        title += f" — {sub_label}"
    if loc:
        title += f" ({loc})"
    if timeslice:
        title += f" [TS={timeslice}]"
    if combustible:
        comb_label = NOMBRES_COMBUSTIBLES.get(combustible, combustible)
        title += f" — {comb_label}"
    if region and agrupar_por != "REGION":
        title += f" — Región {region}"

    es_emision = cfg.get("es_emision", False)
    es_emision_kt = cfg.get("es_emision_kt", False)

    if es_porcentaje or es_factor_planta:
        title += " (%)"
    elif es_emision:
        title += f" ({_emision_unit_label(un, es_emision_kt)})"
    else:
        title += f" ({un})"

    # ── Factor de Planta: pipeline propio ────────────────────────────────
    if es_factor_planta:
        return _build_factor_planta_data(db, job_id, tipo, cfg, title, sub_filtro, loc)

    # ── Cargar datos ─────────────────────────────────────────────────────
    df = _load_variable_data(db, job_id, variable_name)

    if df.empty:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=un,
        )

    # ── Transformación regional (jobs REGIONAL) ──────────────────────────
    # DEBE ir antes del filtro: los outputs REGIONAL llegan con prefijo
    # geográfico de 2 letras (p. ej. ``SE_PWRSOLUGE``) y los filtros de
    # ``CONFIGS`` asumen códigos sin prefijo (``startswith('PWR')``).
    df = _apply_regional_transform(
        db, job_id, df, region_filter=region, agrupar_por=agrupar_por
    )

    # Filtro por combustible (cuando agrupar_por='REGION')
    if combustible and "FUEL" in df.columns:
        df = df[df["FUEL"] == combustible]

    # ── Filtrar ──────────────────────────────────────────────────────────
    filtro_fn = cfg.get("filtro")
    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)

    # Filtro por timeslice (opcional): si el DataFrame tiene la columna
    # TIMESLICE y el caller pasa un código, restringimos antes del groupby.
    # Si no se pasa, se agrega por año (suma de todos los TS), que es el
    # comportamiento histórico.
    if timeslice and "TIMESLICE" in df.columns:
        df = df[df["TIMESLICE"].astype(str) == str(timeslice)]

    if df.empty:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=un,
        )

    # ── Alias de tecnologías (sector eléctrico) ──────────────────────────
    # Para los charts principales del sector eléctrico, consolidamos algunas
    # variantes bajo su tecnología "padre" — la leyenda queda más limpia y
    # la lectura del stack es directa (ver PWR_TECH_ALIASES en configs.py).
    if tipo in CONFIGS_CON_ALIAS_PWR:
        df = _aplicar_alias_pwr(df)

    # ── Conversión por tecnología (kton, etc.) ───────────────────────────
    # Debe ir antes del groupby porque necesita la columna TECHNOLOGY.
    tech_factor_map = cfg.get("unidad_factor_por_tecnologia")
    if tech_factor_map and un in tech_factor_map:
        df = _convertir_por_tecnologia(df, un, tech_factor_map)

    # ── Agrupación ───────────────────────────────────────────────────────
    agrupar_col = agrupar_por if agrupar_por is not None else cfg["agrupar_por"]

    if agrupar_col == "TECNOLOGIA":
        # Algunas configs piden separar las refinerías en (refinería × combustible)
        # mientras dejan el resto agrupado por tecnología (típicamente imports).
        if cfg.get("split_refineries_by_fuel") and "FUEL" in df.columns:
            df["COLOR"] = df.apply(
                lambda r: (
                    f"{r['TECHNOLOGY']}::{r['FUEL']}"
                    if str(r.get("TECHNOLOGY", "")).startswith("UPSREF")
                    and str(r.get("FUEL", "")).strip() != ""
                    else r["TECHNOLOGY"]
                ),
                axis=1,
            )
        else:
            df["COLOR"] = df["TECHNOLOGY"]
    elif agrupar_col == "GROUP":
        if "FUEL" in df.columns:
            df["COLOR"] = (df["TECHNOLOGY"] + "_" + df["FUEL"]).apply(asignar_grupo)
        else:
            df["COLOR"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "FUEL":
        if "FUEL" in df.columns:
            df["COLOR"] = df.apply(_fuel_to_group, axis=1)
        else:
            df["COLOR"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "SECTOR":
        df["COLOR"] = _sector_labels(df["TECHNOLOGY"])
    elif agrupar_col == "EMISION":
        df["COLOR"] = df["FUEL"] if "FUEL" in df.columns else "?"
    elif agrupar_col == "ELECTROLISIS":
        df["COLOR"] = df["TECHNOLOGY"].apply(_map_electrolisis_verde)
    elif agrupar_col == "H2_CONSUMO":
        df["COLOR"] = df["TECHNOLOGY"].apply(_map_h2_consumo_grupo)
    elif agrupar_col == "H2_PRODUCCION":
        df["COLOR"] = df["TECHNOLOGY"].apply(_map_h2_verde_azul_gris)
    elif agrupar_col == "TRANSPORTE_GRUPO":
        df["COLOR"] = df["TECHNOLOGY"].apply(_map_transporte_grupo)
    elif agrupar_col == "MODO":
        df["COLOR"] = df["TECHNOLOGY"].apply(_map_transporte_modo)
    elif agrupar_col == "REGION":
        # transform_regional_df ya añadió la columna REGION con prefijos AN..SO.
        df["COLOR"] = df["REGION"] if "REGION" in df.columns else df["TECHNOLOGY"]
    elif agrupar_col == "YEAR":
        # emisiones_total: solo agrupa por año
        df["COLOR"] = "Total"
    else:
        df["COLOR"] = df["TECHNOLOGY"]

    # ── Agregar ──────────────────────────────────────────────────────────
    df_agg = df.groupby(["COLOR", "YEAR"], as_index=False)["VALUE"].sum()

    # Descartar grupos insignificantes
    df_agg = df_agg[df_agg.groupby("COLOR")["VALUE"].transform("sum") > 1e-5]

    if df_agg.empty:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=un,
        )

    # ── Conversión de unidades ───────────────────────────────────────────
    # Si ya se aplicó conversión por tecnología (kton), no convertir otra vez.
    if tech_factor_map and un in tech_factor_map:
        pass
    elif es_emision:
        if not es_emision_kt:
            df_agg = _convertir_unidades_emision(df_agg, un)
        # es_emision_kt: base = kt, sin conversión
    else:
        df_agg = _convertir_unidades(df_agg, un)

    # ── Porcentaje (prd_electricidad) ────────────────────────────────────
    if es_porcentaje:
        total_por_año = df_agg.groupby("YEAR")["VALUE"].transform("sum")
        df_agg["VALUE"] = df_agg["VALUE"] / total_por_año * 100.0

    # ── Colores ──────────────────────────────────────────────────────────
    # Si agrupar_por fue overridden, ajustar color_fn según agrupación
    if agrupar_por is not None and agrupar_por != cfg.get("agrupar_por"):
        if agrupar_col in ("FUEL", "GROUP"):
            color_fn = _color_por_grupo_fijo
        elif agrupar_col == "SECTOR":
            color_fn = _color_por_sector
        elif agrupar_col == "EMISION":
            color_fn = _color_por_emision
        elif agrupar_col == "TRANSPORTE_GRUPO":
            color_fn = _color_transporte_grupo
        elif agrupar_col == "MODO":
            color_fn = _color_por_modo
        elif agrupar_col == "REGION":
            color_fn = _color_por_region
        elif agrupar_col == "ELECTROLISIS":
            color_fn = _color_electrolisis
        elif agrupar_col == "H2_CONSUMO":
            color_fn = _color_h2_consumo
        else:
            color_fn = (
                cfg.get("color_fn")
                if cfg.get("color_fn") == _color_electricidad
                else generar_colores_tecnologias
            )
    else:
        color_fn = cfg.get("color_fn")
    if color_fn is not None:
        colores_ordenados, orden_color = color_fn(df_agg, "COLOR")
    else:
        orden_color = sorted(df_agg["COLOR"].unique())
        _palette = get_colores_grupos()
        colores_ordenados = [_palette.get(c, "#999999") for c in orden_color]

    color_dict = dict(zip(orden_color, colores_ordenados))

    # ── Construir respuesta ──────────────────────────────────────────────
    años = [int(a) for a in sorted(df_agg["YEAR"].unique())]
    # Extender el eje X hasta el último año del horizonte de simulación. Los
    # años faltantes quedan en 0 (vía ``valor_por_año.get(a, 0.0)`` abajo).
    # Para porcentajes (mix eléctrico, etc.) extender no tiene sentido — el
    # 0 absoluto haría que la suma anual sea 0 y rompería el cálculo.
    if not es_porcentaje:
        años = _extend_years_to_sim_end(años, _get_simulation_max_year(db, job_id))
    categories = [str(a) for a in años]

    def _composite_label(code: str) -> str:
        # COLOR de la forma "UPSREF_XXX::FUEL" → "Refinería ... — Combustible"
        if agrupar_col == "REGION":
            return REGION_LABELS.get(code, code)
        if "::" in code:
            left, right = code.split("::", 1)
            return f"{get_label(left)} — {get_label(right)}"
        return get_label(code)

    from app.services.chart_series_config_service import (
        apply_global_series_config,
        normalize_agrupar_por,
    )

    stack_items = apply_global_series_config(
        db,
        tipo=tipo,
        agrupar_por=normalize_agrupar_por(agrupar_col, agrupar_col),
        orden_color=orden_color,
        color_dict=color_dict,
        default_name=_composite_label,
    )

    series: list[ChartSeries] = []
    for tech, series_color, series_name in stack_items:
        df_tech = df_agg[df_agg["COLOR"] == tech]
        valor_por_año = {
            int(row["YEAR"]): row["VALUE"] for _, row in df_tech.iterrows()
        }
        data = [round(valor_por_año.get(a, 0.0), 6) for a in años]
        series.append(
            ChartSeries(
                name=series_name,
                data=data,
                color=series_color,
                stack="default",
            )
        )

    return ChartDataResponse(
        categories=categories,
        series=series,
        title=title,
        yAxisLabel="%"
        if es_porcentaje
        else (_emision_unit_label(un, es_emision_kt) if es_emision else un),
    )


# ═══════════════════════════════════════════════════════════════════════════
# Helper: comparación recursos por año (by-year)
# ═══════════════════════════════════════════════════════════════════════════

_RECURSOS_TIPOS_COMPARACION = frozenset({
    "recursos_vs_demanda_gas",
    "recursos_vs_demanda_carbon",
})


def _build_comparison_recursos_by_year(
    db: Session,
    job_ids: list[int],
    tipo: str,
    years_to_plot: list[int] | None,
    scenario_names: dict[int, str],
    un: str = "PJ",
) -> CompareChartResponse:
    """Construye comparación by-year para gráficas de recursos (gas, carbón).

    Subplot por año clave, cada barra = producción total de un escenario.
    """
    if years_to_plot is None:
        years_to_plot = [2024, 2030, 2050]

    title = f"{_get_recursos_title(tipo)} — Comparación ({un})"

    prod_by_job_year: dict[int, dict[int, float]] = {}
    has_data = False
    for jid in job_ids:
        totals = _build_recursos_production_total(db, jid, tipo, un=un)
        if totals is None:
            prod_by_job_year[jid] = {}
            continue
        has_data = True
        prod_by_job_year[jid] = totals

    if not has_data:
        return CompareChartResponse(title=title, subplots=[], yAxisLabel=un)

    subplots: list[SubplotData] = []
    color = "#10b981" if tipo == "recursos_vs_demanda_gas" else "#dc2626"
    prod_label = {
        "recursos_vs_demanda_gas": "Producción gas natural nacional",
        "recursos_vs_demanda_carbon": "Demanda total de carbón",
    }.get(tipo, "Producción")

    for año in years_to_plot:
        data = [
            round(prod_by_job_year.get(jid, {}).get(año, 0.0), 6)
            for jid in job_ids
        ]
        categories = [scenario_names.get(jid, f"Job {jid}") for jid in job_ids]

        subplots.append(
            SubplotData(
                year=año,
                categories=categories,
                series=[
                    ChartSeries(
                        name=prod_label,
                        data=data,
                        color=color,
                        stack="default",
                    )
                ],
            )
        )

    return CompareChartResponse(
        title=title, subplots=subplots, yAxisLabel=un
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4. build_comparison_data — MULTI-ESCENARIO
# ═══════════════════════════════════════════════════════════════════════════
def build_comparison_data(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    years_to_plot: list[int] | None = None,
    agrupacion: str | None = None,
    sub_filtro: str | None = None,
    loc: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
    es_porcentaje_override: bool = False,
) -> CompareChartResponse:
    """Construye la respuesta de comparación multi-escenario.

    Genera subplots por año clave, con barras apiladas por categoría.

    Parámetros
    ----------
    db : Session
    job_ids : list[int]
        IDs de simulation_job a comparar (max 10).
    tipo : str
        Clave en ``CONFIGS_COMPARACION``.
    un : str
        Unidades.
    years_to_plot : list[int] | None
        Años clave (default: [2024, 2030, 2050]).
    agrupacion : str | None
        Override de agrupación (TECNOLOGIA, COMBUSTIBLE, SECTOR).
    sub_filtro, loc : str | None
        Filtros opcionales.
    """
    # MAPEO DE SECTORES A COMPARACIÓN (Si el usuario escoge la tabla normal, la mappeamos a su configuración de comparación)
    MAPEO_COMPARACION = {
        "tra_total": "tra_comparacion",
        "ind_total": "ind_comparacion",
        "res_total": "res_comparacion",
        "ter_total": "ter_comparacion",
    }

    es_generico = False
    if tipo in MAPEO_COMPARACION:
        tipo = MAPEO_COMPARACION[tipo]

    if tipo not in CONFIGS_COMPARACION:
        if tipo in CONFIGS:
            es_generico = True
            cfg = CONFIGS[tipo]
            es_emision = cfg.get("es_emision", False)
        else:
            raise ValueError(
                f"tipo='{tipo}' no existe ni en CONFIGS ni en CONFIGS_COMPARACION."
            )
    else:
        cfg = CONFIGS_COMPARACION[tipo]

    if years_to_plot is None:
        years_to_plot = [2024, 2030, 2050]

    # Resolver agrupación y año histórico
    if not es_generico:
        prefijo = cfg["prefijo"]
        agrupacion_fija = cfg.get("agrupacion_fija")
        if agrupacion_fija is not None:
            agrupacion_usar = agrupacion_fija
        elif agrupacion is not None:
            agrupacion_usar = agrupacion
        else:
            agrupacion_usar = cfg["agrupacion_default"]

        usa_historico = cfg["año_historico_unico"]
        año_historico = years_to_plot[0] if years_to_plot else 2024

        label_agrupacion = {
            "TECNOLOGIA": "Tecnologías",
            "COMBUSTIBLE": "Combustibles",
            "FUEL": "Combustibles",
            "SECTOR": "Sectores",
            "MODO": "Modo",
        }.get(agrupacion_usar, agrupacion_usar)
        title_base = f"{cfg['titulo_base']} por {label_agrupacion}"
    else:
        usa_historico = False
        año_historico = years_to_plot[0] if years_to_plot else 2024
        # agrupacion_usar = "TECNOLOGIA"  # Fallback a agrupar por tecnología para cualquier otra gráfica
        agrupacion_usar = (
            agrupacion
            if agrupacion is not None
            else cfg.get("agrupar_por", "TECNOLOGIA")
        )
        title_base = cfg.get("titulo", cfg.get("titulo_base", tipo)) + " (Comparación)"

    title = title_base
    if sub_filtro:
        sub_label = NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)
        title += f" — {sub_label}"
    if loc:
        title += f" ({loc})"
    title += f" ({un})"

    # ── Cargar nombres de escenarios ─────────────────────────────────────
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            from app.models import Scenario

            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"
        # Override por alias del reporte, si se proporcionó.
        ov = (job_display_overrides or {}).get(jid)
        if isinstance(ov, str) and ov.strip():
            scenario_names[jid] = ov.strip()

    variable_name = cfg["variable_default"]

    # ── Ruta especial: recursos vs demanda ────────────────────────────────
    if tipo in _RECURSOS_TIPOS_COMPARACION:
        return _build_comparison_recursos_by_year(
            db, job_ids, tipo, years_to_plot, scenario_names, un,
        )

    # ── Procesar datos ───────────────────────────────────────────────────
    all_data: list[pd.DataFrame] = []

    # Paso 1: Año histórico (solo del escenario primario — el primero en la lista)
    if usa_historico and año_historico in years_to_plot and job_ids:
        first_job_id = job_ids[0]
        df_var = _load_variable_data(db, first_job_id, variable_name)
        # Strip prefijos regionales si el job es REGIONAL (acumulado nacional).
        df_var = _apply_regional_transform(db, first_job_id, df_var)

        if not df_var.empty:
            df_hist = _procesar_bloque_comparacion(
                df_var,
                prefijo,
                sub_filtro,
                loc,
                agrupacion_usar,
                [año_historico],
                un,
            )
            if df_hist is not None and not df_hist.empty:
                df_hist["SCENARIO"] = scenario_names.get(
                    first_job_id, str(año_historico)
                )
                all_data.append(df_hist)

    # Paso 2: Años proyectados (todos los escenarios)
    años_a_procesar = (
        [y for y in years_to_plot if y != año_historico]
        if usa_historico
        else years_to_plot
    )

    for jid in job_ids:
        df_var = _load_variable_data(db, jid, variable_name)
        # Strip prefijos regionales si el job es REGIONAL (acumulado nacional).
        df_var = _apply_regional_transform(db, jid, df_var)
        if df_var.empty:
            continue

        if not es_generico:
            df = _procesar_bloque_comparacion(
                df_var,
                prefijo,
                sub_filtro,
                loc,
                agrupacion_usar,
                años_a_procesar,
                un,
            )
        else:
            df = _procesar_bloque_single(
                df_var,
                cfg,
                sub_filtro,
                loc,
                años_a_procesar,
                un,
                agrupacion_override=agrupacion_usar,
                tipo=tipo,
            )

        if df is None or df.empty:
            continue

        df["SCENARIO"] = scenario_names.get(jid, f"Job {jid}")
        all_data.append(df)

    if not all_data:
        return CompareChartResponse(title=title, subplots=[], yAxisLabel=un)

    df_final = pd.concat(all_data, ignore_index=True)

    # ── Porcentaje override ──────────────────────────────────────────────
    if es_porcentaje_override:
        total_por_año_escenario = df_final.groupby(["YEAR", "SCENARIO"])[
            "VALUE"
        ].transform("sum")
        df_final["VALUE"] = df_final["VALUE"] / total_por_año_escenario * 100.0

    # ── Colores ──────────────────────────────────────────────────
    categorias_unicas = sorted(df_final["CATEGORIA"].dropna().unique())
    if not es_generico:
        mapa_colores = _color_map_comparison(agrupacion_usar, categorias_unicas)
    else:
        # Para gráficas genéricas: usar color_fn según la agrupación REAL
        if agrupacion_usar != cfg.get("agrupar_por"):
            # Hubo override → usar función de color según agrupacion_usar
            if agrupacion_usar == "FUEL":
                color_fn = _color_por_grupo_fijo
            elif agrupacion_usar == "SECTOR":
                color_fn = _color_por_sector
            elif agrupacion_usar == "EMISION":
                color_fn = _color_por_emision
            elif agrupacion_usar == "ELECTROLISIS":
                color_fn = _color_electrolisis
            else:
                color_fn = cfg.get("color_fn") or generar_colores_tecnologias
        else:
            # Sin override → usar color_fn original del config
            color_fn = cfg.get("color_fn")

        if color_fn is not None:
            df_tmp = pd.DataFrame({"COLOR": list(categorias_unicas)})
            colores_lista, orden_lista = color_fn(df_tmp, "COLOR")
            mapa_colores = dict(zip(orden_lista, colores_lista))
        else:
            _palette = get_colores_grupos()
            mapa_colores = {c: _palette.get(c, "#999999") for c in categorias_unicas}

    from app.services.chart_series_config_service import (
        apply_global_series_config,
        normalize_agrupar_por,
    )

    agrup_key = normalize_agrupar_por(agrupacion_usar, agrupacion_usar)
    ordered_stack = apply_global_series_config(
        db,
        tipo=tipo,
        agrupar_por=agrup_key,
        orden_color=list(categorias_unicas),
        color_dict=mapa_colores,
        default_name=lambda c: get_label(str(c)),
    )

    # ── Construir subplots por año ───────────────────────────────────────
    años_ordenados = sorted(df_final["YEAR"].unique())
    subplots: list[SubplotData] = []

    for año in años_ordenados:
        df_año = df_final[df_final["YEAR"] == año]

        if año == years_to_plot[0]:
            # Año histórico: mostrar una sola barra (etiquetada con el nombre
            # del primer escenario para que los alias/display_name funcionen).
            if not usa_historico:
                historical_job_id = job_ids[0]
                historical_name = scenario_names.get(
                    historical_job_id, f"Job {historical_job_id}"
                )
                df_año = df_año[
                    df_año["SCENARIO"] == historical_name
                ].copy()
                df_año["SCENARIO"] = scenario_names.get(
                    job_ids[0], str(año)
                )
            escenarios_en_año = [
                scenario_names.get(job_ids[0], str(año))
            ]
        else:
            # Preservar el orden de selección del usuario (job_ids)
            escenarios_en_año = [
                scenario_names[jid]
                for jid in job_ids
                if jid in scenario_names
                and scenario_names[jid] in df_año["SCENARIO"].values
            ]

        series: list[ChartSeries] = []
        for categoria, col_cat, name_cat in ordered_stack:
            df_cat = df_año[df_año["CATEGORIA"] == categoria]
            if df_cat.empty:
                series.append(
                    ChartSeries(
                        name=name_cat,
                        data=[0.0] * len(escenarios_en_año),
                        color=col_cat,
                        stack="default",
                    )
                )
                continue

            valor_por_escenario = {
                row["SCENARIO"]: row["VALUE"]
                for _, row in df_cat.groupby("SCENARIO", as_index=False)["VALUE"]
                .sum()
                .iterrows()
            }
            data = [
                round(valor_por_escenario.get(esc, 0.0), 6) for esc in escenarios_en_año
            ]

            series.append(
                ChartSeries(
                    name=name_cat,
                    data=data,
                    color=col_cat,
                    stack="default",
                )
            )

        subplots.append(
            SubplotData(
                year=int(año),
                categories=escenarios_en_año,
                series=series,
            )
        )

    return CompareChartResponse(
        title=title, subplots=subplots, yAxisLabel="%" if es_porcentaje_override else un
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4b. build_comparison_facet_data — ESCENARIOS COMPLETOS (FACETS)
# ═══════════════════════════════════════════════════════════════════════════


def build_comparison_facet_data(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
    variable: str | None = None,
    agrupar_por: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
    es_porcentaje_override: bool = False,
    region: str | None = None,
    combustible: str | None = None,
) -> CompareChartFacetResponse:
    """Construye datos para comparación por escenarios completos (facets).

    Cada facet muestra la evolución temporal completa de un escenario.
    Usa CONFIGS (no CONFIGS_COMPARACION). Una query por job_id.

    Parámetros
    ----------
    db : Session
    job_ids : list[int]
        IDs de simulation_job a comparar (max 10).
    tipo : str
        Clave en CONFIGS (ej: 'cap_electricidad', 'gas_consumo').
    un : str
        Unidades de salida (PJ, GW, etc.).
    sub_filtro, loc, variable : str | None
        Filtros y override de variable.
    """
    if tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no encontrado en CONFIGS.")

    cfg = CONFIGS[tipo]
    title_base = cfg.get("titulo", cfg.get("titulo_base", tipo))
    title = title_base
    if sub_filtro:
        sub_label = NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)
        title += f" — {sub_label}"
    if loc:
        title += f" ({loc})"
    title += f" ({un})"

    facets: list[FacetData] = []
    y_label = un
    from app.models import Scenario

    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if not job:
            continue
        scenario = None
        if job.scenario_id is not None:
            scenario = db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
        scenario_name = scenario.name if scenario else (job.input_name or f"Job {jid}")
        tag_name = None
        if scenario is not None:
            from app.services.simulation_service import SimulationService as _SS

            primary = _SS._batch_scenario_tags_by_scenario_ids(
                db, {int(scenario.id)}
            ).get(int(scenario.id))
            if primary:
                tag_name = primary.get("name")
        override = (
            (job_display_overrides or {}).get(jid) if job_display_overrides else None
        )
        has_alias_override = isinstance(override, str) and bool(override.strip())
        job_display = (
            override.strip()
            if has_alias_override
            else (getattr(job, "display_name", None) or None)
        )
        # Cuando el alias reemplaza el nombre del escenario, no queremos
        # concatenar la etiqueta al subtítulo (quedaría "Alias — Tag").
        effective_tag_name = None if has_alias_override else tag_name
        effective_scenario_name = (
            override.strip() if has_alias_override else scenario_name
        )

        chart = build_chart_data(
            db=db,
            job_id=jid,
            tipo=tipo,
            un=un,
            sub_filtro=sub_filtro,
            loc=loc,
            variable=variable,
            agrupar_por=agrupar_por,
            es_porcentaje_override=es_porcentaje_override,
            region=region,
            combustible=combustible,
        )

        if not facets:
            y_label = chart.yAxisLabel
        facets.append(
            FacetData(
                scenario_name=effective_scenario_name,
                job_id=jid,
                display_name=job_display,
                scenario_tag_name=effective_tag_name,
                categories=chart.categories,
                series=chart.series,
            )
        )
    # Unifica el eje X entre todos los facets: si un escenario llega hasta
    # 2030 y otro hasta 2033, ambos paneles muestran 2022-2033 con 0 en los
    # años faltantes para el escenario corto. Mantener ejes idénticos hace
    # que las comparaciones visuales sean directas.
    _align_facet_x_axis(facets)
    # Forzar rango completo 2022-2055 para imp_oil (aunque no haya datos)
    if tipo == "imp_oil" and facets:
        full = [str(y) for y in range(2022, 2056)]
        if not all(c in set(facets[0].categories) for c in full):
            idx = {c: i for i, c in enumerate(full)}
            for f in facets:
                old_cats = set(f.categories)
                old_idx_map = {c: i for i, c in enumerate(f.categories)}
                for s in f.series:
                    s.data = [
                        s.data[old_idx_map[c]] if c in old_cats and old_idx_map[c] < len(s.data) else None
                        for c in full
                    ]
                f.categories = full
    return CompareChartFacetResponse(
        title=title,
        facets=facets,
        yAxisLabel=y_label,
    )


def _inject_exogenous_data_into_facets(
    facet_response: CompareChartFacetResponse,
    exogenous_data_json: str,
) -> CompareChartFacetResponse:
    """Inyecta datos exógenos (ej: Refinerías) como nueva serie en cada facet.

    Espeja la lógica de ``injectExogenousDataFacet`` en el frontend.
    Recibe el JSON tal cual lo envía el frontend (``ExogenousDataConfig``).
    """
    try:
        exo = json.loads(exogenous_data_json)
    except (json.JSONDecodeError, ValueError, TypeError):
        return facet_response

    if not isinstance(exo, dict):
        return facet_response
    if not exo.get("active"):
        return facet_response
    scenarios = exo.get("scenarios", [])
    if not scenarios:
        return facet_response

    category_label = exo.get("categoryLabel", "Refinerías")
    color = exo.get("color", "#808080")

    exo_by_job: dict[int, dict[int, float]] = {}
    for s in scenarios:
        jid = s.get("jobId")
        if jid is None:
            continue
        data_pairs = s.get("data", [])
        exo_by_job[int(jid)] = {int(k): float(v) for k, v in data_pairs if k is not None}

    new_facets: list[FacetData] = []
    for facet in facet_response.facets:
        exo_map = exo_by_job.get(facet.job_id)
        if not exo_map:
            new_facets.append(facet)
            continue
        ref_data: list[float | None] = []
        has_any = False
        for cat in facet.categories:
            try:
                year = int(cat)
            except (ValueError, TypeError):
                ref_data.append(None)
                continue
            val = exo_map.get(year)
            if val is not None:
                ref_data.append(val)
                has_any = True
            else:
                ref_data.append(None)
        if not has_any:
            new_facets.append(facet)
            continue
        new_facets.append(
            FacetData(
                scenario_name=facet.scenario_name,
                job_id=facet.job_id,
                display_name=facet.display_name,
                scenario_tag_name=facet.scenario_tag_name,
                categories=facet.categories,
                series=[
                    *facet.series,
                    ChartSeries(
                        name=category_label,
                        data=ref_data,
                        color=color,
                        stack="default",
                    ),
                ],
            )
        )

    return CompareChartFacetResponse(
        title=facet_response.title,
        facets=new_facets,
        yAxisLabel=facet_response.yAxisLabel,
    )


def _inject_exogenous_contaminantes_data(
    facet_response: CompareChartFacetResponse,
    exogenous_json: str,
) -> CompareChartFacetResponse:
    """Inyecta datos exógenos de contaminantes (BC, CO, etc.) sumándolos a las
    series existentes en cada facet.

    Espeja la lógica de ``injectContaminantesExogenousFacet`` en el frontend.
    """
    try:
        exo = json.loads(exogenous_json)
    except (json.JSONDecodeError, ValueError, TypeError):
        return facet_response

    if not isinstance(exo, dict):
        return facet_response
    if not exo.get("active"):
        return facet_response
    scenarios = exo.get("scenarios", [])
    if not scenarios:
        return facet_response

    # Build exo_by_job: {job_id: {year: {pollutant_key: value}}}
    exo_by_job: dict[int, dict[int, dict[str, float]]] = {}
    for s in scenarios:
        jid = s.get("jobId")
        if jid is None:
            continue
        year_dict: dict[int, dict[str, float]] = {}
        raw_data = s.get("data", {})
        if not isinstance(raw_data, dict):
            continue
        for pollutant_key, pairs in raw_data.items():
            if not isinstance(pairs, list):
                continue
            for pair in pairs:
                if not isinstance(pair, (list, tuple)) or len(pair) < 2:
                    continue
                year = pair[0]
                val = pair[1]
                if year is None or val is None:
                    continue
                y = int(year)
                v = float(val)
                if y not in year_dict:
                    year_dict[y] = {}
                year_dict[y][str(pollutant_key)] = v
        exo_by_job[int(jid)] = year_dict

    def _strip_emi(name: str) -> str:
        return name[3:] if name.startswith("EMI") else name

    new_facets: list[FacetData] = []
    for facet in facet_response.facets:
        job_exo = exo_by_job.get(facet.job_id)
        if not job_exo:
            new_facets.append(facet)
            continue

        changed = False
        new_series: list[ChartSeries] = []
        for series in facet.series:
            key = _strip_emi(series.name)
            new_data: list[float | None] = []
            for cat_idx, cat in enumerate(facet.categories):
                try:
                    year = int(cat)
                except (ValueError, TypeError):
                    new_data.append(series.data[cat_idx] if cat_idx < len(series.data) else None)
                    continue
                val = series.data[cat_idx] if cat_idx < len(series.data) else None
                if val is None:
                    new_data.append(None)
                    continue
                year_exo = job_exo.get(year)
                if year_exo is not None and key in year_exo:
                    summed = val + year_exo[key]
                    new_data.append(summed)
                    changed = True
                else:
                    new_data.append(val)
            new_series.append(ChartSeries(
                name=series.name,
                data=new_data,
                color=series.color,
                stack=series.stack,
            ))
        if not changed:
            new_facets.append(facet)
        else:
            new_facets.append(FacetData(
                scenario_name=facet.scenario_name,
                job_id=facet.job_id,
                display_name=facet.display_name,
                scenario_tag_name=facet.scenario_tag_name,
                categories=facet.categories,
                series=new_series,
            ))

    return CompareChartFacetResponse(
        title=facet_response.title,
        facets=new_facets,
        yAxisLabel=facet_response.yAxisLabel,
    )


def _align_facet_x_axis(facets: list[FacetData]) -> None:
    """Unifica el eje X entre todos los facets, in-place.

    1. Construye la unión ordenada de todas las categorías (ordena como int
       si todas son parseables como año, sino orden lexicográfico).
    2. Para cada facet, reescribe ``categories`` con la unión y rellena
       ``series.data`` con **None** en las posiciones que ese escenario no
       tenía.

    ``None`` (serializado como ``null`` en JSON) es importante porque:
      - **Líneas**: ``null`` crea un *gap* (Highcharts/matplotlib no traza
        el punto) en vez de hacer caer la línea a 0.
      - **Barras apiladas**: ``null`` no se dibuja como barra ni contribuye
        al total apilado — los totales (``StackItemObject.total``) no se
        contaminan.

    Si solo hay 0 o 1 facet, o todos comparten exactamente el mismo eje, no
    hace nada.
    """
    if not facets or len(facets) < 2:
        return
    # Cortocircuito: si todos los facets ya comparten exactamente las mismas
    # categorías en el mismo orden, no hacemos nada.
    first_cats = list(facets[0].categories)
    if all(list(f.categories) == first_cats for f in facets):
        return

    # 1) Construir la unión.
    union_set: set[str] = set()
    for f in facets:
        for c in f.categories:
            union_set.add(str(c))
    # ¿Todas las categorías son parseables como entero (año)?
    try:
        as_ints = sorted({int(c) for c in union_set})
        union: list[str] = [str(y) for y in as_ints]
    except ValueError:
        union = sorted(union_set)

    union_index: dict[str, int] = {c: i for i, c in enumerate(union)}
    n = len(union)

    # 2) Reescribir cada facet con la unión. Posiciones faltantes → None.
    for f in facets:
        old_cats = [str(c) for c in f.categories]
        old_index_in_union = [union_index.get(c) for c in old_cats]
        for s in f.series:
            new_data: list[float | None] = [None] * n
            for i, target in enumerate(old_index_in_union):
                if target is None or i >= len(s.data):
                    continue
                v = s.data[i]
                if v is None:
                    new_data[target] = None
                    continue
                try:
                    fv = float(v)
                    # NaN/Infinity también se representan como None.
                    import math as _math

                    new_data[target] = None if not _math.isfinite(fv) else fv
                except (TypeError, ValueError):
                    new_data[target] = None
            s.data = new_data
        f.categories = list(union)


def _procesar_bloque_comparacion(
    df_var: pd.DataFrame,
    prefijo: str | tuple[str, ...],
    sub_filtro: str | None,
    loc: str | None,
    agrupacion: str,
    años: list[int],
    un: str,
    es_emision: bool = False,
    tech_factor_map: dict | None = None,
) -> pd.DataFrame | None:
    """Pipeline para un bloque de datos de comparación.

    filtrar → filtrar años → asignar categorías → agregar → convertir.
    Port de ``graficas_comparacion._procesar_bloque``.
    """
    if df_var is None or df_var.empty:
        return None

    if "TECHNOLOGY" not in df_var.columns or "YEAR" not in df_var.columns:
        return None

    df = _filtrar_df(df_var, prefijo, sub_filtro, loc)
    if df.empty:
        return None

    # Conversión por tecnología (kton, etc.) — antes del groupby.
    if tech_factor_map and un in tech_factor_map:
        df = _convertir_por_tecnologia(df, un, tech_factor_map)

    df = df[df["YEAR"].isin(años)]
    if df.empty:
        return None

    df = _asignar_categoria(df, agrupacion)

    df = df.groupby(["CATEGORIA", "YEAR"], as_index=False)["VALUE"].sum()

    # Descartar categorías insignificantes
    df = df[df.groupby("CATEGORIA")["VALUE"].transform("sum") > 1e-5]
    if df.empty:
        return None

    # df = _convertir_unidades(df, un)

    if tech_factor_map and un in tech_factor_map:
        pass
    elif not es_emision:
        df = _convertir_unidades(df, un)

    return df


def _procesar_bloque_single(
    df_var: pd.DataFrame,
    cfg: dict,
    sub_filtro: str | None,
    loc: str | None,
    años: list[int],
    un: str,
    agrupacion_override: str | None = None,
    tipo: str | None = None,
) -> pd.DataFrame | None:
    """Procesador genérico que emula la agrupación de build_chart_data para comparación."""
    if df_var is None or df_var.empty:
        return None

    if "TECHNOLOGY" not in df_var.columns or "YEAR" not in df_var.columns:
        return None

    df = df_var.copy()

    filtro_fn = cfg.get("filtro")
    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)

    if df.empty:
        return None

    df = df[df["YEAR"].isin(años)]
    if df.empty:
        return None

    # Alias de tecnologías del sector eléctrico (mismas reglas que build_chart_data).
    if tipo and tipo in CONFIGS_CON_ALIAS_PWR:
        df = _aplicar_alias_pwr(df)

    # Conversión por tecnología (kton, etc.) — antes del groupby.
    tech_factor_map = cfg.get("unidad_factor_por_tecnologia")
    if tech_factor_map and un in tech_factor_map:
        df = _convertir_por_tecnologia(df, un, tech_factor_map)

    agrupar_col = agrupacion_override if agrupacion_override is not None else cfg["agrupar_por"]

    if agrupar_col == "TECNOLOGIA":
        df["CATEGORIA"] = df["TECHNOLOGY"]
    elif agrupar_col == "GROUP":
        if "FUEL" in df.columns:
            df["CATEGORIA"] = (df["TECHNOLOGY"] + "_" + df["FUEL"]).apply(asignar_grupo)
        else:
            df["CATEGORIA"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "FUEL":
        if "FUEL" in df.columns:
            df["CATEGORIA"] = df.apply(_fuel_to_group, axis=1)
        else:
            df["CATEGORIA"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "SECTOR":
        df["CATEGORIA"] = _sector_labels(df["TECHNOLOGY"])
    elif agrupar_col == "EMISION":
        df["CATEGORIA"] = df["FUEL"] if "FUEL" in df.columns else "?"
    elif agrupar_col == "TRANSPORTE_GRUPO":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_transporte_grupo)
    elif agrupar_col == "MODO":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_transporte_modo)
    elif agrupar_col == "ELECTROLISIS":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_electrolisis_verde)
    elif agrupar_col == "H2_CONSUMO":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_h2_consumo_grupo)
    elif agrupar_col == "H2_PRODUCCION":
        df["CATEGORIA"] = df["TECHNOLOGY"].apply(_map_h2_verde_azul_gris)
    elif agrupar_col == "YEAR":
        df["CATEGORIA"] = "Total"
    else:
        df["CATEGORIA"] = df["TECHNOLOGY"]

    df = df.groupby(["CATEGORIA", "YEAR"], as_index=False)["VALUE"].sum()
    df = df[df.groupby("CATEGORIA")["VALUE"].transform("sum") > 1e-5]

    if df.empty:
        return None

    es_emision = cfg.get("es_emision", False)
    es_emision_kt = cfg.get("es_emision_kt", False)
    if tech_factor_map and un in tech_factor_map:
        pass
    elif es_emision:
        if not es_emision_kt:
            df = _convertir_unidades_emision(df, un)
    else:
        df = _convertir_unidades(df, un)

    return df


# ═══════════════════════════════════════════════════════════════════════════
# 4c. build_comparison_line_data — LÍNEAS MULTI-ESCENARIO CONSOLIDADAS
# ═══════════════════════════════════════════════════════════════════════════

_SCENARIO_LINE_COLORS = [
    "#3b82f6",
    "#f59e0b",
    "#10b981",
    "#ef4444",
    "#8b5cf6",
    "#06b6d4",
    "#f97316",
    "#84cc16",
    "#ec4899",
    "#6366f1",
]


def build_comparison_line_data(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
) -> ChartDataResponse:
    """Construye líneas totales multi-escenario sobre el mismo eje.

    Todos los escenarios se trazan en la misma figura (sin subplots).
    X = años, Y = total agregado (suma de todas las tecnologías), una línea por escenario.

    Parámetros
    ----------
    db : Session
    job_ids : list[int]
        IDs de simulation_job a comparar (max 10).
    tipo : str
        Clave en CONFIGS_COMPARACION o CONFIGS.
    un : str
        Unidades de salida (PJ, GW, etc.).
    sub_filtro, loc : str | None
        Filtros opcionales de tecnología/localización.
    """
    MAPEO_COMPARACION = {
        "tra_total": "tra_comparacion",
        "ind_total": "ind_comparacion",
        "res_total": "res_comparacion",
        "ter_total": "ter_comparacion",
    }
    if tipo in MAPEO_COMPARACION:
        tipo = MAPEO_COMPARACION[tipo]

    if tipo in CONFIGS_COMPARACION:
        cfg = CONFIGS_COMPARACION[tipo]
        variable_name: str = cfg["variable_default"]
        prefijo = cfg["prefijo"]
        es_emision = False
        es_emision_kt = False
        title_base = cfg["titulo_base"]

        def _apply_filter(df: pd.DataFrame) -> pd.DataFrame:
            return _filtrar_df(df, prefijo, sub_filtro, loc)

    elif tipo in CONFIGS:
        cfg = CONFIGS[tipo]
        variable_name = cfg["variable_default"]
        es_emision = cfg.get("es_emision", False)
        es_emision_kt = cfg.get("es_emision_kt", False)
        title_base = cfg.get("titulo", cfg.get("titulo_base", tipo))
        filtro_fn = cfg.get("filtro")

        def _apply_filter(df: pd.DataFrame) -> pd.DataFrame:
            if filtro_fn is not None:
                return filtro_fn(df, sub_filtro=sub_filtro, loc=loc)
            return df
    else:
        raise ValueError(
            f"tipo='{tipo}' no existe en CONFIGS ni en CONFIGS_COMPARACION."
        )

    title = title_base
    if sub_filtro:
        title += f" — {NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)}"
    if loc:
        title += f" ({loc})"
    title += f" — Total ({un})"

    # Cargar nombres de escenarios
    from app.models import Scenario
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"
        ov = (job_display_overrides or {}).get(jid)
        if isinstance(ov, str) and ov.strip():
            scenario_names[jid] = ov.strip()

    # Agregar total por año para cada job
    all_years: set[int] = set()
    totals_per_job: dict[int, dict[int, float]] = {}

    for jid in job_ids:
        df = _load_variable_data(db, jid, variable_name)
        # Strip prefijos regionales si el job es REGIONAL (acumulado nacional).
        df = _apply_regional_transform(db, jid, df)
        if df.empty:
            totals_per_job[jid] = {}
            continue
        df = _apply_filter(df)
        if df.empty:
            totals_per_job[jid] = {}
            continue
        if not es_emision:
            df = _convertir_unidades(df, un)
        year_totals = df.groupby("YEAR")["VALUE"].sum()
        totals_per_job[jid] = {
            int(y): round(float(v), 6) for y, v in year_totals.items()
        }
        all_years.update(totals_per_job[jid].keys())

    if not all_years:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=_emision_unit_label(un, es_emision_kt) if es_emision else un,
        )

    years_sorted = sorted(all_years)
    # Extender el eje X hasta el mayor "último año de simulación" entre los
    # escenarios comparados, para que la línea no se corte antes de tiempo.
    sim_max_years = [
        m for m in (_get_simulation_max_year(db, jid) for jid in job_ids) if m is not None
    ]
    if sim_max_years:
        years_sorted = _extend_years_to_sim_end(years_sorted, max(sim_max_years))
    categories = [str(y) for y in years_sorted]

    series: list[ChartSeries] = []
    for idx, jid in enumerate(job_ids):
        year_data = totals_per_job.get(jid, {})
        if not year_data:
            continue
        data = [year_data.get(y, 0.0) for y in years_sorted]
        series.append(
            ChartSeries(
                name=scenario_names.get(jid, f"Job {jid}"),
                data=data,
                color=_SCENARIO_LINE_COLORS[idx % len(_SCENARIO_LINE_COLORS)],
            )
        )

    es_emision_kt_line = (
        CONFIGS.get(tipo, {}).get("es_emision_kt", False) if tipo in CONFIGS else False
    )
    y_label = _emision_unit_label(un, es_emision_kt_line) if es_emision else un
    return ChartDataResponse(
        categories=categories, series=series, title=title, yAxisLabel=y_label
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4d. build_pareto_data — PARETO POR TECNOLOGÍA
# ═══════════════════════════════════════════════════════════════════════════


def build_pareto_data(
    db: Session,
    job_id: int,
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
) -> ParetoChartResponse:
    """Construye los datos para un gráfico de Pareto por tecnología.

    Agrega valores de todas las tecnologías sumando sobre todos los años,
    ordena de mayor a menor y calcula el porcentaje acumulado.

    Parámetros
    ----------
    db : Session
    job_id : int
    tipo : str
        Clave en CONFIGS con agrupar_por='TECNOLOGIA' y no es_capacidad.
    un : str
        Unidades de salida.
    sub_filtro, loc : str | None
        Filtros opcionales.
    """
    if tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no encontrado en CONFIGS.")

    cfg = CONFIGS[tipo]
    es_capacidad = cfg.get("es_capacidad", False)
    es_emision = cfg.get("es_emision", False)
    es_emision_kt = cfg.get("es_emision_kt", False)
    variable_name: str = cfg["variable_default"]
    filtro_fn = cfg.get("filtro")
    title_base = cfg.get("titulo", cfg.get("titulo_base", tipo))

    _emi_label = _emision_unit_label(un, es_emision_kt)
    title = f"{title_base} — Pareto por Tecnología ({_emi_label if es_emision else un})"
    if sub_filtro:
        title += f" [{NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)}]"

    df = _load_variable_data(db, job_id, variable_name)
    if df.empty:
        return ParetoChartResponse(
            categories=[],
            values=[],
            cumulative_percent=[],
            title=title,
            yAxisLabel=_emi_label if es_emision else un,
        )

    # Strip prefijos regionales antes del filtro (jobs REGIONAL).
    df = _apply_regional_transform(db, job_id, df)

    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)
    if df.empty:
        return ParetoChartResponse(
            categories=[],
            values=[],
            cumulative_percent=[],
            title=title,
            yAxisLabel=_emi_label if es_emision else un,
        )

    if not es_emision and not es_capacidad:
        df = _convertir_unidades(df, un)

    # Agregar por tecnología (suma sobre todos los años)
    tech_totals = df.groupby("TECHNOLOGY")["VALUE"].sum().reset_index()
    tech_totals = tech_totals[tech_totals["VALUE"] > 1e-5]
    tech_totals = tech_totals.sort_values("VALUE", ascending=False).reset_index(
        drop=True
    )

    if tech_totals.empty:
        return ParetoChartResponse(
            categories=[],
            values=[],
            cumulative_percent=[],
            title=title,
            yAxisLabel="MtCO₂eq" if es_emision else un,
        )

    total = tech_totals["VALUE"].sum()
    tech_totals["CUMSUM"] = tech_totals["VALUE"].cumsum()
    tech_totals["CUM_PCT"] = (tech_totals["CUMSUM"] / total * 100).round(2)

    categories = [get_label(str(t)) for t in tech_totals["TECHNOLOGY"]]
    values = [round(float(v), 6) for v in tech_totals["VALUE"]]
    cumulative_percent = [float(p) for p in tech_totals["CUM_PCT"]]

    y_label = _emi_label if es_emision else un
    return ParetoChartResponse(
        categories=categories,
        values=values,
        cumulative_percent=cumulative_percent,
        title=title,
        yAxisLabel=y_label,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 5. get_result_summary — KPIs
# ═══════════════════════════════════════════════════════════════════════════


def get_result_summary(
    db: Session,
    job_id: int,
    current_user_id=None,
) -> ResultSummaryResponse:
    """Devuelve resumen de KPIs para el header de visualización.

    Si se provee ``current_user_id`` (UUID), se completan los flags
    ``is_favorite`` para ese usuario; ``is_public``/``is_infeasible_result``
    se derivan del job mismo.
    """
    from app.models import Scenario, SimulationJobFavorite, User
    from app.services.simulation_service import SimulationService

    job = db.query(SimulationJob).filter(SimulationJob.id == job_id).first()

    if not job:
        raise ValueError(f"Job {job_id} no encontrado.")

    scenario = None
    if job.scenario_id is not None:
        scenario = db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
    scenario_name = scenario.name if scenario else job.input_name
    scenario_tag = None
    scenario_tags_list: list[ScenarioTagPublic] = []
    if scenario is not None:
        all_tags = SimulationService._batch_all_scenario_tags_by_scenario_ids(
            db, {int(scenario.id)}
        ).get(int(scenario.id), [])
        scenario_tags_list = [ScenarioTagPublic.model_validate(t) for t in all_tags]
        if scenario_tags_list:
            scenario_tag = scenario_tags_list[0]

    solver_status = (job.model_timings_json or {}).get("solver_status", "unknown")

    total_co2 = (
        db.query(func.coalesce(func.sum(OsemosysOutputParamValue.value), 0))
        .filter(
            OsemosysOutputParamValue.id_simulation_job == job_id,
            OsemosysOutputParamValue.variable_name == "AnnualEmissions",
        )
        .scalar()
    ) or 0.0

    is_favorite = False
    if current_user_id is not None:
        is_favorite = (
            db.query(SimulationJobFavorite)
            .filter(
                SimulationJobFavorite.user_id == current_user_id,
                SimulationJobFavorite.job_id == job_id,
            )
            .first()
            is not None
        )

    owner = (
        db.query(User.username).filter(User.id == job.user_id).scalar()
        if job.user_id
        else None
    )

    return ResultSummaryResponse(
        job_id=job.id,
        scenario_id=job.scenario_id,
        scenario_name=scenario_name,
        scenario_description=(scenario.description if scenario else None),
        scenario_tag=scenario_tag,
        scenario_tags=scenario_tags_list,
        display_name=getattr(job, "display_name", None) or None,
        solver_name=job.solver_name,
        solver_status=solver_status,
        objective_value=job.objective_value or 0.0,
        coverage_ratio=job.coverage_ratio or 0.0,
        reserve_margin_dual=job.reserve_margin_dual,
        total_demand=job.total_demand or 0.0,
        total_dispatch=job.total_dispatch or 0.0,
        total_unmet=job.total_unmet or 0.0,
        total_co2=float(total_co2),
        is_public=bool(getattr(job, "is_public", True)),
        is_favorite=bool(is_favorite),
        is_infeasible_result=SimulationService._is_infeasible_succeeded_job(job),
        owner_username=owner,
        model_defaults_version_id=getattr(job, "model_defaults_version_id", None),
    )


# ═══════════════════════════════════════════════════════════════════════════
# 6. get_chart_catalog — CATÁLOGO DE GRÁFICAS
# ═══════════════════════════════════════════════════════════════════════════


def get_chart_catalog() -> list[ChartCatalogItem]:
    """Devuelve la lista de gráficas disponibles para el selector del frontend."""
    from app.schemas.visualization import DataExplorerFilters
    from app.visualization.data_explorer_filters import get_data_explorer_filters

    items: list[ChartCatalogItem] = []

    for config_id, cfg in CONFIGS.items():
        label = cfg.get("titulo", cfg.get("titulo_base", config_id))
        de_raw = get_data_explorer_filters(config_id, cfg.get("variable_default"))
        de_filters = (
            DataExplorerFilters(**{k: v for k, v in de_raw.items() if v})
            if de_raw
            else None
        )
        items.append(
            ChartCatalogItem(
                id=config_id,
                label=label,
                variable_default=cfg["variable_default"],
                has_sub_filtro=_config_has_sub_filtro(cfg),
                has_loc=_config_has_loc(cfg),
                sub_filtros=_config_sub_filtros(cfg),
                es_capacidad=cfg.get("es_capacidad", False),
                soporta_pareto=_config_soporta_pareto(cfg),
                data_explorer_filters=de_filters,
            )
        )

    return items


def _config_has_sub_filtro(cfg: dict) -> bool:
    """Determina si un config single-scenario soporta sub_filtro.

    Los configs de residencial, industrial, transporte y terciario
    aceptan sub_filtro a través de su función filtro.
    """
    filtro = cfg.get("filtro")
    if filtro is None:
        return False
    # Funciones que soportan sub_filtro por su signature
    filtro_name = getattr(filtro, "__name__", "")
    return filtro_name in (
        "_filtro_residencial",
        "_filtro_industrial",
        "_filtro_transporte",
        "_filtro_terciario",
        "_filtro_prefijo_con_sub",
        "_filtro_construccion",
        "_filtro_agroforestal",
        "_filtro_mineria",
        "_filtro_coquerias",
    )


def _config_has_loc(cfg: dict) -> bool:
    """Determina si un config single-scenario soporta localización."""
    filtro = cfg.get("filtro")
    if filtro is None:
        return False
    filtro_name = getattr(filtro, "__name__", "")
    return filtro_name == "_filtro_residencial"


def _config_soporta_pareto(cfg: dict) -> bool:
    """Pareto disponible para configs con agrupación TECNOLOGIA, no capacidad, no ratios."""
    return (
        cfg.get("agrupar_por") == "TECNOLOGIA"
        and not cfg.get("es_capacidad", False)
        and not cfg.get("es_factor_planta", False)
        and not cfg.get("es_porcentaje", False)
        and cfg.get("variable_default") not in ("AnnualEmissions",)
    )


def _config_sub_filtros(cfg: dict) -> list[str] | None:
    """Devuelve la lista de sub_filtros conocidos para un config, o None."""
    filtro = cfg.get("filtro")
    if filtro is None:
        return None
    filtro_name = getattr(filtro, "__name__", "")
    if filtro_name == "_filtro_residencial":
        return ["CKN", "WHT", "AIR", "REF", "ILU", "TV", "OTH"]
    if filtro_name == "_filtro_industrial":
        return ["BOI", "FUR", "MPW", "AIR", "REF", "ILU", "OTH"]
    if filtro_name == "_filtro_transporte":
        return [
            "CARRETERA",
            "AVI",
            "BOT",
            "SHP",
            "LDV",
            "FWD",
            "BUS",
            "TCK_C2P",
            "TCK_CSG",
            "MOT",
            "MIC",
            "TAX",
            "STT",
            "MET",
        ]
    if filtro_name == "_filtro_terciario":
        return ["ACL", "AIR", "CKN", "DAT", "FAN", "ILU", "MPW", "REF", "OTH"]
    return None


# ═══════════════════════════════════════════════════════════════════════════
# 7. EXPORT ALL — ZIP con gráficas como imágenes
# ═══════════════════════════════════════════════════════════════════════════


def export_all_charts_zip(
    db: Session,
    job_id: int,
    un: str = "PJ",
    fmt: str = "svg",
    *,
    view_mode: str = "column",
    clean: bool = False,
) -> "io.BytesIO":
    """Genera un ZIP con todas las gráficas renderizadas como SVG o PNG.

    Para configs de capacidad genera 3 figuras (Total, Nueva, Acumulada).
    Retorna un BytesIO listo para streaming.

    ``view_mode``: ``"column"`` (default), ``"line"`` o ``"area"``.

    Si ``clean=True`` las gráficas se renderizan sin título ni etiquetas
    numéricas sobre las barras.
    """
    import io
    import zipfile

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    CAPACITY_VARIABLES = [
        ("TotalCapacityAnnual", "Cap_Total"),
        ("NewCapacity", "Cap_Nueva"),
        ("AccumulatedNewCapacity", "Cap_Acumulada"),
    ]

    ext = "svg" if fmt == "svg" else "png"
    output = io.BytesIO()

    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        file_count = 0

        for config_id, cfg in CONFIGS.items():
            es_capacidad = cfg.get("es_capacidad", False)
            label = cfg.get("titulo", cfg.get("titulo_base", config_id))

            charts_to_render: list[tuple[str, ChartDataResponse]] = []

            if es_capacidad:
                for var_name, var_suffix in CAPACITY_VARIABLES:
                    chart = build_chart_data(
                        db,
                        job_id,
                        config_id,
                        un=un,
                        variable=var_name,
                    )
                    if chart.series:
                        charts_to_render.append((f"{label} — {var_suffix}", chart))
            else:
                chart = build_chart_data(db, job_id, config_id, un=un)
                if chart.series:
                    charts_to_render.append((label, chart))

            for chart_label, chart_data in charts_to_render:
                if view_mode == "line":
                    img_buf = _render_line_chart(
                        chart_data, chart_label,
                        fmt=ext, clean=clean,
                    )
                elif view_mode == "area":
                    img_buf = _render_stacked_area(
                        chart_data, chart_label,
                        fmt=ext, clean=clean,
                    )
                else:
                    img_buf = _render_stacked_bar(
                        chart_data, chart_label,
                        fmt=ext, clean=clean,
                    )
                safe_name = _safe_filename(chart_label)
                if clean:
                    safe_name += "_clean"
                zf.writestr(f"{safe_name}.{ext}", img_buf.getvalue())
                file_count += 1

    output.seek(0)
    return output


def _legend_ncols_for_labels(labels: list[str], hard_cap: int = 5) -> int:
    """Cap de columnas de leyenda según el largo máximo de las etiquetas.

    Evita que leyendas con etiquetas muy largas se expandan más allá del ancho
    de la figura — lo que con ``bbox_inches="tight"`` haría que matplotlib
    estire la imagen final horizontalmente, dejando el panel pequeño y con
    huecos blancos a los lados.
    """
    n = len(labels)
    if n == 0:
        return 1
    max_len = max(len(lab) for lab in labels)
    if max_len > 48:
        cap = 2
    elif max_len > 34:
        cap = 3
    elif max_len > 22:
        cap = 4
    else:
        cap = hard_cap
    return max(1, min(cap, hard_cap, n))


def _is_line_series(s: Any) -> bool:
    """True si la serie tiene ``chart_type == 'line'``."""
    return bool(getattr(s, "chart_type", None) == "line")


def _filter_hidden_series(
    series_list: list[Any],
    hidden_names: set[str] | None,
) -> list[Any]:
    """Filtra series cuyo nombre está en ``hidden_names``."""
    if not hidden_names:
        return series_list
    return [s for s in series_list if s.name not in hidden_names]


def _render_stacked_bar(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como gráfica de barras apiladas con matplotlib.

    Soporta **series mixtas**: las series con ``chart_type='line'`` se dibujan
    como líneas sobre las barras en lugar de apilarse.
    """
    import io
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib.lines import Line2D as _Line2D

    categories = chart.categories
    n_cats = len(categories)
    x = np.arange(n_cats)

    fig, ax = plt.subplots(figsize=(max(12, n_cats * 0.5), 7))

    # Separar series en barras y líneas (respetando chart_type)
    all_series = _filter_hidden_series(list(chart.series), hidden_series)
    bar_series = [s for s in all_series if not _is_line_series(s)]
    line_series = [s for s in all_series if _is_line_series(s)]

    # ── Barras apiladas ────────────────────────────────────────────────
    bottom = np.zeros(n_cats)
    for s in reversed(bar_series):
        raw = np.array(s.data, dtype=float)
        values = np.nan_to_num(raw, nan=0.0, posinf=0.0, neginf=0.0)
        ax.bar(x, values, bottom=bottom, color=s.color, width=0.7)
        bottom += values

    # ── Líneas sobre las barras ────────────────────────────────────────
    for s in line_series:
        raw = np.array(s.data, dtype=float)
        values = np.nan_to_num(raw, nan=float("nan"), posinf=float("nan"), neginf=float("nan"))
        ax.plot(x, values, color=s.color, linewidth=2.5, marker="o", markersize=4)

    # ── Handles de leyenda mixtos ──────────────────────────────────────
    legend_handles: list[_Line2D] = []
    legend_labels: list[str] = []

    # Barras: círculos (mismo orden que la leyenda de facets — primera
    # serie del array = primer ítem de leyenda)
    for s in bar_series:
        legend_handles.append(
            _Line2D(
                [0], [0],
                marker="o", color=s.color, linestyle="None",
                markersize=10, markerfacecolor=s.color, markeredgecolor=s.color,
            )
        )
        legend_labels.append(s.name)
    # Líneas: segmento de línea con marcador
    for s in line_series:
        legend_handles.append(
            _Line2D(
                [0], [0],
                color=s.color, linewidth=2.5, marker="o", markersize=5,
                markeredgecolor=s.color, markerfacecolor=s.color,
            )
        )
        legend_labels.append(s.name)

    # Stack totals on top — 1 decimal máx, cada 2 categorías (0, 2, 4, …).
    if not clean:
        for i, total in enumerate(bottom):
            if i % 2 != 0:
                continue
            if total > 0:
                ax.text(
                    i,
                    total,
                    f"{total:,.1f}",
                    ha="center",
                    va="bottom",
                    fontsize=18,
                    color="#333",
                )

    ax.set_xticks(x)
    ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=20)
    ax.set_ylabel(chart.yAxisLabel, fontsize=24)
    if not clean:
        ax.set_title(title, fontsize=28, fontweight="bold", pad=12)
    ax.legend(
        legend_handles,
        legend_labels,
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=_legend_ncols_for_labels(legend_labels),
        fontsize=20,
        frameon=False,
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.55,
    )
    ax.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax.tick_params(axis="y", labelsize=18)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # Ajustar límite Y superior para que las líneas no queden cortadas
    if not line_series:
        y_top = bottom.max() if len(bottom) > 0 else 1.0
    else:
        line_max = max(
            np.nanmax(np.array(s.data, dtype=float)) for s in line_series
        ) if line_series else 0.0
        y_top = max(bottom.max(), line_max) if len(bottom) > 0 else line_max
    y_top = max(y_top, 1.0) * 1.20  # 20% headroom — asegura último tick ≥ max dato

    if y_axis_min is not None or y_axis_max is not None:
        cur_lo = float(y_axis_min) if y_axis_min is not None else 0.0
        cur_hi = float(y_axis_max) if y_axis_max is not None else y_top
        ax.set_ylim(cur_lo, cur_hi)
    else:
        ax.set_ylim(0.0, y_top)

    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)
    buf.seek(0)
    return buf


def _series_style_for_render(s: ChartSeries) -> dict:
    """Produce kwargs de estilo para ax.plot() a partir de atributos opcionales de la serie.

    Soporta:
    - ``lineStyle``: estilo de línea (``"solid"``, ``"dashed"``, ``"dotted"``, etc.)
    - ``markerSymbol``: símbolo del marcador (``"circle"``, ``"diamond"``, ``"square"``, etc.)
    - ``markerRadius``: tamaño del marcador
    - ``lineWidth``: ancho de línea
    """
    style: dict = {}
    ls = getattr(s, "lineStyle", None)
    if ls and isinstance(ls, str):
        dash_map = {
            "solid": "-", "dashed": "--", "dotted": ":", "dashdot": "-.",
            "ShortDash": "--", "ShortDot": ":", "ShortDashDot": "-.",
            "LongDash": "--",
        }
        style["linestyle"] = dash_map.get(ls, ls)
    ms = getattr(s, "markerSymbol", None)
    if ms and isinstance(ms, str) and ms != "none":
        marker_map = {
            "circle": "o", "diamond": "D", "square": "s", "triangle": "^",
            "triangle-down": "v", "cross": "x", "plus": "+",
        }
        style["marker"] = marker_map.get(ms, ms)
    mr = getattr(s, "markerRadius", None)
    if mr is not None:
        try:
            style["markersize"] = float(mr)
        except (ValueError, TypeError):
            pass
    lw = getattr(s, "lineWidth", None)
    if lw is not None:
        try:
            style["linewidth"] = float(lw)
        except (ValueError, TypeError):
            pass
    return style


def _render_line_chart(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como gráfica de líneas con matplotlib."""
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    categories = chart.categories
    n_cats = len(categories)
    x = np.arange(n_cats)

    fig, ax = plt.subplots(figsize=(max(12, n_cats * 0.5), 7))

    for s in _filter_hidden_series(list(chart.series), hidden_series):
        values = np.array(s.data, dtype=float)
        style = _series_style_for_render(s)
        ax.plot(
            x,
            values,
            label=s.name,
            color=s.color,
            **style,
        )

    ax.set_xticks(x)
    ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=20)
    ax.set_ylabel(chart.yAxisLabel, fontsize=24)
    if not clean:
        ax.set_title(title, fontsize=28, fontweight="bold", pad=12)
    # Orden de leyenda:
    #   1) Series naturales en su orden natural (misma convención que la
    #      leyenda del app — primera serie del array primero).
    #   2) Series manuales (sintéticas) SIEMPRE al final, en su orden natural.
    _line_handles, _line_labels = ax.get_legend_handles_labels()
    _synth_flags = [bool(getattr(s, "is_synthetic", False)) for s in chart.series]
    _natural = [
        (h, l) for (h, l, f) in zip(_line_handles, _line_labels, _synth_flags) if not f
    ]
    _synth = [
        (h, l) for (h, l, f) in zip(_line_handles, _line_labels, _synth_flags) if f
    ]
    _ordered = _natural + _synth
    ax.legend(
        [h for h, _ in _ordered],
        [l for _, l in _ordered],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=_legend_ncols_for_labels([s.name for s in chart.series]),
        fontsize=20,
        frameon=False,
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.55,
    )
    ax.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax.tick_params(axis="y", labelsize=18)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if y_axis_min is not None or y_axis_max is not None:
        cur_lo, cur_hi = ax.get_ylim()
        ax.set_ylim(
            float(y_axis_min) if y_axis_min is not None else cur_lo,
            float(y_axis_max) if y_axis_max is not None else cur_hi,
        )

    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)
    buf.seek(0)
    return buf


def _render_stacked_area(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como áreas apiladas con matplotlib.

    Soporta **series mixtas**: las series con ``chart_type='line'`` se dibujan
    como líneas sobre las áreas en lugar de apilarse.
    """
    import io
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib.lines import Line2D as _Line2D

    categories = chart.categories
    n_cats = len(categories)
    x = np.arange(n_cats)

    fig, ax = plt.subplots(figsize=(max(12, n_cats * 0.5), 7))

    # Separar series en áreas y líneas (respetando chart_type)
    all_series = _filter_hidden_series(list(chart.series), hidden_series)
    area_series = [s for s in all_series if not _is_line_series(s)]
    line_series = [s for s in all_series if _is_line_series(s)]

    # ── Áreas apiladas ────────────────────────────────────────────────
    if area_series:
        # stackplot dibuja la primera serie al fondo. Para que la convención
        # coincida con Highcharts (primera serie del array → arriba),
        # invertimos el orden antes de pasarlo a stackplot.
        rev_series = list(reversed(area_series))
        ys = [
            np.nan_to_num(
                np.array(s.data, dtype=float),
                nan=0.0,
                posinf=0.0,
                neginf=0.0,
            )
            for s in rev_series
        ]
        labels = [s.name for s in rev_series]
        colors = [getattr(s, "color", None) for s in rev_series]
        ax.stackplot(
            x,
            np.vstack(ys) if ys else np.zeros((0, n_cats)),
            labels=labels,
            colors=[c if c else None for c in colors],
            alpha=0.9,
            linewidth=0.5,
            edgecolor="white",
        )

    # ── Líneas sobre las áreas ────────────────────────────────────────
    for s in line_series:
        raw = np.array(s.data, dtype=float)
        values = np.nan_to_num(raw, nan=float("nan"), posinf=float("nan"), neginf=float("nan"))
        ax.plot(x, values, color=s.color, linewidth=2.5, marker="o", markersize=4)

    # ── Handles de leyenda mixtos ─────────────────────────────────────
    legend_handles: list[_Line2D] = []
    legend_labels: list[str] = []

    # Áreas: círculos (orden invertido = primero arriba del stack)
    for s in reversed(area_series):
        legend_handles.append(
            _Line2D(
                [0], [0],
                marker="o", color=s.color, linestyle="None",
                markersize=10, markerfacecolor=s.color, markeredgecolor=s.color,
            )
        )
        legend_labels.append(s.name)
    # Líneas: segmento de línea con marcador
    for s in line_series:
        legend_handles.append(
            _Line2D(
                [0], [0],
                color=s.color, linewidth=2.5, marker="o", markersize=5,
                markeredgecolor=s.color, markerfacecolor=s.color,
            )
        )
        legend_labels.append(s.name)

    ax.set_xticks(x)
    ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=20)
    ax.tick_params(axis="y", labelsize=18)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax.set_ylabel(chart.yAxisLabel, fontsize=24)
    if not clean:
        ax.set_title(title, fontsize=28, fontweight="bold", pad=12)

    # Orden de leyenda: naturales invertidos + sintéticas al final.
    _synth_flags = [bool(getattr(s, "is_synthetic", False)) for s in chart.series]
    _natural = [
        (h, l)
        for (h, l, f) in zip(legend_handles, legend_labels, _synth_flags)
        if not f
    ]
    _synth = [
        (h, l) for (h, l, f) in zip(legend_handles, legend_labels, _synth_flags) if f
    ]
    _ordered = list(reversed(_natural)) + _synth
    ax.legend(
        [h for h, _ in _ordered],
        [l for _, l in _ordered],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=_legend_ncols_for_labels(legend_labels),
        fontsize=20,
        frameon=False,
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.55,
    )
    ax.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlim(x.min() if n_cats > 0 else 0, x.max() if n_cats > 0 else 0)

    # Ajustar límite Y para que las líneas no queden cortadas
    if not area_series:
        area_top = 0.0
    else:
        area_top = np.sum(ys, axis=0).max() if len(ys) > 1 else ys[0].max()
    if not line_series:
        line_max = 0.0
    else:
        line_max = max(
            np.nanmax(np.array(s.data, dtype=float)) for s in line_series
        ) if line_series else 0.0
    y_top = max(area_top, line_max, 1.0) * 1.20

    if y_axis_min is not None or y_axis_max is not None:
        cur_lo = float(y_axis_min) if y_axis_min is not None else 0.0
        cur_hi = float(y_axis_max) if y_axis_max is not None else y_top
        ax.set_ylim(cur_lo, cur_hi)
    else:
        ax.set_ylim(0.0, y_top)

    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)
    buf.seek(0)
    return buf


def render_comparison_by_year_bytes(
    data: CompareChartResponse,
    fmt: str = "svg",
    *,
    view_mode: str = "column",
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
) -> bytes:
    """Renderiza una comparación por año (subplots, un panel por año).

    ``view_mode``:
    - ``"column"`` (default): barras agrupadas (una barra por escenario).
    - ``"area"``: áreas apiladas (aporta por escenario dentro de cada categoría).
    - ``"line"``: líneas (un trazado por escenario).

    Si ``clean=True``, omite el título general.
    """
    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    subplots = [sp for sp in data.subplots if sp.series]
    if not subplots:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center")
        ax.set_axis_off()
        buf = io.BytesIO()
        fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()


    n = len(subplots)
    cols = min(3, n)
    rows = (n + cols - 1) // cols
    fig, axes = plt.subplots(
        rows, cols, figsize=(max(6, cols * 5), max(4, rows * 4)), squeeze=False
    )

    # Paleta consistente por nombre de serie a través de subplots.
    all_names: list[str] = []
    for sp in subplots:
        for s in _filter_hidden_series(list(sp.series), hidden_series):
            if s.name not in all_names:
                all_names.append(s.name)
    name_to_color: dict[str, str] = {}
    for sp in subplots:
        for s in _filter_hidden_series(list(sp.series), hidden_series):
            if s.name not in name_to_color and getattr(s, "color", None):
                name_to_color[s.name] = s.color  # type: ignore[assignment]

    # ── Calcular el máximo global Y para compartir escala entre subplots ──
    global_max = 0.0
    for sp in subplots:
        visible_series = _filter_hidden_series(list(sp.series), hidden_series)
        nc = len(sp.categories)
        if nc == 0:
            continue
        if view_mode == "area":
            stack = np.zeros(nc)
            for s in visible_series:
                values = np.nan_to_num(np.array(s.data, dtype=float), nan=0.0)
                if values.size < nc:
                    values = np.pad(values, (0, nc - values.size))
                stack += values[:nc]
            if stack.size:
                global_max = max(global_max, float(stack.max()))
        else:
            for s in visible_series:
                values = np.array(s.data, dtype=float)
                finite = values[np.isfinite(values)]
                if finite.size:
                    global_max = max(global_max, float(finite.max()))
    if global_max <= 0:
        global_max = 1.0
    y_top = global_max * 1.20

    for idx, sp in enumerate(subplots):
        ax = axes[idx // cols][idx % cols]
        categories = list(sp.categories)
        nc = len(categories)
        ns = len(sp.series)
        if nc == 0 or ns == 0:
            ax.set_axis_off()
            continue
        visible_series = _filter_hidden_series(list(sp.series), hidden_series)
        nvs = len(visible_series)
        x = np.arange(nc)
        if view_mode == "area":
            ys = [
                np.nan_to_num(np.array(s.data, dtype=float), nan=0.0)
                for s in reversed(visible_series)
            ]
            labels = [s.name for s in reversed(visible_series)]
            colors = [name_to_color.get(s.name) or getattr(s, "color", None) for s in reversed(visible_series)]
            ax.stackplot(x, ys, labels=labels, colors=colors, alpha=0.9, linewidth=0.5, edgecolor="white")
        elif view_mode == "line":
            for s in visible_series:
                values = np.array(s.data, dtype=float)
                ax.plot(x, values, label=s.name, color=name_to_color.get(s.name) or getattr(s, "color", None),
                        linewidth=2.0, marker="o", markersize=4)
        else:
            width = 0.8 / (nvs or 1)
            for si, s in enumerate(visible_series):
                offset = (si - (max(nvs, 1) - 1) / 2) * width
                ax.bar(
                    x + offset,
                    s.data,
                    width=width,
                    label=s.name,
                    color=name_to_color.get(s.name) or getattr(s, "color", None),
                )
        ax.set_title(f"Año {sp.year}", fontsize=22)
        ax.set_xticks(x)
        ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=20)
        ax.set_ylabel(data.yAxisLabel, fontsize=24)
        ax.tick_params(axis="y", labelsize=18)
        from matplotlib.ticker import FuncFormatter as _FuncFormatter

        ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
        ax.grid(axis="y", alpha=0.3, linewidth=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        if idx == 0:
            ax.legend(
                loc="upper center",
                bbox_to_anchor=(0.5, -0.2),
                ncol=_legend_ncols_for_labels([s.name for s in sp.series], hard_cap=4),
            fontsize=20,
                frameon=False,
                handlelength=1.0,
                handletextpad=0.6,
                columnspacing=1.85,
                labelspacing=0.55,
            )

    # Forzar la misma escala Y en todos los subplots.
    effective_y_lo = float(y_axis_min) if y_axis_min is not None else 0.0
    effective_y_hi = float(y_axis_max) if y_axis_max is not None else y_top
    for j in range(n):
        ax = axes[j // cols][j % cols]
        ax.set_ylim(effective_y_lo, effective_y_hi)

    # Ocultar axes sobrantes
    for j in range(n, rows * cols):
        axes[j // cols][j % cols].set_axis_off()

    if not clean:
        fig.suptitle(data.title, fontsize=28, fontweight="bold")
    fig.tight_layout(rect=[0, 0.02, 1, 0.96])

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def render_pareto_chart_bytes(
    pareto: ParetoChartResponse,
    fmt: str = "svg",
    *,
    clean: bool = False,
) -> bytes:
    """Renderiza un ParetoChartResponse (barras descendentes + % acumulado).

    Si ``clean=True`` se omite el título.
    """
    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    categories = list(pareto.categories)
    values = list(pareto.values)
    cum_pct = list(pareto.cumulative_percent)
    n = len(categories)
    if n == 0:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center")
        ax.set_axis_off()
        buf = io.BytesIO()
        fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    x = np.arange(n)
    fig, ax1 = plt.subplots(figsize=(max(12, n * 0.5), 7))
    ax1.bar(x, values, color="#60a5fa", edgecolor="#1e3a8a", linewidth=0.5)
    ax1.set_ylabel(pareto.yAxisLabel, fontsize=24, color="#1e3a8a")
    ax1.set_xticks(x)
    # Eje X a 45° (más legible que vertical para etiquetas largas tipo
    # tecnología/sector). ``ha="right"`` ancla el final de la etiqueta al tick
    # para que no se solape con la barra siguiente.
    ax1.set_xticklabels(
        categories,
        rotation=45,
        ha="right",
        rotation_mode="anchor",
        fontsize=20,
    )
    ax1.tick_params(axis="y", labelsize=18)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax1.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax1.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax1.spines["top"].set_visible(False)

    ax2 = ax1.twinx()
    ax2.plot(x, cum_pct, color="#dc2626", marker="o", linewidth=2)
    ax2.set_ylabel("% acumulado", fontsize=24, color="#dc2626")
    ax2.tick_params(axis="y", labelsize=18)
    ax2.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax2.set_ylim(0, 110)
    ax2.spines["top"].set_visible(False)

    if not clean:
        ax1.set_title(pareto.title, fontsize=28, fontweight="bold", pad=12)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _facet_x_axis_label_step(categories: list[Any]) -> int:
    """Paso entre etiquetas visibles del eje X (responsive).

    Si las categorías son años (1800–2200): como mínimo cada **2 años** y a lo sumo
    unas **12** marcas con texto para horizontes largos.
    En otro caso, limita a ~12 etiquetas sobre el total de categorías.
    """
    n = len(categories)
    if n <= 1:
        return 1

    try:
        years = [int(str(c).strip()) for c in categories]
    except (ValueError, TypeError):
        years = None
    else:
        if not years or not all(1800 <= y <= 2200 for y in years):
            years = None

    target_visible = 12
    if years is not None:
        return max(2, (n + target_visible - 1) // target_visible)

    return max(1, (n + target_visible - 1) // target_visible)


def _facet_x_ticklabels_thinned(
    categories: list[Any],
    step: int,
    fixed_labels: set[str] | None = None,
) -> list[str]:
    """Etiquetas con cadena vacía en índices omitidos; asegura inicio y fin legibles.

    Si se provee *fixed_labels*, solo se etiquetan las categorías que aparezcan
    en ese conjunto (en vez del muestreo por *step*).
    """
    n = len(categories)
    if step < 1:
        step = 1
    out: list[str] = [""] * n
    if fixed_labels is not None:
        for i, cat in enumerate(categories):
            if str(cat) in fixed_labels:
                out[i] = str(cat)
    else:
        for i in range(0, n, step):
            out[i] = str(categories[i])
    if n > 1:
        if not out[0]:
            out[0] = str(categories[0])
        if not out[-1]:
            out[-1] = str(categories[-1])
    return out


def render_comparison_facet_figure_bytes(
    data: CompareChartFacetResponse,
    fmt: str = "png",
    *,
    layout: str = "horizontal",
    view_mode: str = "column",
    legend_title: str | None = None,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    series_order: list[str] | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
    fixed_labels: set[str] | None = None,
) -> bytes:
    """Una sola figura: facetas en fila/columna, título global, leyenda inferior (Matplotlib).

    ``layout`` puede ser ``"horizontal"`` (1 fila × N columnas) o
    ``"vertical"`` (N filas × 1 columna).

    ``view_mode``:
    - ``"column"`` (default): barras apiladas.
    - ``"area"``: áreas apiladas.
    - ``"line"``: líneas (un trazado por serie).

    Prioriza **legibilidad**: misma escala Y entre escenarios, tipografía clara, leyenda
    con marco y números formateados en ejes y totales de barra cuando aportan.
    """
    import io

    import matplotlib

    assert layout in ("horizontal", "vertical"), f"layout inválido: {layout}"

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib import patheffects as pe
    from matplotlib.patches import Rectangle
    from matplotlib.ticker import AutoMinorLocator, FuncFormatter, MaxNLocator

    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")

    facets = [f for f in data.facets if f.series]
    if not facets:
        raise ValueError("Sin facetas con series para exportar")

    n = len(facets)
    legend_order: list[tuple[str, str]] = []
    seen_names: set[str] = set()
    for facet in facets:
        for s in _filter_hidden_series(list(facet.series), hidden_series):
            if s.name not in seen_names:
                seen_names.add(s.name)
                legend_order.append((s.name, s.color))

    # La leyenda compartida sigue el orden de primera aparición entre facets
    # (mismo algoritmo que buildSharedLegendItems en el frontend). Con un
    # orden custom se usa ese como criterio principal y las series fuera de él
    # conservan su orden natural al final (sort estable).
    if series_order:
        custom_rank: dict[str, int] = {n: i for i, n in enumerate(series_order)}
        legend_order.sort(key=lambda item: custom_rank.get(item[0], len(series_order)))

    n_leg_items = len(legend_order)
    legend_labels_full = [name for name, _c in legend_order]
    max_leg_label_len = max((len(lab) for lab in legend_labels_full), default=0)

    # Leyenda: **siempre** texto completo; menos columnas si los nombres son largos.
    LEG_NCOL_MAX = 8
    if max_leg_label_len > 48:
        leg_ncol = max(1, min(2, n_leg_items))
    elif max_leg_label_len > 34:
        leg_ncol = max(1, min(3, n_leg_items))
    elif max_leg_label_len > 22:
        leg_ncol = max(1, min(4, n_leg_items))
    else:
        leg_ncol = max(1, min(LEG_NCOL_MAX, n_leg_items))
    n_leg_rows = max(1, (n_leg_items + leg_ncol - 1) // leg_ncol)

    # Ancho por panel: presupuesto horizontal de ~26″ repartido entre `n` paneles.
    inter_panel = 0.55
    fig_w_target = 26.0
    w_per_facet = max(
        5.0,
        min(9.0, (fig_w_target - inter_panel * max(0, n - 1)) / n),
    )

    leg_font_estimate = (
        16.0
        if max_leg_label_len > 52 or n_leg_items > 14
        else (17.0 if max_leg_label_len > 36 or n_leg_items > 10 else 18.0)
    )
    leg_font_estimate = float(min(leg_font_estimate, 19.0))

    title_band_inch = 1.25
    x_label_inch = 1.10
    gap_inch = 0.24
    legend_pad_inch = 0.12
    line_h_inch = leg_font_estimate * 1.35 / 72.0
    legend_h_inch = line_h_inch * n_leg_rows + 0.12
    bottom_margin_inch = legend_pad_inch + legend_h_inch + gap_inch + x_label_inch

    if layout == "vertical":
        panel_h_inch = max(4.0, w_per_facet * 0.45)
        fig_w = max(14.0, min(w_per_facet + 4.0, 18.0))
        fig_h = panel_h_inch * n + inter_panel * max(0, n - 1) + title_band_inch + bottom_margin_inch
        fig_h = float(min(max(fig_h, 6.0), 26.0))
        fig, axes = plt.subplots(n, 1, figsize=(fig_w, fig_h), squeeze=False)
        facet_axes = axes[:, 0]
    else:
        panel_h_inch = max(4.0, w_per_facet * 0.55)
        fig_w = w_per_facet * n + inter_panel * max(0, n - 1)
        fig_w = max(9.0, fig_w)
        fig_h = panel_h_inch + title_band_inch + bottom_margin_inch
        fig_h = float(min(max(fig_h, 6.0), 14.0))
        fig, axes = plt.subplots(1, n, figsize=(fig_w, fig_h), squeeze=False)
        facet_axes = axes[0]

    y_label = data.yAxisLabel or "Valor"
    stack_tops: list[np.ndarray] = []
    line_maxes: list[float] = []

    for idx, (ax, facet) in enumerate(zip(facet_axes, facets)):
        categories = list(facet.categories)
        n_cats = len(categories)
        x = np.arange(n_cats, dtype=float)

        facet_series = _filter_hidden_series(list(facet.series), hidden_series)
        bar_series = [s for s in facet_series if not _is_line_series(s)]
        line_series = [s for s in facet_series if _is_line_series(s)]

        # Render según view_mode
        facet_bar_max = 0.0
        if view_mode == "area":
            # Áreas apiladas: todas las series (bar + line) como áreas
            area_series = list(reversed(facet_series))
            ys = []
            for s in area_series:
                raw = np.array(s.data, dtype=float)
                if raw.size < n_cats:
                    raw = np.pad(raw, (0, n_cats - int(raw.size)))
                elif raw.size > n_cats:
                    raw = raw[:n_cats]
                ys.append(np.nan_to_num(raw, nan=0.0, posinf=0.0, neginf=0.0))
            labels = [s.name for s in area_series]
            colors = [s.color for s in area_series]
            ax.stackplot(x, ys, labels=labels, colors=colors, alpha=0.9, linewidth=0.5, edgecolor="white")
            for v in ys:
                facet_bar_max = max(facet_bar_max, float(np.max(v)))
            stack_tops.append(np.sum(ys, axis=0))
            line_maxes.append(0.0)
        elif view_mode == "line":
            # Líneas: todas las series como líneas
            for s in facet_series:
                raw = np.array(s.data, dtype=float)
                if raw.size < n_cats:
                    raw = np.pad(raw, (0, n_cats - int(raw.size)))
                elif raw.size > n_cats:
                    raw = raw[:n_cats]
                values = np.nan_to_num(raw, nan=float("nan"), posinf=float("nan"), neginf=float("nan"))
                ax.plot(x, values, color=s.color, linewidth=2.0, marker="o", markersize=4)
            stack_tops.append(np.zeros(n_cats))
            line_maxes.append(0.0)
        else:
            # Barras apiladas (comportamiento original)
            bottom = np.zeros(n_cats, dtype=float)
            for s in reversed(bar_series):
                raw = np.array(s.data, dtype=float)
                if raw.size < n_cats:
                    raw = np.pad(raw, (0, n_cats - int(raw.size)))
                elif raw.size > n_cats:
                    raw = raw[:n_cats]
                values = np.nan_to_num(raw, nan=0.0, posinf=0.0, neginf=0.0)
                facet_bar_max = max(facet_bar_max, float(np.max(values)))
                ax.bar(
                    x,
                    values,
                    bottom=bottom,
                    color=s.color,
                    width=0.74,
                    edgecolor="#ffffff",
                    linewidth=0.45,
                )
                bottom = bottom + values
            stack_tops.append(bottom.copy())

            # Líneas sobre las barras
            facet_line_max = 0.0
            for s in line_series:
                raw = np.array(s.data, dtype=float)
                if raw.size < n_cats:
                    raw = np.pad(raw, (0, n_cats - int(raw.size)))
                elif raw.size > n_cats:
                    raw = raw[:n_cats]
                values = np.nan_to_num(raw, nan=float("nan"), posinf=float("nan"), neginf=float("nan"))
                ax.plot(x, values, color=s.color, linewidth=2.5, marker="o", markersize=4)
                finite_values = values[np.isfinite(values)]
                if finite_values.size:
                    facet_line_max = max(facet_line_max, float(finite_values.max()))
            line_maxes.append(facet_line_max)

        ax.set_xticks(x)
        if fixed_labels is not None:
            x_labels = _facet_x_ticklabels_thinned(categories, 1, fixed_labels=fixed_labels)
        else:
            x_step = _facet_x_axis_label_step(categories)
            x_labels = _facet_x_ticklabels_thinned(categories, x_step)
        n_labeled = sum(1 for lb in x_labels if lb)
        x_fs = (
            20
            if n_labeled > 14 or n_cats > 36
            else (20 if n_cats > 22 or n_labeled > 11 else 22)
        )
        ax.set_xticklabels(
            x_labels,
            rotation=90,
            ha="center",
            fontsize=x_fs,
            color="#1e293b",
        )
        ax.set_ylabel(
            y_label,
            fontsize=22,
            color="#0f172a",
            labelpad=8,
        )
        sim_lbl = (
            facet.display_name or facet.scenario_name or f"Job {facet.job_id}"
        ).strip()
        tag_lbl = (facet.scenario_tag_name or "").strip()
        facet_title = f"{sim_lbl} — {tag_lbl}" if tag_lbl else sim_lbl
        if layout == "vertical":
            ax.text(
                -0.20, 0.5, facet_title,
                transform=ax.transAxes,
                fontsize=22, fontweight="bold",
                va="center", ha="right",
                color="#0f172a",
            )
        else:
            ax.set_title(
                facet_title,
                fontsize=28,
                color="#0f172a",
                pad=10,
            )
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        for _side in ("left", "bottom"):
            ax.spines[_side].set_visible(True)
            ax.spines[_side].set_color("#1e293b")
            ax.spines[_side].set_linewidth(1.35)
        ax.tick_params(
            axis="y",
            labelsize=20,
            colors="#0f172a",
            width=1.15,
            length=6,
            labelcolor="#0f172a",
        )
        ax.tick_params(
            axis="x",
            colors="#0f172a",
            width=1.15,
            length=5,
            labelcolor="#1e293b",
        )
        ax.set_facecolor("#ffffff")

    global_max = 0.0
    for b in stack_tops:
        if b.size:
            b_clean = b[np.isfinite(b)]
            if b_clean.size:
                global_max = max(global_max, float(b_clean.max()))
    for lm in line_maxes:
        global_max = max(global_max, lm)
    if global_max <= 0:
        global_max = 1.0
    y_top = global_max * 1.20

    show_stack_totals = view_mode != "line" and all(len(b) <= 18 for b in stack_tops)

    effective_y_lo = float(y_axis_min) if y_axis_min is not None else 0.0
    effective_y_hi = float(y_axis_max) if y_axis_max is not None else y_top

    for ax, bottom in zip(facet_axes, stack_tops):
        ax.set_ylim(effective_y_lo, effective_y_hi)
        ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _p: format_axis_3sig(v)))
        ax.yaxis.set_major_locator(
            MaxNLocator(nbins=7, min_n_ticks=5, steps=[1, 2, 2.5, 5, 10]),
        )
        ax.yaxis.set_minor_locator(AutoMinorLocator(2))

        ax.grid(which="major", axis="y", linestyle="-", linewidth=0.95, color="#64748b", alpha=0.55, zorder=0)
        ax.grid(which="minor", axis="y", linestyle=":", linewidth=0.7, color="#94a3b8", alpha=0.5, zorder=0)
        ax.grid(which="major", axis="x", linestyle="-", linewidth=0.75, color="#94a3b8", alpha=0.42, zorder=0)

        if not show_stack_totals:
            continue
        if clean:
            continue
        n_cats = len(bottom)
        for i in range(n_cats):
            if i % 2 != 0:
                continue
            total = float(bottom[i])
            if total <= 0 or total < global_max * 0.018:
                continue
            t = ax.text(
                i,
                total,
                f"{total:,.1f}",
                ha="center",
                va="bottom",
                fontsize=18,
                color="#0f172a",
            )
            t.set_path_effects([pe.withStroke(linewidth=2.5, foreground="white")])

    fig.patch.set_facecolor("#ffffff")
    if not clean:
        suptitle_y = 1.0 - 0.28 / fig_h
        fig.suptitle(
            data.title,
            fontsize=32,
            fontweight="bold",
            color="#020617",
            y=suptitle_y,
        )

    from matplotlib.lines import Line2D as _Line2D

    handles = []
    is_line_mode = view_mode == "line"
    for name, c in legend_order:
        if is_line_mode:
            handles.append(
                _Line2D(
                    [0], [0],
                    color=c, linewidth=2.0, marker="o", markersize=5,
                    markeredgecolor=c, markerfacecolor=c,
                )
            )
        elif view_mode == "area":
            handles.append(
                _Line2D(
                    [0], [0],
                    marker="s", color=c, linestyle="None",
                    markersize=10, markerfacecolor=c, markeredgecolor=c,
                )
            )
        else:
            line_names: set[str] = set()
            for facet in facets:
                for s in _filter_hidden_series(list(facet.series), hidden_series):
                    if _is_line_series(s):
                        line_names.add(s.name)
            if name in line_names:
                handles.append(
                    _Line2D(
                        [0], [0],
                        color=c, linewidth=2.5, marker="o", markersize=5,
                        markeredgecolor=c, markerfacecolor=c,
                    )
                )
            else:
                handles.append(
                    _Line2D(
                        [0], [0],
                        marker="o", color=c, linestyle="None",
                        markersize=10, markerfacecolor=c, markeredgecolor=c,
                    )
                )
    leg_font = leg_font_estimate

    bottom_margin = bottom_margin_inch / fig_h
    top_margin = 1.0 - title_band_inch / fig_h
    legend_anchor_y = legend_pad_inch / fig_h

    fig.legend(
        handles=handles,
        labels=legend_labels_full,
        loc="lower center",
        bbox_to_anchor=(0.5, legend_anchor_y),
        ncol=leg_ncol,
        fontsize=leg_font,
        frameon=False,
        labelcolor="#0f172a",
        handlelength=0.8,
        handletextpad=0.4,
        columnspacing=1.2,
        labelspacing=0.28,
    )

    if layout == "vertical":
        left_margin = max(0.15, min(0.35, 3.0 / fig_w))
        hspace = 0.28 if n >= 4 else 0.24
        plt.subplots_adjust(
            left=left_margin,
            right=0.995,
            top=top_margin,
            bottom=bottom_margin,
            hspace=hspace,
        )
    else:
        wspace = 0.20 if n >= 4 else 0.16
        plt.subplots_adjust(
            left=0.07,
            right=0.995,
            top=top_margin,
            bottom=bottom_margin,
            wspace=wspace,
        )

    buf = io.BytesIO()
    fig.savefig(
        buf,
        format=fmt,
        dpi=200,
        facecolor="#ffffff",
        edgecolor="none",
        bbox_inches="tight",
        pad_inches=0.3,
    )
    plt.close(fig)
    return buf.getvalue()


def render_chart_visualization_bytes(
    chart: ChartDataResponse,
    fmt: str = "png",
    view_mode: str = "column",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
) -> bytes:
    """Genera PNG o SVG con Matplotlib (sin navegador).

    ``view_mode``: ``"column"`` | ``"line"`` | ``"area"``.
    Soporta filtro de series ocultas vía ``hidden_series``.
    """
    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")
    title = chart.title
    if view_mode == "line":
        buf = _render_line_chart(chart, title, fmt=fmt, y_axis_min=y_axis_min, y_axis_max=y_axis_max, clean=clean, hidden_series=hidden_series)
    elif view_mode == "area":
        buf = _render_stacked_area(chart, title, fmt=fmt, y_axis_min=y_axis_min, y_axis_max=y_axis_max, clean=clean, hidden_series=hidden_series)
    else:
        buf = _render_stacked_bar(chart, title, fmt=fmt, y_axis_min=y_axis_min, y_axis_max=y_axis_max, clean=clean, hidden_series=hidden_series)
    return buf.getvalue()


def chart_data_to_csv_bytes(chart: ChartDataResponse) -> bytes:
    """Tabla categorías × series como CSV UTF-8 con BOM (compatible con Excel)."""
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Categoria"] + [s.name for s in chart.series])
    for i, cat in enumerate(chart.categories):
        row: list[Any] = [cat]
        for s in chart.series:
            val = s.data[i] if i < len(s.data) else None
            row.append(val)
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def chart_data_to_xlsx_bytes(chart: ChartDataResponse) -> bytes:
    """Tabla categorías × series como XLSX (sin imagen embebida)."""
    import io

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_excel_sheet_name(chart.title[:31] if chart.title else "Datos", set())

    headers = [_category_column_label(chart.categories)] + [s.name for s in chart.series]
    ws.append(headers)
    for i, cat in enumerate(chart.categories):
        row: list[Any] = [cat]
        for s in chart.series:
            val = s.data[i] if i < len(s.data) else None
            row.append(val)
        ws.append(row)

    for idx in range(len(headers)):
        col_letter = get_column_letter(idx + 1)
        max_len = len(str(headers[idx]))
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _safe_excel_sheet_name(name: str, used: set[str]) -> str:
    """Nombre de hoja válido para Excel (máx. 31 caracteres, único en el libro)."""
    clean = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in name)
    clean = clean.strip()[:31] or "Hoja"
    base = clean
    counter = 1
    while clean in used:
        suffix = f"_{counter}"
        max_base = 31 - len(suffix)
        clean = f"{base[:max_base]}{suffix}"
        counter += 1
    used.add(clean)
    return clean


def _looks_like_year(value: Any) -> bool:
    """True si el valor parece un año (1900–2200)."""
    try:
        year = int(str(value).strip())
    except (ValueError, TypeError):
        return False
    return 1900 <= year <= 2200


def _category_column_label(categories: list[Any], *, default: str = "Categoría") -> str:
    """Etiqueta de columna para categorías del eje X (Año, Escenario, etc.)."""
    if categories and all(_looks_like_year(c) for c in categories):
        return "Año"
    return default


def _estimate_table_start_row(image_height_px: int) -> int:
    """Fila inicial de contenido debajo de una imagen embebida."""
    return max(int(image_height_px / 15) + 3, 28)


def _estimate_chart_anchor_row(data_end_row: int, chart_height_rows: int = 14) -> int:
    """Fila donde anclar la gráfica nativa de Excel debajo de la tabla."""
    return data_end_row + 2


def _estimate_image_anchor_row(chart_anchor_row: int, chart_height_rows: int = 14) -> int:
    """Fila donde anclar la imagen PNG debajo de la gráfica nativa."""
    return chart_anchor_row + chart_height_rows


def _hex_to_excel_color(color: str) -> str:
    """Normaliza color hex (#RRGGBB) al formato openpyxl (RRGGBB)."""
    clean = (color or "").strip().lstrip("#").upper()
    if len(clean) == 3:
        clean = "".join(ch * 2 for ch in clean)
    if len(clean) != 6:
        return "808080"
    return clean


def _apply_series_colors(chart: Any, series_colors: list[str]) -> None:
    """Aplica colores sólidos a cada serie de un chart openpyxl."""
    for idx, ser in enumerate(chart.series):
        color = _hex_to_excel_color(series_colors[idx] if idx < len(series_colors) else "#808080")
        ser.graphicalProperties.solidFill = color
        ser.graphicalProperties.line.solidFill = color


def _add_native_excel_chart(
    ws: Any,
    *,
    categories_col: int,
    data_start_col: int,
    data_end_col: int,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    series_colors: list[str],
    chart_type: str = "column",
    title: str = "",
    y_axis_label: str = "",
    anchor_cell: str = "A1",
    width: int = 24,
    height: int = 14,
    categories_end_col: int | None = None,
    multi_level_categories: bool = False,
) -> None:
    """Inserta una gráfica nativa de Excel referenciando la tabla ya escrita."""
    from openpyxl.chart import AreaChart, BarChart, LineChart, Reference

    if data_end_row < data_start_row or data_end_col < data_start_col:
        return

    if chart_type == "line":
        chart = LineChart()
    elif chart_type == "area":
        chart = AreaChart()
        chart.grouping = "stacked"
        chart.overlap = 100
    else:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100

    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_axis_label
    chart.width = width
    chart.height = height
    chart.legend.position = "b"
    # openpyxl por defecto coloca catAx en axPos="l", lo que rompe barras apiladas
    # (etiquetas de años apiladas verticalmente sin datos visibles en Excel).
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    if multi_level_categories:
        chart.x_axis.noMultiLvlLbl = False

    data_ref = Reference(
        ws,
        min_col=data_start_col,
        max_col=data_end_col,
        min_row=header_row,
        max_row=data_end_row,
    )
    cat_end = categories_end_col if categories_end_col is not None else categories_col
    cats_ref = Reference(
        ws,
        min_col=categories_col,
        max_col=cat_end,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    _apply_series_colors(chart, series_colors)
    ws.add_chart(chart, anchor_cell)


def _write_data_table(
    ws: Any,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    y_axis_label: str | None = None,
) -> tuple[int, int, int, int]:
    """Escribe tabla de datos. Retorna (header_row, data_start_row, data_end_row, last_row)."""
    from openpyxl.utils import get_column_letter

    row_idx = start_row
    if y_axis_label:
        ws.cell(row=row_idx, column=1, value=f"Unidad: {y_axis_label}")
        row_idx += 1
    header_row = row_idx
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row_idx, column=col_idx, value=header)
    row_idx += 1
    data_start_row = row_idx
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
    data_end_row = row_idx - 1
    last_row = data_end_row

    for idx in range(len(headers)):
        col_letter = get_column_letter(idx + 1)
        max_len = len(str(headers[idx]))
        for r in range(start_row, row_idx):
            val = ws.cell(row=r, column=idx + 1).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    return header_row, data_start_row, data_end_row, last_row


def _comparison_has_two_dimension_headers(headers: list[str]) -> bool:
    """True si la tabla de comparación tiene dos columnas de dimensión (Escenario × Año)."""
    if len(headers) < 3:
        return False
    dim_labels = {"Categoria", "Categoría", "Año", "Escenario"}
    return headers[0] in dim_labels and headers[1] in dim_labels


def _write_comparison_sheet(
    ws: Any,
    *,
    image_bytes: bytes,
    headers: list[str],
    rows: list[list[Any]],
    y_axis_label: str,
    sheet_title: str,
    view_mode: str = "column",
    chart_title: str = "",
    series_colors: list[str] | None = None,
) -> None:
    """Hoja de comparación: tabla, gráfica nativa (categorías multinivel) e imagen PNG."""
    import io

    from openpyxl.drawing.image import Image as XLImage

    ws.title = sheet_title
    header_row, data_start_row, data_end_row, _last_row = _write_data_table(
        ws, 1, headers, rows, y_axis_label
    )

    colors = list(series_colors or [])
    two_dims = _comparison_has_two_dimension_headers(headers)
    if two_dims:
        categories_col = 1
        categories_end_col = 2
        data_start_col = 3
        data_end_col = len(headers)
        multi_level = True
        if not colors:
            colors = ["#808080"] * (data_end_col - data_start_col + 1)
    else:
        categories_col = 1
        categories_end_col = None
        data_start_col = 2
        data_end_col = len(headers)
        multi_level = False

    chart_anchor_row = _estimate_chart_anchor_row(data_end_row)
    if data_end_col >= data_start_col and data_end_row >= data_start_row:
        try:
            _add_native_excel_chart(
                ws,
                categories_col=categories_col,
                categories_end_col=categories_end_col,
                data_start_col=data_start_col,
                data_end_col=data_end_col,
                header_row=header_row,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                series_colors=colors,
                chart_type=view_mode if view_mode in ("column", "line", "area") else "column",
                title=chart_title or sheet_title,
                y_axis_label=y_axis_label,
                anchor_cell=f"A{chart_anchor_row}",
                multi_level_categories=multi_level,
            )
            image_row = _estimate_image_anchor_row(chart_anchor_row)
        except Exception:
            image_row = _estimate_chart_anchor_row(data_end_row)
    else:
        image_row = _estimate_chart_anchor_row(data_end_row)

    img = XLImage(io.BytesIO(image_bytes))
    max_width = 900
    if img.width > max_width:
        scale = max_width / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
    ws.add_image(img, f"A{image_row}")


def _write_chart_sheet(
    ws: Any,
    *,
    image_bytes: bytes,
    categories: list[str],
    series: list[ChartSeries],
    y_axis_label: str,
    sheet_title: str | None = None,
    view_mode: str = "column",
    chart_title: str = "",
    category_header: str | None = None,
) -> None:
    """Escribe hoja: tabla arriba, gráfica nativa Excel, imagen PNG debajo."""
    import io

    from openpyxl.drawing.image import Image as XLImage

    if sheet_title:
        ws.title = sheet_title

    cat_label = category_header or _category_column_label(categories)
    headers = [cat_label] + [s.name for s in series]
    rows: list[list[Any]] = []
    for i, cat in enumerate(categories):
        row: list[Any] = [cat]
        for s in series:
            val = s.data[i] if i < len(s.data) else 0
            row.append(val if val is not None else 0)
        rows.append(row)

    header_row, data_start_row, data_end_row, _last_row = _write_data_table(
        ws, 1, headers, rows, y_axis_label
    )

    series_colors = [s.color for s in series]
    chart_anchor_row = _estimate_chart_anchor_row(data_end_row)
    chart_anchor = f"A{chart_anchor_row}"
    try:
        _add_native_excel_chart(
            ws,
            categories_col=1,
            data_start_col=2,
            data_end_col=1 + len(series),
            header_row=header_row,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
            series_colors=series_colors,
            chart_type=view_mode if view_mode in ("column", "line", "area") else "column",
            title=chart_title or sheet_title or "",
            y_axis_label=y_axis_label,
            anchor_cell=chart_anchor,
        )
    except Exception:
        chart_anchor_row = data_end_row + 1

    img = XLImage(io.BytesIO(image_bytes))
    max_width = 900
    if img.width > max_width:
        scale = max_width / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
    image_row = _estimate_image_anchor_row(chart_anchor_row)
    ws.add_image(img, f"A{image_row}")


def _facet_display_name(facet: FacetData) -> str:
    sim = (facet.display_name or facet.scenario_name or str(facet.job_id)).strip()
    tag = (facet.scenario_tag_name or "").strip()
    return f"{sim} — {tag}" if tag else sim


def _subplot_display_name(sp: SubplotData) -> str:
    if sp.scenario_name:
        return sp.scenario_name.strip()
    return f"Año {sp.year}"


def _collect_series_names_from_series_lists(
    series_lists: list[list[ChartSeries]],
) -> list[str]:
    names: list[str] = []
    for series in series_lists:
        for s in series:
            if s.name not in names:
                names.append(s.name)
    return names


def _series_colors_for_names(
    series_names: list[str],
    series_lists: list[list[ChartSeries]],
) -> list[str]:
    by_name: dict[str, str] = {}
    for sl in series_lists:
        for s in sl:
            if s.name not in by_name and s.color:
                by_name[s.name] = s.color
    return [by_name.get(name, "#808080") for name in series_names]


def _build_facet_combined_table(facets: list[FacetData]) -> tuple[list[str], list[list[Any]]]:
    series_names = _collect_series_names_from_series_lists([f.series for f in facets])
    headers = ["Escenario", "Año"] + series_names
    rows: list[list[Any]] = []
    for facet in facets:
        scenario = _facet_display_name(facet)
        series_by_name = {s.name: s for s in facet.series}
        for i, cat in enumerate(facet.categories):
            row: list[Any] = [scenario, cat]
            for name in series_names:
                s = series_by_name.get(name)
                val = s.data[i] if s is not None and i < len(s.data) else 0
                row.append(val if val is not None else 0)
            rows.append(row)
    return headers, rows


def _build_by_year_combined_table(subplots: list[SubplotData]) -> tuple[list[str], list[list[Any]]]:
    series_names = _collect_series_names_from_series_lists([sp.series for sp in subplots])
    headers = ["Año", "Escenario"] + series_names
    rows: list[list[Any]] = []
    for sp in subplots:
        series_by_name = {s.name: s for s in sp.series}
        for i, cat in enumerate(sp.categories):
            row: list[Any] = [sp.year, cat]
            for name in series_names:
                s = series_by_name.get(name)
                val = s.data[i] if s is not None and i < len(s.data) else 0
                row.append(val if val is not None else 0)
            rows.append(row)
    return headers, rows


def _build_by_year_alt_combined_table(subplots: list[SubplotData]) -> tuple[list[str], list[list[Any]]]:
    series_names = _collect_series_names_from_series_lists([sp.series for sp in subplots])
    headers = ["Escenario", "Año"] + series_names
    rows: list[list[Any]] = []
    for sp in subplots:
        scenario = _subplot_display_name(sp)
        series_by_name = {s.name: s for s in sp.series}
        for i, cat in enumerate(sp.categories):
            row: list[Any] = [scenario, cat]
            for name in series_names:
                s = series_by_name.get(name)
                val = s.data[i] if s is not None and i < len(s.data) else 0
                row.append(val if val is not None else 0)
            rows.append(row)
    return headers, rows


def _subplot_to_chart_response(
    sp: SubplotData,
    title: str,
    y_axis_label: str,
) -> ChartDataResponse:
    return ChartDataResponse(
        categories=list(sp.categories),
        series=list(sp.series),
        title=title,
        yAxisLabel=y_axis_label,
    )


def _render_single_subplot_bytes(
    sp: SubplotData,
    title: str,
    y_axis_label: str,
    *,
    view_mode: str = "column",
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    hidden_series: set[str] | None = None,
) -> bytes:
    payload = CompareChartResponse(
        title=title,
        subplots=[sp],
        yAxisLabel=y_axis_label,
    )
    return render_comparison_by_year_bytes(
        payload,
        fmt="png",
        view_mode=view_mode,
        y_axis_min=y_axis_min,
        y_axis_max=y_axis_max,
        clean=clean,
        hidden_series=hidden_series,
    )


def export_compare_xlsx(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    compare_mode: str = "facet",
    years_to_plot: list[int] | None = None,
    sub_filtro: str | None = None,
    loc: str | None = None,
    variable: str | None = None,
    agrupar_por: str | None = None,
    view_mode: str = "column",
    job_display_overrides: dict[int, str] | None = None,
    region: str | None = None,
    combustible: str | None = None,
    hidden_series: set[str] | None = None,
    series_order: list[str] | None = None,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
    clean: bool = False,
    es_porcentaje_override: bool = False,
    facet_placement: str = "inline",
    legend_title: str | None = None,
    exogenous_data: str | None = None,
    exogenous_contaminantes_data: str | None = None,
) -> "io.BytesIO":
    """Genera XLSX multi-hoja para comparación: hoja Comparación + hojas por escenario/año."""
    import io

    from openpyxl import Workbook

    if compare_mode not in ("facet", "by-year", "by-year-alt", "line-total"):
        raise ValueError(
            "compare_mode debe ser 'facet', 'by-year', 'by-year-alt' o 'line-total'"
        )

    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names: set[str] = set()
    render_view_mode = view_mode if view_mode in ("column", "line", "area") else "column"
    year_list = years_to_plot or [2024, 2030, 2050]

    if compare_mode == "facet":
        facet_payload = build_comparison_facet_data(
            db=db,
            job_ids=job_ids,
            tipo=tipo,
            un=un,
            sub_filtro=sub_filtro,
            loc=loc,
            variable=variable,
            agrupar_por=agrupar_por,
            es_porcentaje_override=es_porcentaje_override,
            region=region,
            combustible=combustible,
            job_display_overrides=job_display_overrides,
        )
        if exogenous_data:
            facet_payload = _inject_exogenous_data_into_facets(
                facet_payload, exogenous_data
            )
        if exogenous_contaminantes_data:
            facet_payload = _inject_exogenous_contaminantes_data(
                facet_payload, exogenous_contaminantes_data
            )
        if not facet_payload.facets or not any(f.series for f in facet_payload.facets):
            raise ValueError("Sin datos para exportar con los filtros actuales")

        if series_order:
            for facet in facet_payload.facets:
                reorder_chart_series(facet, series_order)

        layout = "vertical" if facet_placement == "stacked" else "horizontal"
        comparison_img = render_comparison_facet_figure_bytes(
            facet_payload,
            fmt="png",
            layout=layout,
            view_mode=render_view_mode,
            legend_title=legend_title,
            series_order=series_order,
            clean=clean,
            hidden_series=hidden_series,
        )
        comp_headers, comp_rows = _build_facet_combined_table(facet_payload.facets)
        comp_series_names = comp_headers[2:] if len(comp_headers) > 2 else []
        comp_series_colors = _series_colors_for_names(
            comp_series_names,
            [f.series for f in facet_payload.facets],
        )
        ws_comp = wb.create_sheet(
            _safe_excel_sheet_name("Comparación", used_sheet_names)
        )
        _write_comparison_sheet(
            ws_comp,
            image_bytes=comparison_img,
            headers=comp_headers,
            rows=comp_rows,
            y_axis_label=facet_payload.yAxisLabel,
            sheet_title=ws_comp.title,
            view_mode=render_view_mode,
            chart_title=facet_payload.title,
            series_colors=comp_series_colors,
        )

        for facet in facet_payload.facets:
            if not facet.series:
                continue
            scenario_label = _facet_display_name(facet)
            chart = ChartDataResponse(
                categories=list(facet.categories),
                series=list(facet.series),
                title=f"{facet_payload.title} — {scenario_label}",
                yAxisLabel=facet_payload.yAxisLabel,
            )
            if series_order:
                reorder_chart_series(chart, series_order)
            individual_img = render_chart_visualization_bytes(
                chart,
                fmt="png",
                view_mode=render_view_mode,
                y_axis_min=y_axis_min,
                y_axis_max=y_axis_max,
                clean=clean,
                hidden_series=hidden_series,
            )
            sheet_name = _safe_excel_sheet_name(scenario_label, used_sheet_names)
            ws = wb.create_sheet(sheet_name)
            _write_chart_sheet(
                ws,
                image_bytes=individual_img,
                categories=chart.categories,
                series=chart.series,
                y_axis_label=chart.yAxisLabel,
                sheet_title=sheet_name,
                view_mode=render_view_mode,
                chart_title=chart.title,
                category_header="Año",
            )

    elif compare_mode in ("by-year", "by-year-alt"):
        if compare_mode == "by-year-alt":
            cmp_data = build_comparison_data_by_year_alt(
                db=db,
                job_ids=job_ids,
                tipo=tipo,
                un=un,
                years_to_plot=year_list,
                agrupacion=agrupar_por,
                sub_filtro=sub_filtro,
                loc=loc,
                es_porcentaje_override=es_porcentaje_override,
                region=region,
                job_display_overrides=job_display_overrides,
            )
            build_combined = _build_by_year_alt_combined_table
        else:
            cmp_data = build_comparison_data(
                db=db,
                job_ids=job_ids,
                tipo=tipo,
                un=un,
                years_to_plot=year_list,
                agrupacion=agrupar_por,
                sub_filtro=sub_filtro,
                loc=loc,
                es_porcentaje_override=es_porcentaje_override,
                region=region,
                job_display_overrides=job_display_overrides,
            )
            build_combined = _build_by_year_combined_table

        if not cmp_data.subplots or not any(s.series for s in cmp_data.subplots):
            raise ValueError("Sin datos para exportar con los filtros actuales")

        if series_order:
            for sp in cmp_data.subplots:
                reorder_chart_series(sp, series_order)

        comparison_img = render_comparison_by_year_bytes(
            cmp_data,
            fmt="png",
            view_mode=render_view_mode,
            y_axis_min=y_axis_min,
            y_axis_max=y_axis_max,
            clean=clean,
            hidden_series=hidden_series,
        )
        comp_headers, comp_rows = build_combined(cmp_data.subplots)
        comp_series_names = comp_headers[2:] if len(comp_headers) > 2 else []
        comp_series_colors = _series_colors_for_names(
            comp_series_names,
            [sp.series for sp in cmp_data.subplots],
        )
        ws_comp = wb.create_sheet(
            _safe_excel_sheet_name("Comparación", used_sheet_names)
        )
        _write_comparison_sheet(
            ws_comp,
            image_bytes=comparison_img,
            headers=comp_headers,
            rows=comp_rows,
            y_axis_label=cmp_data.yAxisLabel,
            sheet_title=ws_comp.title,
            view_mode=render_view_mode,
            chart_title=cmp_data.title,
            series_colors=comp_series_colors,
        )

        for sp in cmp_data.subplots:
            if not sp.series:
                continue
            panel_label = _subplot_display_name(sp)
            panel_title = f"{cmp_data.title} — {panel_label}"
            if series_order:
                reorder_chart_series(sp, series_order)
            panel_img = _render_single_subplot_bytes(
                sp,
                panel_title,
                cmp_data.yAxisLabel,
                view_mode=render_view_mode,
                y_axis_min=y_axis_min,
                y_axis_max=y_axis_max,
                clean=clean,
                hidden_series=hidden_series,
            )
            sheet_name = _safe_excel_sheet_name(panel_label, used_sheet_names)
            ws = wb.create_sheet(sheet_name)
            _write_chart_sheet(
                ws,
                image_bytes=panel_img,
                categories=list(sp.categories),
                series=list(sp.series),
                y_axis_label=cmp_data.yAxisLabel,
                sheet_title=sheet_name,
                view_mode=render_view_mode,
                chart_title=panel_title,
                category_header="Escenario" if compare_mode == "by-year" else "Año",
            )

    else:  # line-total
        line_data = build_comparison_line_data(
            db=db,
            job_ids=job_ids,
            tipo=tipo,
            un=un,
            sub_filtro=sub_filtro,
            loc=loc,
            job_display_overrides=job_display_overrides,
        )
        if not line_data.series:
            raise ValueError("Sin datos para exportar con los filtros actuales")

        if series_order:
            reorder_chart_series(line_data, series_order)

        comparison_img = render_chart_visualization_bytes(
            line_data,
            fmt="png",
            view_mode="line",
            y_axis_min=y_axis_min,
            y_axis_max=y_axis_max,
            clean=clean,
            hidden_series=hidden_series,
        )
        sheet_name = _safe_excel_sheet_name("Comparación", used_sheet_names)
        ws = wb.create_sheet(sheet_name)
        _write_chart_sheet(
            ws,
            image_bytes=comparison_img,
            categories=line_data.categories,
            series=line_data.series,
            y_axis_label=line_data.yAxisLabel,
            sheet_title=sheet_name,
            view_mode="line",
            chart_title=line_data.title,
            category_header="Año",
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def pareto_data_to_csv_bytes(pareto: ParetoChartResponse) -> bytes:
    """CSV UTF-8 con BOM: categoría, valor, % acumulado."""
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Categoria", pareto.yAxisLabel, "% Acumulado"])
    for i, cat in enumerate(pareto.categories):
        val = pareto.values[i] if i < len(pareto.values) else None
        cum = (
            pareto.cumulative_percent[i] if i < len(pareto.cumulative_percent) else None
        )
        writer.writerow([cat, val, cum])
    return buffer.getvalue().encode("utf-8-sig")


def _safe_filename(name: str) -> str:
    """Genera un nombre de archivo seguro."""
    clean = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in name)
    return clean.strip()[:80]


# ═══════════════════════════════════════════════════════════════════════════
# 8. EXPORT RAW DATA — Excel
# ═══════════════════════════════════════════════════════════════════════════


def export_raw_data_excel(
    db: Session,
    job_id: int,
) -> "io.BytesIO":
    """Exporta todos los datos crudos del job a un archivo Excel (.xlsx)."""
    import io

    rows = (
        db.query(OsemosysOutputParamValue)
        .filter(OsemosysOutputParamValue.id_simulation_job == job_id)
        .all()
    )

    if not rows:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=404,
            detail="No hay datos crudos disponibles para este escenario. La simulación puede no haber guardado resultados en la base de datos.",
        )

    records = []
    for r in rows:
        records.append(
            {
                "VariableName": r.variable_name,
                "Technology": r.technology_name or "",
                "Fuel": r.fuel_name or "",
                "Emission": r.emission_name or "",
                "Year": r.year,
                "Value": float(r.value),
                "IndexJSON": str(r.index_json) if r.index_json else "",
            }
        )

    df = pd.DataFrame(records)

    output = io.BytesIO()
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        worksheet = writer.sheets["Raw Data"]
        # Autofit columns without depending on xlsxwriter.
        for idx, col in enumerate(df):
            series = df[col]
            value_len_max = int(
                series.apply(lambda x: len(str(x)) if pd.notna(x) else 0).max()
            )
            max_len = max(value_len_max, len(str(series.name))) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════════════════════
# 9. EXPORT RESULTS — ZIP de CSVs por variable (formato OSeMOSYS estándar)
# ═══════════════════════════════════════════════════════════════════════════

# Dimensiones por variable. Extiende VARIABLE_INDEX_NAMES con las legacy
# tipadas (Dispatch / UnmetDemand) que se persisten directamente vía
# pipeline._build_output_rows pero no aparecen en el registry.
_LEGACY_TYPED_INDEX_NAMES: dict[str, tuple[str, ...]] = {
    "Dispatch": ("REGION", "TECHNOLOGY", "FUEL", "YEAR"),
    "UnmetDemand": ("REGION", "YEAR"),
}

# Overrides de orden de columnas para el CSV exportado, donde el orden
# OSeMOSYS estándar difiere del usado en VARIABLE_INDEX_NAMES.
# (No modificar VARIABLE_INDEX_NAMES: se usa para interpretar índices al
# persistir resultados.)
_EXPORT_INDEX_OVERRIDES: dict[str, tuple[str, ...]] = {
    "ProductionByTechnology": ("REGION", "TIMESLICE", "TECHNOLOGY", "FUEL", "YEAR"),
    "UseByTechnology": ("REGION", "TIMESLICE", "TECHNOLOGY", "FUEL", "YEAR"),
    "RateOfProductionByTechnology": (
        "REGION",
        "TIMESLICE",
        "TECHNOLOGY",
        "FUEL",
        "YEAR",
    ),
    "RateOfUseByTechnology": ("REGION", "TIMESLICE", "TECHNOLOGY", "FUEL", "YEAR"),
}


def export_results_csv_zip(
    db: Session,
    job_id: int,
) -> "io.BytesIO":
    """Exporta resultados de un job a un ZIP con un CSV por variable.

    Cada CSV usa el formato estándar OSeMOSYS: columnas de dimensión en el
    orden declarado por ``VARIABLE_INDEX_NAMES``, seguidas de ``VALUE``.
    Un archivo por ``variable_name`` presente en BD.
    """
    import io
    import zipfile

    from app.models import (
        Dailytimebracket,
        Daytype,
        Emission,
        Fuel,
        ModeOfOperation,
        Region,
        Season,
        StorageSet,
        Technology,
        Timeslice,
    )
    from app.simulation.core.results_processing import VARIABLE_INDEX_NAMES

    rows = (
        db.query(
            OsemosysOutputParamValue.variable_name,
            Region.name.label("region"),
            Technology.name.label("technology"),
            Fuel.name.label("fuel"),
            Emission.name.label("emission"),
            Timeslice.code.label("timeslice"),
            ModeOfOperation.code.label("mode_of_operation"),
            StorageSet.code.label("storage"),
            Season.code.label("season"),
            Daytype.code.label("daytype"),
            Dailytimebracket.code.label("dailytimebracket"),
            OsemosysOutputParamValue.year,
            OsemosysOutputParamValue.value,
            OsemosysOutputParamValue.technology_name,
            OsemosysOutputParamValue.fuel_name,
            OsemosysOutputParamValue.emission_name,
        )
        .outerjoin(Region, OsemosysOutputParamValue.id_region == Region.id)
        .outerjoin(Technology, OsemosysOutputParamValue.id_technology == Technology.id)
        .outerjoin(Fuel, OsemosysOutputParamValue.id_fuel == Fuel.id)
        .outerjoin(Emission, OsemosysOutputParamValue.id_emission == Emission.id)
        .outerjoin(Timeslice, OsemosysOutputParamValue.id_timeslice == Timeslice.id)
        .outerjoin(
            ModeOfOperation,
            OsemosysOutputParamValue.id_mode_of_operation == ModeOfOperation.id,
        )
        .outerjoin(StorageSet, OsemosysOutputParamValue.id_storage == StorageSet.id)
        .outerjoin(Season, OsemosysOutputParamValue.id_season == Season.id)
        .outerjoin(Daytype, OsemosysOutputParamValue.id_daytype == Daytype.id)
        .outerjoin(
            Dailytimebracket,
            OsemosysOutputParamValue.id_dailytimebracket == Dailytimebracket.id,
        )
        .filter(OsemosysOutputParamValue.id_simulation_job == job_id)
        .all()
    )

    if not rows:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=404,
            detail="No hay resultados para este escenario.",
        )

    _DIM_TO_ROW_ATTR = {
        "REGION": "region",
        "TECHNOLOGY": "technology",
        "FUEL": "fuel",
        "EMISSION": "emission",
        "TIMESLICE": "timeslice",
        "MODE_OF_OPERATION": "mode_of_operation",
        "STORAGE": "storage",
        "SEASON": "season",
        "DAYTYPE": "daytype",
        "DAILYTIMEBRACKET": "dailytimebracket",
    }

    def _index_names_for(var_name: str) -> tuple[str, ...]:
        if var_name in _EXPORT_INDEX_OVERRIDES:
            return _EXPORT_INDEX_OVERRIDES[var_name]
        if var_name in VARIABLE_INDEX_NAMES:
            return VARIABLE_INDEX_NAMES[var_name]
        if var_name in _LEGACY_TYPED_INDEX_NAMES:
            return _LEGACY_TYPED_INDEX_NAMES[var_name]
        return ()

    def _fallback_dims_from_row(r) -> list[str]:
        """Para variables desconocidas, deduce columnas no nulas observadas."""
        dims: list[str] = []
        for dim, attr in _DIM_TO_ROW_ATTR.items():
            if getattr(r, attr, None) not in (None, ""):
                dims.append(dim)
        if r.year is not None:
            dims.append("YEAR")
        return dims

    by_var: dict[str, list] = {}
    for r in rows:
        by_var.setdefault(r.variable_name, []).append(r)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for var_name, var_rows in sorted(by_var.items()):
            dims = list(_index_names_for(var_name))
            if not dims:
                dims = _fallback_dims_from_row(var_rows[0])

            header = list(dims) + ["VALUE"]
            records: list[dict[str, object]] = []
            for r in var_rows:
                rec: dict[str, object] = {}
                for dim in dims:
                    if dim == "YEAR":
                        rec["YEAR"] = r.year if r.year is not None else ""
                        continue
                    attr = _DIM_TO_ROW_ATTR.get(dim)
                    val = getattr(r, attr, None) if attr else None
                    if val is None:
                        if dim == "TECHNOLOGY":
                            val = r.technology_name
                        elif dim == "FUEL":
                            val = r.fuel_name
                        elif dim == "EMISSION":
                            val = r.emission_name
                    rec[dim] = val if val is not None else ""
                rec["VALUE"] = float(r.value)
                records.append(rec)

            df = pd.DataFrame(records, columns=header)
            csv_bytes = df.to_csv(index=False).encode("utf-8")
            zf.writestr(f"{var_name}.csv", csv_bytes)

    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════════
# Helper: comparación recursos por escenario (by-year-alt)
# ═══════════════════════════════════════════════════════════════════════════

def _build_comparison_recursos_by_year_alt(
    db: Session,
    job_ids: list[int],
    tipo: str,
    years_to_plot: list[int] | None,
    un: str = "PJ",
) -> CompareChartResponse:
    """Construye comparación by-year-alt para gráficas de recursos.

    Subplot por escenario, categories = años, serie = producción total.
    """
    if years_to_plot is None:
        years_to_plot = [2024, 2030, 2050]

    title = (
        f"{_get_recursos_title(tipo)} — "
        f"Por Años (Alternativo) ({un})"
    )

    # Cargar nombres de escenarios
    from app.models import Scenario
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"

    subplots: list[SubplotData] = []
    color = "#10b981" if tipo == "recursos_vs_demanda_gas" else "#dc2626"
    prod_label = {
        "recursos_vs_demanda_gas": "Producción gas natural nacional",
        "recursos_vs_demanda_carbon": "Demanda total de carbón",
    }.get(tipo, "Producción")

    for jid in job_ids:
        totals = _build_recursos_production_total(db, jid, tipo, un=un)
        if totals is None:
            continue

        data = [round(totals.get(a, 0.0), 6) for a in years_to_plot]
        subplots.append(
            SubplotData(
                year=jid,
                scenario_name=scenario_names.get(jid, f"Job {jid}"),
                categories=[str(a) for a in years_to_plot],
                series=[
                    ChartSeries(
                        name=prod_label,
                        data=data,
                        color=color,
                        stack="default",
                    )
                ],
            )
        )

    if not subplots:
        return CompareChartResponse(title=title, subplots=[], yAxisLabel=un)

    return CompareChartResponse(
        title=title, subplots=subplots, yAxisLabel=un
    )

    # Cargar nombres de escenarios
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            from app.models import Scenario
            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"

    subplots: list[SubplotData] = []
    color = "#10b981"
    name = "Producción gas natural nacional"

    for jid in job_ids:
        df_gas = _build_recursos_vs_demanda_gas_df(db, jid, un=un)
        if df_gas is None or df_gas.empty:
            continue

        tech_rows = df_gas.groupby("YEAR", as_index=False)["PRODUCTION"].sum()
        prod_by_year = dict(zip(tech_rows["YEAR"], tech_rows["PRODUCTION"]))

        data = [round(prod_by_year.get(a, 0.0), 6) for a in years_to_plot]
        subplots.append(
            SubplotData(
                year=jid,
                scenario_name=scenario_names.get(jid, f"Job {jid}"),
                categories=[str(a) for a in years_to_plot],
                series=[
                    ChartSeries(
                        name=name,
                        data=data,
                        color=color,
                        stack="default",
                    )
                ],
            )
        )

    if not subplots:
        return CompareChartResponse(title=title, subplots=[], yAxisLabel=un)

    return CompareChartResponse(
        title=title, subplots=subplots, yAxisLabel=un
    )


def build_comparison_data_by_year_alt(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    years_to_plot: list[int] | None = None,
    agrupacion: str | None = None,
    sub_filtro: str | None = None,
    loc: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
    es_porcentaje_override: bool = False,
) -> CompareChartResponse:
    """Construye respuesta agrupada por ESCENARIO (no por año).

    Lógica:
    - Cada subplot = un escenario
    - categories = años seleccionados
    - series = una por cada categoría/tecnología (según agrupación)
    """
    # Mapeo de tabla normal a comparación si aplica
    MAPEO_COMPARACION = {
        "tra_total": "tra_comparacion",
        "ind_total": "ind_comparacion",
        "res_total": "res_comparacion",
        "ter_total": "ter_comparacion",
    }

    es_generico = False
    if tipo in MAPEO_COMPARACION:
        tipo = MAPEO_COMPARACION[tipo]

    if tipo not in CONFIGS_COMPARACION and tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no encontrado")

    if years_to_plot is None:
        years_to_plot = [2024, 2030, 2050]

    # ── Ruta especial: recursos vs demanda ────────────────────────────────
    if tipo in _RECURSOS_TIPOS_COMPARACION:
        return _build_comparison_recursos_by_year_alt(
            db, job_ids, tipo, years_to_plot, un,
        )

    # Resolver configuración
    if tipo in CONFIGS_COMPARACION:
        cfg = CONFIGS_COMPARACION[tipo]
        prefijo = cfg["prefijo"]
        agrupacion_fija = cfg.get("agrupacion_fija")
        if agrupacion_fija is not None:
            agrupacion_usar = agrupacion_fija
        elif agrupacion is not None:
            agrupacion_usar = agrupacion
        else:
            agrupacion_usar = cfg["agrupacion_default"]
        variable_name = cfg["variable_default"]
    else:
        es_generico = True
        cfg = CONFIGS[tipo]
        variable_name = cfg["variable_default"]
        agrupacion_usar = (
            agrupacion
            if agrupacion is not None
            else cfg.get("agrupar_por", "TECNOLOGIA")
        )

    # Cargar nombres de escenarios
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            from app.models import Scenario
            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"
        ov = (job_display_overrides or {}).get(jid)
        if isinstance(ov, str) and ov.strip():
            scenario_names[jid] = ov.strip()

    # Procesar datos para cada escenario
    all_data: list[pd.DataFrame] = []
    for jid in job_ids:
        df_var = _load_variable_data(db, jid, variable_name)
        # Strip prefijos regionales si el job es REGIONAL (acumulado nacional).
        df_var = _apply_regional_transform(db, jid, df_var)
        if df_var.empty:
            continue

        if not es_generico:
            df = _procesar_bloque_comparacion(
                df_var, prefijo, sub_filtro, loc,
                agrupacion_usar, years_to_plot, un,
            )
        else:
            df = _procesar_bloque_single(
                df_var, cfg, sub_filtro, loc,
                years_to_plot, un,
                agrupacion_override=agrupacion_usar,
                tipo=tipo,
            )

        if df is None or df.empty:
            continue

        df["ESCENARIO"] = scenario_names.get(jid, f"Job {jid}")
        df["JOB_ID"] = jid
        all_data.append(df)

    if not all_data:
        title_base = cfg.get("titulo_base", cfg.get("titulo", tipo))
        return CompareChartResponse(
            title=f"{title_base} (Comparación Alternativa)",
            subplots=[],
            yAxisLabel=un,
        )

    df_final = pd.concat(all_data, ignore_index=True)

    # Aplicar porcentaje si corresponde
    if es_porcentaje_override:
        total_por_escenario_año = df_final.groupby(["JOB_ID", "YEAR"])["VALUE"].transform("sum")
        df_final["VALUE"] = df_final["VALUE"] / total_por_escenario_año * 100.0

    # ── Colores ──────────────────────────────────────────────────
    categorias_unicas = sorted(df_final["CATEGORIA"].dropna().unique())
    if not es_generico:
        mapa_colores = _color_map_comparison(agrupacion_usar, categorias_unicas)
    else:
        # Para gráficas genéricas: usar color_fn según la agrupación REAL
        if agrupacion_usar != cfg.get("agrupar_por"):
            if agrupacion_usar == "FUEL":
                color_fn = _color_por_grupo_fijo
            elif agrupacion_usar == "SECTOR":
                color_fn = _color_por_sector
            elif agrupacion_usar == "EMISION":
                color_fn = _color_por_emision
            elif agrupacion_usar == "ELECTROLISIS":
                color_fn = _color_electrolisis
            else:
                color_fn = cfg.get("color_fn") or generar_colores_tecnologias
        else:
            color_fn = cfg.get("color_fn")

        if color_fn is not None:
            df_tmp = pd.DataFrame({"COLOR": list(categorias_unicas)})
            colores_lista, orden_lista = color_fn(df_tmp, "COLOR")
            mapa_colores = dict(zip(orden_lista, colores_lista))
        else:
            _palette = get_colores_grupos()
            mapa_colores = {c: _palette.get(c, "#999999") for c in categorias_unicas}

    from app.services.chart_series_config_service import (
        apply_global_series_config,
        normalize_agrupar_por,
    )
    agrup_key = normalize_agrupar_por(agrupacion_usar, agrupacion_usar)
    ordered_stack = apply_global_series_config(
        db,
        tipo=tipo,
        agrupar_por=agrup_key,
        orden_color=list(categorias_unicas),
        color_dict=mapa_colores,
        default_name=lambda c: get_label(str(c)),
    )

    # Construir subplots por escenario
    subplots: list[SubplotData] = []
    for jid in job_ids:
        df_escenario = df_final[df_final["JOB_ID"] == jid]
        if df_escenario.empty:
            continue

        escenario_nombre = scenario_names.get(jid, f"Job {jid}")

        series: list[ChartSeries] = []
        for categoria, col_cat, name_cat in ordered_stack:
            df_cat = df_escenario[df_escenario["CATEGORIA"] == categoria]
            if df_cat.empty:
                series.append(
                    ChartSeries(
                        name=name_cat,
                        data=[0.0] * len(years_to_plot),
                        color=col_cat,
                        stack="default",
                    )
                )
                continue

            valor_por_año = {
                int(row["YEAR"]): row["VALUE"]
                for _, row in df_cat.iterrows()
            }
            data = [round(valor_por_año.get(a, 0.0), 6) for a in years_to_plot]

            series.append(
                ChartSeries(
                    name=name_cat,
                    data=data,
                    color=col_cat,
                    stack="default",
                )
            )

        subplots.append(
            SubplotData(
                year=jid,
                scenario_name=escenario_nombre,
                categories=[str(a) for a in years_to_plot],
                series=series,
            )
        )

    title_base = cfg.get("titulo_base", cfg.get("titulo", tipo))
    title = f"{title_base} — Por Años (Alternativo)"
    if sub_filtro:
        title += f" — {NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)}"
    if loc:
        title += f" ({loc})"
    title += f" ({un})"

    return CompareChartResponse(
        title=title,
        subplots=subplots,
        yAxisLabel="%" if es_porcentaje_override else un,
    )
