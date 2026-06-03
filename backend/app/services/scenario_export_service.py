"""
Exporta un escenario a Excel en formato SAND (hoja Parameters).

Estructura idéntica a la plantilla SAND oficial: Parameter, dimensiones
(REGION, TECHNOLOGY, EMISSION, MODE_OF_OPERATION, FUEL, TIMESLICE, STORAGE,
REGION2), Time indipendent variables, y columnas por año (2022–2055 fijo).
Permite descargar, editar en Excel y volver a subir manteniendo el contrato
con el importador.

Implementación: usa openpyxl en modo `write_only` y escribe a un path en disco
para mantener bajo el uso de memoria y reducir el tiempo de generación —
necesario para escenarios grandes que de otro modo timeoutean en el proxy.
"""

from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.simulation.core.data_processing import PARAM_INDEX, _resolved_query
from app.simulation.core.mode_of_operation_normalize import normalize_mode_of_operation_scalar

# Cabeceras SAND canónicas — orden y casing tomados de SAND/SAND_*_REF.xlsm
# (verificado en SAND_15_12_2025_REF, SAND_18_11_2025_REF, SAND_26_08_2025, SAND_01_04_2025).
SAND_DIMENSION_HEADERS = [
    "Parameter",
    "REGION",
    "TECHNOLOGY",
    "EMISSION",
    "MODE_OF_OPERATION",
    "FUEL",
    "TIMESLICE",
    "STORAGE",
    "REGION2",
]
TIME_INDEPENDENT_HEADER = "Time indipendent variables"
SAND_YEARS: list[int] = list(range(2022, 2056))

RAW_HEADERS = [
    "Parameter",
    "REGION",
    "TECHNOLOGY",
    "EMISSION",
    "MODE_OF_OPERATION",
    "FUEL",
    "TIMESLICE",
    "STORAGE",
    "REGION2",
    "YEAR",
    "VALUE",
]

# Dimensiones que SAND no representa: los parámetros que las usen se omiten
# del export SAND para no introducir cabeceras espurias.
_SAND_FORBIDDEN_DIMS = {"SEASON", "DAYTYPE", "DAILYTIMEBRACKET", "UDC"}

# Índices del row devuelto por _resolved_query():
# 0:param 1:region 2:tech 3:fuel 4:emission 5:timeslice 6:mode 7:season 8:daytype 9:dtb 10:storage 11:udc 12:year 13:value
_ROW_PARAM = 0
_ROW_REGION = 1
_ROW_TECH = 2
_ROW_FUEL = 3
_ROW_EMISSION = 4
_ROW_TIMESLICE = 5
_ROW_MODE = 6
_ROW_STORAGE = 10
_ROW_YEAR = 12
_ROW_VALUE = 13


def _row_to_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _param_is_sand_compatible(pname: str) -> bool:
    dims = PARAM_INDEX.get(pname)
    if dims is None:
        return False
    return not (set(dims) & _SAND_FORBIDDEN_DIMS)


def _sand_key(row) -> tuple:
    """Construye la key SAND (orden de columnas de la plantilla)."""
    return (
        row[_ROW_PARAM],
        _row_to_str(row[_ROW_REGION]),
        _row_to_str(row[_ROW_TECH]),
        _row_to_str(row[_ROW_EMISSION]),
        normalize_mode_of_operation_scalar(row[_ROW_MODE]),
        _row_to_str(row[_ROW_FUEL]),
        _row_to_str(row[_ROW_TIMESLICE]),
        _row_to_str(row[_ROW_STORAGE]),
        "",  # REGION2 — sin columna en el esquema BD actual
    )


def export_scenario_to_excel_file(
    db: Session, *, scenario_id: int, scenario_name: str, output_path: str
) -> None:
    """Escribe el Excel SAND directamente a `output_path` (modo write_only)."""
    result_proxy = db.execute(_resolved_query(), {"scenario_id": scenario_id})

    grouped: dict[tuple, dict[int | None, float]] = defaultdict(dict)
    for row in result_proxy.yield_per(50_000):
        pname = row[_ROW_PARAM]
        if not _param_is_sand_compatible(pname):
            continue
        value = float(row[_ROW_VALUE])
        year_raw = row[_ROW_YEAR]
        year_val: int | None = int(year_raw) if year_raw is not None else None
        grouped[_sand_key(row)][year_val] = value

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Parameters")

    headers = list(SAND_DIMENSION_HEADERS) + [TIME_INDEPENDENT_HEADER] + [str(y) for y in SAND_YEARS]
    ws.append(headers)

    n_year_cols = len(SAND_YEARS)
    for key, year_to_val in sorted(grouped.items(), key=lambda x: x[0]):
        row_out: list = [v or None for v in key]
        row_out.append(year_to_val.get(None))
        row_out.extend(year_to_val.get(yr) for yr in SAND_YEARS)
        # Sanity: longitud total = dims + time-indep + years
        assert len(row_out) == len(SAND_DIMENSION_HEADERS) + 1 + n_year_cols
        ws.append(row_out)

    wb.save(output_path)


def export_scenario_raw_to_excel_file(
    db: Session, *, scenario_id: int, scenario_name: str, output_path: str
) -> None:
    """Escribe el Excel RAW (1 fila por registro) directamente a `output_path`."""
    result_proxy = db.execute(_resolved_query(), {"scenario_id": scenario_id})

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="RawParameters")
    ws.append(RAW_HEADERS)

    for row in result_proxy.yield_per(50_000):
        ws.append(
            [
                _row_to_str(row[_ROW_PARAM]) or None,
                _row_to_str(row[_ROW_REGION]) or None,
                _row_to_str(row[_ROW_TECH]) or None,
                _row_to_str(row[_ROW_EMISSION]) or None,
                normalize_mode_of_operation_scalar(row[_ROW_MODE]) or None,
                _row_to_str(row[_ROW_FUEL]) or None,
                _row_to_str(row[_ROW_TIMESLICE]) or None,
                _row_to_str(row[_ROW_STORAGE]) or None,
                None,  # REGION2
                int(row[_ROW_YEAR]) if row[_ROW_YEAR] is not None else None,
                float(row[_ROW_VALUE]) if row[_ROW_VALUE] is not None else None,
            ]
        )

    wb.save(output_path)
